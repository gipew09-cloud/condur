"""
Полноценные тесты по доработкам: акт (формат/итог/прописью/название),
JWT-авторизация (вход владельца/админа + мгновенный отзыв), и рендер всех
изменённых страниц кабинета через настоящее Jinja-окружение с реальными фильтрами.

Запуск: `pip install pytest openpyxl jinja2` затем `pytest` в корне проекта.
ENV ботов/БД фиктивные — нужны только для импорта настроек; до сети/БД не доходим.
"""
import os
import io
from datetime import date, datetime, timezone
from decimal import Decimal
from types import SimpleNamespace as NS

os.environ.setdefault("OWNER_BOT_TOKEN", "test")
os.environ.setdefault("DRIVER_BOT_TOKEN", "test")
os.environ.setdefault("DATABASE_URL", "postgresql+asyncpg://u:p@localhost/db")
os.environ.setdefault("JWT_SECRET", "test-secret-please-set-a-long-one-in-prod")

import pytest  # noqa: E402
from jinja2 import Environment, FileSystemLoader  # noqa: E402
from openpyxl import Workbook  # noqa: E402

from app.services import act_service as A  # noqa: E402
from app.services import auth_service as AU  # noqa: E402
from app.services import rc_service as RC  # noqa: E402
from app.services.timeutil import fmt_dt  # noqa: E402

# Jinja-окружение с теми же фильтрами, что в app/web/router.py (чтобы не тянуть
# в тест весь FastAPI/SQLAlchemy-стек ради рендера шаблонов).
_VT = {"truck": "Грузовик", "gazelle": "Газель / фургон", "refrigerator": "Рефрижератор"}
_TS = {"created": "создан", "in_transit": "в пути", "unloading": "на выгрузке",
       "completed": "завершён", "cancelled": "отменён"}
_PILL = {"completed": "pill--success", "paid": "pill--success", "pending": "pill--warn",
         "approved": "pill--success", "rejected": "pill--danger", "in_transit": "pill--info"}
_ENV = Environment(loader=FileSystemLoader("app/web/templates"))
_ENV.filters["vtype"] = lambda c: _VT.get(c or "", c or "—")
_ENV.filters["tstatus"] = lambda c: _TS.get(c or "", c or "—")
_ENV.filters["pillclass"] = lambda s: _PILL.get((s or "").lower(), "pill--neutral")
_ENV.filters["statusru"] = lambda s: s or "—"
_ENV.filters["localdt"] = lambda dt, tz=None, fmt="%d.%m.%Y %H:%M": fmt_dt(dt, tz, fmt)


# ---------------------------------------------------------------- сумма прописью
@pytest.mark.parametrize("amount,expected", [
    (Decimal("292000"), "Двести девяносто две тысячи рублей 00 копеек."),
    (Decimal("178000"), "Сто семьдесят восемь тысяч рублей 00 копеек."),
    (Decimal("0"), "Ноль рублей 00 копеек."),
    (Decimal("1"), "Один рубль 00 копеек."),
    (Decimal("1234.56"), "Одна тысяча двести тридцать четыре рубля 56 копеек."),
    (Decimal("21"), "Двадцать один рубль 00 копеек."),
])
def test_rubles_in_words(amount, expected):
    assert A.rubles_in_words(amount) == expected


def test_money_str_ru_format():
    # разделитель тысяч — неразрывный пробел (\xa0), нормализуем для сравнения
    assert A._money_str(Decimal("292000")).replace("\xa0", " ") == "292 000,00"
    assert A._money_str(Decimal("1234.5")).replace("\xa0", " ") == "1 234,50"


# ------------------------------------------------------------------ генератор акта
def _sample_rows(n=3):
    return [
        {"date": date(2026, 5, d), "origin": "Агропарк Софийская 151",
         "destination": f"РЦ {d}", "plate": "Т 557 ОС 178", "driver": "Саломов",
         "amount": Decimal("19000")}
        for d in range(1, n + 1)
    ]


def _executor():
    return {"full_name": "ИП Кибиткина", "inn": "781699567368", "ogrnip": "319",
            "address": "СПб", "bank_name": "УРАЛСИБ", "account": "40802",
            "corr_account": "30101", "bik": "044030706", "signer_name": "Кибиткина"}


def _customer():
    return {"name": 'ООО "Рузисеть"', "inn": "7802533479", "kpp": "781701001",
            "address": "СПб", "bank_name": "ПСКБ", "account": "40702",
            "corr_account": "30101", "bik": "044030852", "contract_number": "№521",
            "contract_date": date(2020, 6, 29), "signer_name": "Аллахвердиев"}


def _build(title="Акт", rows=None):
    return A.build_act_101rs(
        title=title, act_number="102", act_date=date(2026, 6, 25),
        period_from=date(2026, 6, 1), period_to=date(2026, 6, 25),
        executor=_executor(), customer=_customer(),
        rows=_sample_rows(3) if rows is None else rows,
    )


def _find_header_row(ws):
    for r in range(1, 12):
        if ws.cell(r, 1).value == "№":
            return r
    raise AssertionError("шапка таблицы не найдена")


def test_act_title_default_and_custom():
    assert _build().active["A1"].value == "Акт № 102 от 25.06.2026 г."
    assert _build(title="Акт сверки").active["A1"].value == "Акт сверки № 102 от 25.06.2026 г."


def test_act_header_has_no_fill_and_9pt():
    ws = _build().active
    h = ws.cell(_find_header_row(ws), 1)
    assert h.fill.patternType is None, "шапка не должна быть с заливкой (как в образце 101)"
    assert h.font.bold is True and h.font.size == 9


def test_act_money_format_and_totals():
    rows = _sample_rows(3)  # 3 × 19000 = 57000
    ws = _build(rows=rows).active
    hr = _find_header_row(ws)
    price_cell = ws.cell(hr + 1, 5)
    assert price_cell.number_format == "#,##0.00" and "₽" not in price_cell.number_format
    # ИТОГО и «Всего оказано услуг N»
    joined = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    ).replace("\xa0", " ")  # неразрывный пробел → обычный для сравнения
    assert "Всего оказано услуг 3." in joined
    assert "57 000,00 руб." in joined
    assert "Пятьдесят семь тысяч рублей 00 копеек." in joined
    assert "Без налога (НДС)" in joined


def test_act_empty_rows_ok():
    ws = _build(rows=[]).active
    joined = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "Всего оказано услуг 0." in joined


def test_act_uses_distribution_center_address_when_present():
    rows = [{
        "date": date(2026, 6, 1),
        "origin": "Агропарк Софийская 151",
        "destination": "РЦ Шушары",
        "destination_address": "Санкт-Петербург, Пушкинский район, пос. Шушары, Московское шоссе, 231",
        "plate": "Т 557 ОС 178",
        "driver": "Саломов",
        "amount": Decimal("19000"),
    }]
    joined = "\n".join(
        str(c.value) for row in _build(rows=rows).active.iter_rows() for c in row if c.value
    )
    assert "Московское шоссе, 231" in joined
    assert "РЦ Шушары" not in joined


# ------------------------------------------------------------------ JWT / доступ
def test_jwt_roundtrip_owner_and_admin():
    assert AU.decode_jwt(AU.create_jwt(5, tid=123)) == (5, 123)   # админ вошёл в кабинет 5
    assert AU.decode_jwt(AU.create_jwt(5, tid=999)) == (5, 999)   # владелец
    assert AU.decode_jwt(AU.create_jwt(5)) == (5, None)           # старый токен (совместимость)


def test_jwt_invalid_returns_none():
    assert AU.decode_jwt("not-a-token") is None
    assert AU.decode_jwt("") is None


def test_login_code_is_one_time_and_checked():
    code = AU.issue_code(777)
    assert AU.consume_code(777, "000000") is False   # неверный код
    assert AU.consume_code(777, code) is True         # верный
    assert AU.consume_code(777, code) is False        # повторно — уже израсходован


def _admin_access_ok(owner_tid, jwt_tid, admin_exists):
    """Копия логики current_owner: админ-доступ действует, только пока есть Admin."""
    if jwt_tid is not None and jwt_tid != owner_tid:
        return admin_exists
    return True


def test_admin_revocation_is_immediate():
    assert _admin_access_ok(999, 123, admin_exists=True) is True    # админ активен
    assert _admin_access_ok(999, 123, admin_exists=False) is False  # удалён → доступ закрыт сразу
    assert _admin_access_ok(999, 999, admin_exists=False) is True   # владелец
    assert _admin_access_ok(999, None, admin_exists=False) is True  # старый токен = владелец


# ------------------------------------------------------------------ рендер страниц
def _render(name, **ctx):
    return _ENV.get_template(name).render(**ctx)


OWNER = NS(company_name="БРО ЛОГИСТИК", executor_name="ИП", inn="7816",
           ogrnip="319", legal_address="СПб", bank_name="УРАЛСИБ",
           bank_account="40802", corr_account="30101", bik="044030706",
           signer_name="Иванов", full_name="Иванов", telegram_id=1)


def test_render_login():
    assert "Вход в кабинет" in _render("login.html", error=None)
    assert "Свежий путь" not in _render("login.html", error=None)


def test_mobile_navigation_has_requisites_logout_and_scroll():
    html = _render("base.html", owner=OWNER, active_page="map")
    assert '<nav class="bottom-nav">' in html
    assert 'href="/stats"' in html
    assert 'href="/requisites"' in html
    assert 'href="/logout"' in html

    mobile_css = open("app/web/static/mobile.css", encoding="utf-8").read()
    cabinet_css = open("app/web/static/cabinet.css", encoding="utf-8").read()
    assert "overflow-x: auto" in mobile_css
    assert "flex: 0 0 76px" in mobile_css
    assert "grid-template-columns: auto minmax(0, 1fr)" in cabinet_css


def test_login_clears_legacy_auth_cookie_source():
    source = open("app/web/router.py", encoding="utf-8").read()
    # вход ставит постоянную cookie сессии (через хелпер с Secure/Expires для iOS)
    assert "auth_service.set_session_cookie(response, raw_token)" in source
    # и чистит старый 7-дневный JWT
    assert 'response.delete_cookie("auth")' in source


def test_session_cookie_is_secure_and_persistent():
    """Cookie сессии должна быть Secure + с Expires + Path=/ — иначе Safari/iOS
    сбрасывает её и вход постоянно просит код заново."""
    source = open("app/services/auth_service.py", encoding="utf-8").read()
    assert "secure=settings.cookie_secure" in source
    assert "expires=expires" in source
    assert 'path="/"' in source
    assert 'samesite="lax"' in source


def test_render_finances():
    html = _render(
        "finances.html", owner=OWNER, active_page="finances",
        summary={"total_income": Decimal("258000"), "total_expense": Decimal("267555"),
                 "profit": Decimal("-9555"), "fuel": Decimal("0")},
        margin=-4.0,
        cashflow={"labels": ["май", "июн"], "revenue": [0, 258000],
                  "expenses": [0, 267555], "profit": [0, -9555], "period": "6m"},
        directions=[{"route": "А → Б", "trips": 5, "revenue": Decimal("106000"),
                     "profit": Decimal("106000"), "bar": 100}],
        entries=[], period_from="2026-06-01", period_to="2026-06-25",
        today="2026-06-25")
    assert "Денежный поток" in html and "Прибыльность направлений" in html
    assert ".sp-kpi" in html  # стили встроены в страницу


def test_render_vehicles_card_layout():
    v = NS(id=1, license_plate="ACHICHJ23627", brand="тимьмиь", type="refrigerator",
           osago_expires=date(2026, 9, 1), inspection_expires=None, tacho_expires=None,
           fuel_norm_per_100km=Decimal("32"))
    row = {"vehicle": v, "km": 0, "fuel": Decimal(0), "trips": 0, "revenue": Decimal(0),
           "expense": Decimal(0), "profit": Decimal(0), "margin": Decimal(0), "active": False}
    html = _render("vehicles.html", owner=OWNER, rows=[row], active_page="vehicles",
                   in_minus=0, notice=None,
                   totals={"count": 1, "in_work": 0, "revenue": Decimal(0), "profit": Decimal(0)})
    assert '<div class="grow">' not in html   # номер+статус в ряд, без grow-обёртки
    assert "свободна" in html


def test_render_drivers_all_time():
    d = NS(id=1, full_name="Иванов Иван", phone="+7", salary_type="fixed_per_month",
           salary_rate=Decimal("180000"), per_diem_rub=Decimal("0"),
           shift_start_time=None, telegram_id=1)
    row = {"driver": d, "km": 1477, "shifts": 12, "trips": 12, "revenue": Decimal("258000"),
           "fuel_cost": Decimal("0"), "active_shift": False, "idle_label": "24.06 13:16"}
    html = _render("drivers.html", owner=OWNER, rows=[row], active_page="drivers",
                   invite=None, totals={"count": 1, "in_shift": 0, "idle": 1})
    assert "за всё время" in html and ">12<" in html


def test_render_acts_checklist():
    trips = [{"id": 11, "date": date(2026, 6, 21), "origin": "А", "destination": "Б",
              "driver": "Саломов", "plate": "Т557", "revenue": Decimal("19000")}]
    html = _render("acts.html", owner=OWNER, trips=trips, period_from="2026-06-01",
                   period_to="2026-06-25", customers=[NS(id=2, name='ООО "Рузисеть"')],
                   total_amount=Decimal("19000"), total_trips=1, act_date="2026-06-25",
                   act_title="Акт выполненных работ", act_number_val="102", sel_customer_id="2",
                   requisites_ready=True, active_page="finances")
    assert "Акты выполненных работ" in html and 'name="trip_ids"' in html and "Выбрать все" in html
    assert 'name="selection_mode" value="checklist"' in html


def test_render_requisites_with_admins():
    admins = [NS(id=1, name="Бухгалтер Анна", telegram_id=123456789, created_at=None)]
    html = _render("requisites.html", owner=OWNER, customers=[], admins=admins,
                   sessions=[], is_owner_viewer=True, active_page="requisites")
    assert "Администраторы" in html and "Telegram ID 123456789" in html
    assert "/admins/add" in html
    assert "Устройства" in html  # блок сессий есть даже когда список пуст


def _session_row(id, who, current=False, can_revoke=True):
    return {"id": id, "who": who, "device": "Chrome · Windows", "ip": "1.2.3.4",
            "created": "04.07 10:00", "seen": "04.07 12:00",
            "is_current": current, "can_revoke": can_revoke}


def test_render_requisites_devices_owner_view():
    html = _render("requisites.html", owner=OWNER, customers=[], admins=[],
                   sessions=[_session_row(1, "Владелец", current=True),
                             _session_row(2, "Бухгалтер Анна")],
                   is_owner_viewer=True, active_page="requisites")
    assert "это устройство" in html and "Выйти здесь" in html
    assert "/sessions/2/revoke" in html and "Завершить" in html
    assert "/sessions/revoke-others" in html  # владелец видит «выйти на всех»


def test_render_requisites_devices_admin_view_full_access():
    # Админ — «второй телефон» владельца: видит кнопки завершения у ВСЕХ
    # устройств кабинета и массовый выход (политика изменена 2026-07-13,
    # раньше чужие сессии были только у владельца — кнопки «Убрать» не хватало).
    html = _render("requisites.html", owner=OWNER, customers=[], admins=[],
                   sessions=[_session_row(1, "Владелец"),
                             _session_row(2, "Бухгалтер Анна", current=True)],
                   is_owner_viewer=False, active_page="requisites")
    assert "/sessions/revoke-others" in html
    assert "/sessions/1/revoke" in html
    assert "/sessions/2/revoke" in html


def test_session_token_hash_and_device_label():
    t1, t2 = AU.new_session_token(), AU.new_session_token()
    assert t1 != t2 and len(t1) > 40
    assert AU.session_token_hash(t1) != AU.session_token_hash(t2)
    assert len(AU.session_token_hash(t1)) == 64  # sha256 hex
    assert AU.session_token_hash(t1) == AU.session_token_hash(t1)  # детерминирован

    ua_mac = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/126 Safari/537.36"
    assert AU.device_label_from_user_agent(ua_mac) == "Chrome · macOS"
    ua_iphone = "Mozilla/5.0 (iPhone; CPU iPhone OS 17_5 like Mac OS X) AppleWebKit/605 Version/17.5 Safari/604.1"
    assert AU.device_label_from_user_agent(ua_iphone) == "Safari · iPhone"
    assert AU.device_label_from_user_agent(None) == "Браузер · ?"


def test_render_expenses_donut():
    exp = NS(id=1, created_at=None, category="fuel", amount_rub=Decimal("5000"),
             status="approved", receipt_web_data=None, receipt_photo_url=None)
    html = _render("expenses.html", owner=OWNER, active_page="trips",
                   rows=[(exp, "Саломов", "Т557")],
                   totals={"count": 1, "sum": Decimal("5000"), "pending": 0, "approved": 1},
                   filter_category="", filter_status="", categories=("fuel",),
                   drivers=[NS(id=1, full_name="Саломов")],
                   vehicles=[NS(id=2, license_plate="Т557ОС178")],
                   filter_driver_id=1, filter_vehicle_id=None,
                   filter_date_from="2026-07-01", filter_date_to="2026-07-04",
                   breakdown=[{"label": "Топливо", "amount": 5000.0}])
    assert "Структура расходов по категориям" in html
    # фильтры: водитель выбран, машина в списке, даты подставлены, кнопки на месте
    assert 'name="driver_id"' in html and "selected>Саломов" in html
    assert 'name="vehicle_id"' in html and "Т557ОС178" in html
    assert 'value="2026-07-01"' in html and "Применить" in html and "Сбросить" in html


def test_render_routes_with_distribution_centers():
    center = NS(id=1, name="РЦ Шушары", address="СПб, Московское шоссе, 231",
                aliases="Шушары; РЦ СПб", latitude=Decimal("59.75"),
                longitude=Decimal("30.45"))
    html = _render(
        "routes.html", owner=OWNER, active_page="routes",
        rows=[], centers=[center], rc_imported=1,
    )
    assert "Справочник РЦ" in html
    assert "РЦ Шушары" in html and "СПб, Московское шоссе, 231" in html
    assert "/routes/rc/import" in html and "/routes/rc/1/delete" in html


def test_rc_xlsx_import_and_lookup_preserves_zero_coordinates():
    wb = Workbook()
    ws = wb.active
    ws.append(["РЦ", "Адрес", "Алиасы", "Широта", "Долгота"])
    ws.append(["РЦ Шушары", "СПб, Московское шоссе, 231", "Шушары; РЦ СПб", 0, "30,4501"])
    buf = io.BytesIO()
    wb.save(buf)

    rows = RC.distribution_centers_from_xlsx(buf.getvalue())
    assert rows == [{
        "name": "РЦ Шушары",
        "address": "СПб, Московское шоссе, 231",
        "aliases": "Шушары; РЦ СПб",
        "latitude": "0",
        "longitude": "30,4501",
    }]
    assert RC.decimal_or_none(rows[0]["latitude"]) == Decimal("0")
    assert RC.decimal_or_none(rows[0]["longitude"]) == Decimal("30.4501")

    lookup = RC.distribution_center_lookup([
        NS(name=rows[0]["name"], address=rows[0]["address"], aliases=rows[0]["aliases"])
    ])
    assert RC.canonical_rc_address("доставка в шушары", lookup) == "СПб, Московское шоссе, 231"


def test_rc_xlsx_import_single_column_file():
    """Файл владельца «РЦ Адреса Спб.xlsx»: одна колонка, без шапки,
    название и адрес слиты в одной ячейке. Должен импортироваться."""
    wb = Workbook()
    ws = wb.active
    ws.append(["РЦ Лента Лен.Обл. Тосненский район посёлок Красный бор"])
    ws.append(["РЦ 7 шагов, Санкт-Петербург г., п. Шушары"])
    ws.append([None])  # пустая строка — пропускается
    buf = io.BytesIO()
    wb.save(buf)

    rows = RC.distribution_centers_from_xlsx(buf.getvalue())
    assert len(rows) == 2
    assert rows[0]["name"] == rows[0]["address"]  # адрес = вся ячейка
    assert rows[0]["name"].startswith("РЦ Лента")


def test_render_trips_kpi():
    html = _render("trips.html", owner=OWNER, active_page="trips",
                   drivers=[NS(id=1, full_name="Д")],
                   vehicles=[NS(id=1, license_plate="П", brand="")],
                   filter_driver_id=None, filter_date_from="", filter_date_to="",
                   today="2026-06-25",
                   totals={"count": 12, "revenue": Decimal("292000"), "profit": Decimal("112000")},
                   rows=[], page=1, has_next=False)
    assert "sp-kpi" in html and "Рейсов по фильтру" in html


def test_pending_driver_revenue_is_not_counted_as_final_revenue():
    trip = NS(
        id=39,
        created_at=datetime(2026, 7, 3, 18, 55, tzinfo=timezone.utc),
        origin="непонятно",
        destination="непонятно",
        is_manual=False,
        revenue_rub=None,
        driver_revenue_pending_rub=Decimal("12000"),
        fuel_cost_rub=Decimal("0"),
        profit_rub=Decimal("0"),
        status="completed",
    )
    html = _render(
        "_trips_table.html",
        owner=OWNER,
        rows=[(trip, "овыраовраор", "Е774ЕТ178")],
        totals={"count": 1, "revenue": Decimal("0"), "profit": Decimal("0")},
        filter_driver_id="",
        filter_date_from="",
        filter_date_to="",
        page=1,
        has_next=False,
    )

    assert "на подтверждении 12 000 ₽" in html
    assert "<div class=\"lbl\">Выручка</div><div class=\"val\">0 <span class=\"rub\">₽</span></div>" in html


def test_trip_detail_shows_pending_driver_revenue_separately():
    trip = NS(
        id=39,
        created_at=datetime(2026, 7, 3, 18, 55, tzinfo=timezone.utc),
        completed_at=datetime(2026, 7, 3, 18, 56, tzinfo=timezone.utc),
        origin="непонятно",
        destination="непонятно",
        is_manual=False,
        status="completed",
        cargo_name="РЦ адамнт",
        revenue_rub=None,
        driver_revenue_pending_rub=Decimal("12000"),
        fuel_cost_rub=Decimal("0"),
        other_costs_rub=Decimal("0"),
        profit_rub=Decimal("0"),
        waybill_photo_url=None,
    )
    html = _render(
        "trip_detail.html",
        owner=OWNER,
        active_page="trips",
        trip=trip,
        driver=NS(full_name="овыраовраор"),
        vehicle=NS(license_plate="Е774ЕТ178"),
        travel=None,
        documents=[],
        expenses=[],
        waybill_uploaded_at=None,
    )

    assert "на подтверждении 12 000 ₽" in html
    assert "<strong>0 ₽</strong>" in html
    assert "Карточка рейса · таймлайн" in html
    assert "Удалить рейс" in html


def test_trip_and_shift_delete_routes_are_safe_detach_handlers():
    src = open("app/web/router.py", encoding="utf-8").read()
    assert '@app.post("/trips/{trip_id}/delete")' in src
    assert '@app.post("/shifts/{shift_id}/delete")' in src
    assert "Expense.trip_id == trip.id" in src
    assert "Event.trip_id == trip.id" in src
    assert "TripDocument.trip_id == trip.id" in src
    assert "Expense.shift_id == shift.id" in src
    assert "Event.shift_id == shift.id" in src


# ------------------------------------------------------------------ карта (Яндекс)
def test_map_template_uses_yandex_not_leaflet():
    src = open("app/web/templates/map.html", encoding="utf-8").read()
    for banned in ("leaflet", "Leaflet", "openstreetmap", "cartocdn", "CARTO", "L.map"):
        assert banned not in src, f"старая карта не удалена: {banned}"
    assert "api-maps.yandex.ru" in src

    html = _render("map.html", owner=OWNER, active_page="map",
                   yandex_maps_api_key="test-key-123")
    assert "api-maps.yandex.ru/2.1/?apikey=test-key-123" in html
    assert "vehicle-marker" in html and "visibilitychange" in html
    assert "removeMissing(vehicleMarkers" in html and "removeMissing(driverMarkers" in html


def test_map_template_without_key_shows_hint():
    html = _render("map.html", owner=OWNER, active_page="map", yandex_maps_api_key="")
    assert "YANDEX_MAPS_API_KEY" in html
    assert "api-maps.yandex.ru" not in html


# ------------------------------------------------------------------ метка «с какого времени»
def test_smart_since_label_today_yesterday_older():
    from datetime import datetime, timedelta, timezone as tz

    from app.services.timeutil import now_in_tz, smart_since_label

    now_local = now_in_tz("Europe/Moscow")
    today_dt = now_local.replace(hour=10, minute=5) if now_local.hour >= 11 \
        else now_local.replace(minute=0)
    assert smart_since_label(today_dt, "Europe/Moscow").startswith("с ")
    assert "," not in smart_since_label(today_dt, "Europe/Moscow")

    yesterday = now_local - timedelta(days=1)
    assert smart_since_label(yesterday, "Europe/Moscow").startswith("со вчера, ")

    older = now_local - timedelta(days=5)
    label = smart_since_label(older, "Europe/Moscow")
    assert label.startswith("с ") and older.strftime("%d.%m") in label
    assert smart_since_label(None, "Europe/Moscow") == "—"


# ------------------------------------------------------------------ график: шаг периода
def test_cashflow_buckets_days_weeks_months():
    from app.services.timeutil import cashflow_buckets

    # ≤45 дней — по дням
    labels, keys, key_of = cashflow_buckets(date(2026, 7, 1), date(2026, 7, 7))
    assert len(keys) == 7 and labels[0] == "01.07" and labels[-1] == "07.07"
    assert key_of(date(2026, 7, 3)) == date(2026, 7, 3)

    # 46–200 дней — по неделям, ключ = понедельник
    labels, keys, key_of = cashflow_buckets(date(2026, 4, 1), date(2026, 7, 4))
    assert all(l.startswith("нед. ") for l in labels)
    assert all(k.weekday() == 0 for k in keys)
    assert key_of(date(2026, 7, 2)) == date(2026, 6, 29)  # четверг → его понедельник

    # >200 дней — по месяцам, подпись с полным годом («июл 2026», не «июл 26»)
    labels, keys, key_of = cashflow_buckets(date(2025, 8, 1), date(2026, 7, 4))
    assert labels[-1] == "июл 2026" and labels[0] == "авг 2025"
    assert keys[-1] == (2026, 7)
    assert key_of(date(2026, 7, 2)) == (2026, 7)

    # границы переключения шага
    assert cashflow_buckets(date(2026, 1, 1), date(2026, 2, 14))[2](date(2026, 1, 5)) == date(2026, 1, 5)   # 45 дн — дни
    assert cashflow_buckets(date(2026, 1, 1), date(2026, 2, 15))[2](date(2026, 1, 7)) == date(2026, 1, 5)   # 46 дн — недели


def test_render_finances_period_presets():
    html = _render(
        "finances.html", owner=OWNER, active_page="finances",
        summary={"total_income": Decimal("93000"), "total_expense": Decimal("0"),
                 "profit": Decimal("93000"), "fuel": Decimal("0")},
        margin=100.0,
        cashflow={"labels": ["01.07"], "revenue": [93000], "expenses": [0],
                  "profit": [93000], "period": "custom"},
        directions=[], entries=[], period_from="2026-07-01", period_to="2026-07-05",
        today="2026-07-04")
    assert "Денежный поток" in html and "2026-07-01 — 2026-07-05" in html
    for days in ("7", "30", "91", "182", "365"):
        assert f'data-days="{days}"' in html
    assert "fin-preset" in html and "fin-period-form" in html


# ------------------------------------------------------------------ скрытая команда /wipe
def test_wipe_not_in_help_and_phrase_consistent():
    import ast
    from app.bots import messages as msg

    assert "wipe" not in msg.OWNER_HELP.lower(), "/wipe не должен светиться в /help"

    # фраза подтверждения объявлена в сервисе…
    tree = ast.parse(open("app/services/maintenance_service.py", encoding="utf-8").read())
    phrase = next(
        ast.literal_eval(node.value) for node in tree.body
        if isinstance(node, ast.Assign)
        and any(getattr(t, "id", "") == "WIPE_CONFIRM_PHRASE" for t in node.targets)
    )
    assert phrase == "УДАЛИТЬ ВСЁ"

    # …и бот использует именно её (и не переопределяет свою)
    bot_src = open("app/bots/owner_bot.py", encoding="utf-8").read()
    assert 'Command("wipe")' in bot_src
    assert "maintenance_service.WIPE_CONFIRM_PHRASE" in bot_src
    assert "wipe_owner_data" in bot_src


def test_wipe_deletes_everything_including_owner():
    """Полный ноль: /wipe стирает и сам аккаунт (бот начинает регистрацию заново)."""
    src = open("app/services/maintenance_service.py", encoding="utf-8").read()
    for model in ("Owner", "Admin", "WebSession", "Subscription", "Customer",
                  "DistributionCenter", "Driver", "Vehicle", "Trip", "Shift",
                  "Expense", "Event", "VehicleTelemetryPoint", "VehicleTelemetryRawPacket"):
        assert f"delete({model})" in src, f"wipe должен удалять {model}"
    # аккаунт удаляется последним
    assert src.index("delete(Owner)") > src.index("delete(WebSession)")


def test_receiver_skips_unknown_terminals_and_acks():
    """Пакеты чужих трекеров не пишутся в БД (мусор), но ACK отправляется."""
    src = open("app/telemetry/egts_receiver.py", encoding="utf-8").read()
    assert "EGTS skipped" in src
    skip_block = src.split("EGTS skipped")[1][:400]
    assert "build_response(parsed)" in skip_block  # ACK до какой-либо записи в БД


def test_telemetry_cleanup_job_registered():
    jobs_src = open("app/services/scheduler_jobs.py", encoding="utf-8").read()
    assert "telemetry_cleanup_job" in jobs_src
    assert "RAW_PACKETS_KEEP_DAYS = 7" in jobs_src
    main_src = open("app/main.py", encoding="utf-8").read()
    assert "telemetry_cleanup_job" in main_src


def test_stats_routes_have_correct_decorators():
    """Регресс-защита: GET /stats должен вести на страницу статистики, а не на
    POST-обработчик скрытия стоянки (был баг — два декоратора на одной функции
    → GET /stats требовал event_id и падал с 'Field required')."""
    src = open("app/web/router.py", encoding="utf-8").read()
    # страница статистики — со своим GET-декоратором
    assert '@app.get("/stats", response_class=HTMLResponse)\nasync def stats_page(' in src
    # GET /stats НЕ должен стоять прямо перед POST-обработчиком
    assert '@app.get("/stats", response_class=HTMLResponse)\n@app.post(' not in src
    # обработчики стоянки — отдельные POST-роуты с event_id в ПУТИ
    for action in ("hide", "correct", "unhide"):
        assert f'@app.post("/stats/downtime/{{event_id}}/{action}")' in src


def test_no_route_decorator_on_helper_functions():
    """Регресс-защита от краша деплоя: роут-декоратор @app.get/post/... НЕ должен
    висеть на функции-хелпере (имя с '_'). Такое случается, когда новую функцию
    вставили МЕЖДУ декоратором и его функцией — FastAPI падает на импорте
    (session: AsyncSession трактуется как поле). Ловим статически через ast."""
    import ast
    src = open("app/web/router.py", encoding="utf-8").read()
    tree = ast.parse(src)
    http_methods = {"get", "post", "put", "delete", "patch"}
    offenders = []
    for node in ast.walk(tree):
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            for dec in node.decorator_list:
                if (isinstance(dec, ast.Call) and isinstance(dec.func, ast.Attribute)
                        and isinstance(dec.func.value, ast.Name)
                        and dec.func.value.id == "app" and dec.func.attr in http_methods
                        and node.name.startswith("_")):
                    offenders.append(f"{dec.func.attr.upper()} → {node.name} (строка {node.lineno})")
    assert not offenders, "роут-декоратор на функции-хелпере: " + "; ".join(offenders)


# ------------------------------------------------- смена: фактический выезд по GPS
OWNER_TZ = NS(**{**vars(OWNER), "timezone": "Europe/Moscow"})


def _shift_ns(status="started"):
    return NS(id=7, is_manual=False,
              started_at=datetime(2026, 7, 14, 5, 0, tzinfo=timezone.utc),
              ended_at=None, odometer_start=None, odometer_end=None,
              distance_km=None, status=status, driver_id=1, vehicle_id=1,
              odometer_start_photo_url=None, odometer_end_photo_url=None)


def _render_shift(**over):
    ctx = dict(owner=OWNER_TZ, shift=_shift_ns(), driver=NS(full_name="Саломов"),
               vehicle=NS(license_plate="Т557ОС178"), trips=[], expenses=[],
               photo_start_at=None, photo_end_at=None,
               gps_seen=False, gps_departure_at=None, gps_departure_delay=None,
               active_page="trips")
    ctx.update(over)
    return _render("shift_detail.html", **ctx)


def test_render_shift_detail_gps_departure_shown():
    html = _render_shift(
        gps_seen=True,
        gps_departure_at=datetime(2026, 7, 14, 5, 12, tzinfo=timezone.utc),
        gps_departure_delay="через 12 мин",
    )
    assert "Выехал · по GPS" in html
    assert "08:12" in html            # 05:12 UTC = 08:12 МСК
    assert "через 12 мин" in html


def test_render_shift_detail_gps_not_departed_yet():
    html = _render_shift(gps_seen=True)  # смена активна, движения нет
    assert "ещё не выехал" in html and "машина стоит" in html


def test_render_shift_detail_no_gps_block_hidden():
    html = _render_shift(gps_seen=False)
    assert "по GPS" not in html  # нет телематики — ничего не выдумываем


# ------------------------------------------------- машины: бейдж «в смене/в рейсе»
def _vehicle_row(state):
    v = NS(id=1, license_plate="Т557ОС178", brand="Рено", type="refrigerator",
           stavtrack_object_id=None, fuel_norm_per_100km=None,
           osago_expires=None, inspection_expires=None, tacho_expires=None)
    return {"vehicle": v, "km": 0, "fuel": Decimal(0), "trips": 0,
            "revenue": Decimal(0), "expense": Decimal(0), "profit": Decimal(0),
            "margin": Decimal(0), "active": state != "free", "state": state}


@pytest.mark.parametrize("state,badge", [
    ("trip", "в рейсе"),
    ("shift", "в смене"),
    ("free", "свободна"),
])
def test_render_vehicle_row_badge_reflects_real_state(state, badge):
    html = _render("_vehicle_row.html", row=_vehicle_row(state))
    assert badge in html
    # «в рейсе» не должно показываться, когда машина просто в смене
    if state == "shift":
        assert "в рейсе" not in html


# ------------------------------------------------- хронология смены по GPS
from datetime import timedelta  # noqa: E402

from app.services import telemetry_service as TS  # noqa: E402

_BASE = datetime(2026, 7, 14, 8, 0, tzinfo=timezone.utc)


def _points(*specs):
    """specs: (минута_от_начала, скорость_кмч)."""
    return [(_BASE + timedelta(minutes=m), Decimal(s)) for m, s in specs]


def test_segments_stop_move_stop():
    # стоял 30 мин → ехал час → стоял до конца окна
    pts = _points(*[(m, 0) for m in range(0, 30, 5)],
                  *[(m, 40) for m in range(30, 90, 5)],
                  *[(m, 0) for m in range(90, 121, 5)])
    segs = TS.segment_movements(pts, window_end=_BASE + timedelta(minutes=120))
    assert [s["kind"] for s in segs] == ["stop", "move", "stop"]
    assert segs[0]["start"] == _BASE
    assert segs[1]["start"] == _BASE + timedelta(minutes=30)  # фактический выезд
    assert segs[1]["end"] == _BASE + timedelta(minutes=90)
    assert segs[-1]["end"] == _BASE + timedelta(minutes=120)


def test_segments_traffic_light_does_not_split_trip():
    # минутная остановка на светофоре посреди езды не рвёт поездку
    pts = _points(*[(m, 40) for m in range(0, 20)],
                  (20, 0),  # 1 минута стоянки
                  *[(m, 40) for m in range(21, 40)])
    segs = TS.segment_movements(pts, window_end=_BASE + timedelta(minutes=40))
    assert [s["kind"] for s in segs] == ["move"]


def test_segments_gap_marked_as_nosignal():
    # точки шли до 10-й минуты, потом трекер молчал до 40-й
    pts = _points(*[(m, 0) for m in range(0, 11, 5)],
                  (40, 0))
    segs = TS.segment_movements(pts, window_end=_BASE + timedelta(minutes=45))
    kinds = [s["kind"] for s in segs]
    assert kinds == ["stop", "nosignal", "stop"]
    # «нет сигнала» начинается через 15 мин (grace) после последней точки
    assert segs[1]["start"] == _BASE + timedelta(minutes=25)
    assert segs[1]["end"] == _BASE + timedelta(minutes=40)


def test_segments_tail_open_marks_ongoing():
    pts = _points((0, 0), (5, 0))
    segs = TS.segment_movements(
        pts, window_end=_BASE + timedelta(minutes=10), tail_open=True
    )
    assert segs[-1]["ongoing"] is True


def test_render_shift_detail_gps_timeline():
    timeline = [
        {"icon": "🅿️", "label": "Стоял", "frm": "08:00", "to": "08:30",
         "dur": "30 мин", "kind": "stop", "ongoing": False},
        {"icon": "🚚", "label": "Ехал", "frm": "08:30", "to": "09:30",
         "dur": "1 ч", "kind": "move", "ongoing": False},
        {"icon": "🅿️", "label": "Стоит", "frm": "09:30", "to": "сейчас",
         "dur": "25 мин", "kind": "stop", "ongoing": True},
    ]
    html = _render_shift(gps_seen=True, gps_timeline=timeline)
    assert "Хронология смены" in html
    assert "08:30 → 09:30" in html
    assert "Стоит" in html and "сейчас" in html


def test_render_shift_detail_no_timeline_when_no_gps():
    html = _render_shift(gps_seen=False, gps_timeline=[])
    assert "Хронология смены" not in html
