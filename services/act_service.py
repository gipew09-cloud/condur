"""
Акт оказанных услуг по РЦ за период (близко к форме 101 РС).

Группируем завершённые рейсы по пункту назначения (РЦ) и за выбранные даты
делаем .xlsx: по каждому РЦ — отдельный лист с рейсами, итогом и суммой прописью.
Только openpyxl (уже в проекте), без внешних ключей и сервисов.
"""
import io
import re
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=13)
_MONEY = "#,##0 ₽"
_THIN = Side(style="thin", color="C9CFD9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# --- сумма прописью (рубли) ---
_ONES = [
    "", "один", "два", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять",
    "десять", "одиннадцать", "двенадцать", "тринадцать", "четырнадцать", "пятнадцать",
    "шестнадцать", "семнадцать", "восемнадцать", "девятнадцать",
]
_TENS = ["", "", "двадцать", "тридцать", "сорок", "пятьдесят", "шестьдесят", "семьдесят", "восемьдесят", "девяносто"]
_HUNDREDS = ["", "сто", "двести", "триста", "четыреста", "пятьсот", "шестьсот", "семьсот", "восемьсот", "девятьсот"]


def _ones_word(o: int, female: bool) -> str:
    if female and o == 1:
        return "одна"
    if female and o == 2:
        return "две"
    return _ONES[o]


def _triplet(n: int, female: bool = False) -> list[str]:
    out: list[str] = []
    h, rem = divmod(n, 100)
    t, o = divmod(rem, 10)
    if h:
        out.append(_HUNDREDS[h])
    if t >= 2:
        out.append(_TENS[t])
        if o:
            out.append(_ones_word(o, female))
    elif t == 1:
        out.append(_ONES[10 + o])
    elif o:
        out.append(_ones_word(o, female))
    return out


def _plural(n: int, forms: tuple[str, str, str]) -> str:
    n = abs(n) % 100
    if 10 < n < 20:
        return forms[2]
    n %= 10
    if n == 1:
        return forms[0]
    if 2 <= n <= 4:
        return forms[1]
    return forms[2]


def rubles_in_words(amount: Decimal) -> str:
    amount = Decimal(amount)
    rub = int(amount)
    kop = int(round((amount - rub) * 100))
    parts: list[str] = []
    if rub == 0:
        parts = ["ноль"]
    else:
        millions, rest = divmod(rub, 1_000_000)
        thousands, ones = divmod(rest, 1000)
        if millions:
            parts += _triplet(millions)
            parts.append(_plural(millions, ("миллион", "миллиона", "миллионов")))
        if thousands:
            parts += _triplet(thousands, female=True)
            parts.append(_plural(thousands, ("тысяча", "тысячи", "тысяч")))
        if ones:
            parts += _triplet(ones)
    rub_word = _plural(rub, ("рубль", "рубля", "рублей"))
    kop_word = _plural(kop, ("копейка", "копейки", "копеек"))
    s = " ".join(w for w in parts if w)
    s = s[:1].upper() + s[1:]
    return f"{s} {rub_word} {kop:02d} {kop_word}."


def _safe_sheet_title(name: str, used: set[str]) -> str:
    title = re.sub(r"[\\/\?\*\[\]:]", " ", name or "РЦ").strip()[:28] or "РЦ"
    base, i = title, 2
    while title in used:
        title = f"{base[:26]} {i}"
        i += 1
    used.add(title)
    return title


def build_acts_workbook(
    groups: dict[str, list[dict]], df: date, dt: date, company: str
) -> Workbook:
    """groups: РЦ → список рейсов {date, origin, destination, plate, driver, revenue}."""
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    period = f"{df.strftime('%d.%m.%Y')} — {dt.strftime('%d.%m.%Y')}"

    if not groups:
        ws = wb.create_sheet(_safe_sheet_title("Нет данных", used))
        ws["A1"] = f"За период {period} завершённых рейсов с выручкой нет."
        return wb

    for rc, trips in groups.items():
        ws = wb.create_sheet(_safe_sheet_title(rc, used))
        ws["A1"] = company or "Перевозчик"
        ws["A1"].font = _TITLE_FONT
        ws["A2"] = "АКТ оказанных транспортных услуг"
        ws["A2"].font = Font(bold=True)
        ws["A3"] = f"Заказчик (РЦ): {rc}"
        ws["A4"] = f"Период: {period}"

        header_row = 6
        headers = ["№", "Дата", "Маршрут", "Машина", "Водитель", "Сумма"]
        for col, title in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col, value=title)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _BORDER

        total = Decimal(0)
        r = header_row + 1
        for i, t in enumerate(trips, start=1):
            revenue = Decimal(t["revenue"] or 0)
            total += revenue
            values = [
                i,
                t["date"].strftime("%d.%m.%Y") if t["date"] else "—",
                f"{t['origin'] or '—'} → {t['destination'] or '—'}",
                t["plate"] or "—",
                t["driver"] or "—",
                float(revenue),
            ]
            for col, v in enumerate(values, start=1):
                c = ws.cell(row=r, column=col, value=v)
                c.border = _BORDER
                if col == 6:
                    c.number_format = _MONEY
                    c.alignment = Alignment(horizontal="right")
            r += 1

        # ИТОГО
        ws.cell(row=r, column=5, value="ИТОГО:").font = Font(bold=True)
        total_cell = ws.cell(row=r, column=6, value=float(total))
        total_cell.font = Font(bold=True)
        total_cell.number_format = _MONEY
        total_cell.alignment = Alignment(horizontal="right")
        ws.cell(row=r + 2, column=1, value=f"Всего оказано услуг на сумму: {rubles_in_words(total)}")

        widths = [5, 13, 38, 14, 22, 15]
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w

    return wb


# =====================================================================
# НАСТОЯЩАЯ ФОРМА АКТА (101 РС): один лист, шапка с реквизитами сторон,
# колонки № · Наименование · Кол-во · Ед. · Цена · Сумма, итог прописью,
# подписи. Реплика образца «Рузисеть ИП Кибиткина № 101 РС».
# executor/customer — словари реквизитов; rows — плоский список рейсов.
# =====================================================================
_ACT_NCOL = 6  # A..F
# Стиль «как у образца 101 РС»: мелкий шрифт (9 пт), чёрные тонкие рамки,
# шапка таблицы — без заливки (жирный чёрный текст), деньги «#,##0.00» без ₽.
_LABEL_FONT = Font(name="Calibri", bold=True, size=9)
_ACT_BASE = Font(name="Calibri", size=9)
_TITLE_BIG = Font(name="Calibri", bold=True, size=13)
_ACT_MONEY = "#,##0.00"
_ACT_THIN = Side(style="thin", color="000000")
_ACT_BORDER = Border(left=_ACT_THIN, right=_ACT_THIN, top=_ACT_THIN, bottom=_ACT_THIN)
_ACT_WIDTHS = {"A": 5, "B": 62, "C": 9, "D": 7, "E": 14, "F": 16}


def _money_str(amount: Decimal) -> str:
    """292000 → '292 000,00' (пробел — разряды, запятая — копейки)."""
    s = f"{Decimal(amount):,.2f}"            # '292,000.00'
    return s.replace(",", " ").replace(".", ",")


def _executor_text(e: dict) -> str:
    parts = [e.get("full_name") or ""]
    if e.get("inn"):
        parts.append(f"ИНН {e['inn']}")
    if e.get("ogrnip"):
        parts.append(f"ОГРНИП {e['ogrnip']}")
    if e.get("address"):
        parts.append(e["address"])
    if e.get("bank_name"):
        parts.append(f"в банке {e['bank_name']}")
    if e.get("account"):
        parts.append(f"р/с {e['account']}")
    if e.get("corr_account"):
        parts.append(f"к/с {e['corr_account']}")
    if e.get("bik"):
        parts.append(f"БИК {e['bik']}")
    return ", ".join(p for p in parts if p)


def _customer_text(c: dict) -> str:
    parts = [c.get("name") or ""]
    if c.get("inn"):
        parts.append(f"ИНН {c['inn']}")
    if c.get("kpp"):
        parts.append(f"КПП {c['kpp']}")
    if c.get("address"):
        parts.append(c["address"])
    if c.get("account"):
        parts.append(f"р/с {c['account']}")
    if c.get("bank_name"):
        parts.append(f"в банке {c['bank_name']}")
    if c.get("bik"):
        parts.append(f"БИК {c['bik']}")
    if c.get("corr_account"):
        parts.append(f"к/с {c['corr_account']}")
    return ", ".join(p for p in parts if p)


def _contract_text(c: dict) -> str:
    num = c.get("contract_number")
    d = c.get("contract_date")
    if num and d:
        ds = d.strftime("%d.%m.%Y") if hasattr(d, "strftime") else str(d)
        return f"Договор {num} от {ds}"
    return f"Договор {num}" if num else "—"


def _row_description(row: dict) -> str:
    """'05.05.2026. Агропарк Софийская 151- РЦ ..., Т 557 ОС 178, Саломов Диер.'"""
    d = row.get("date")
    ds = d.strftime("%d.%m.%Y") if hasattr(d, "strftime") else str(d or "")
    origin = (row.get("origin") or "").strip()
    dest = (row.get("destination_address") or row.get("destination") or "").strip()
    plate = (row.get("plate") or "").strip()
    driver = (row.get("driver") or "").strip()
    route = f"{origin}- {dest}" if origin else dest
    tail = ", ".join(x for x in (plate, driver) if x)
    s = f"{ds}. {route}" if ds else route
    if tail:
        s += f", {tail}"
    return s + "."


def build_act_101rs(
    *,
    title: str = "Акт",
    act_number: str,
    act_date: date,
    period_from: date,
    period_to: date,
    executor: dict,
    customer: dict,
    rows: list[dict],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Акт")[:31]
    last = get_column_letter(_ACT_NCOL)
    for col, w in _ACT_WIDTHS.items():
        ws.column_dimensions[col].width = w

    def block(rng: str, value: str, *, font=None, align="left", wrap=False, vtop=False):
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = value
        c.font = font or _ACT_BASE
        c.alignment = Alignment(
            horizontal=align, vertical="top" if vtop else "center", wrap_text=wrap
        )
        return c

    r = 1
    block(
        f"A{r}:{last}{r}",
        f"{title or 'Акт'} № {act_number} от {act_date.strftime('%d.%m.%Y')} г.",
        font=_TITLE_BIG, align="center",
    )
    ws.row_dimensions[r].height = 22
    r += 2

    for label, text, height in (
        ("Исполнитель:", _executor_text(executor), 46),
        ("Заказчик:", _customer_text(customer), 46),
        ("Основание:", _contract_text(customer), 16),
    ):
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = _LABEL_FONT
        cell.alignment = Alignment(vertical="top")
        block(f"B{r}:{last}{r}", text, wrap=True, vtop=True)
        ws.row_dimensions[r].height = height
        r += 1
    r += 1

    headers = ["№", "Наименование работ, услуг", "Кол-во", "Ед.", "Цена, руб", "Сумма, руб"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = _LABEL_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _ACT_BORDER
    r += 1

    total = Decimal(0)
    for idx, row in enumerate(rows, start=1):
        amount = Decimal(row.get("amount") or 0)
        total += amount
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=_row_description(row))
        ws.cell(row=r, column=3, value=1)
        ws.cell(row=r, column=4, value="усл.")
        ws.cell(row=r, column=5, value=float(amount))
        ws.cell(row=r, column=6, value=float(amount))
        for col in range(1, _ACT_NCOL + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = _ACT_BORDER
            cell.font = _ACT_BASE
            if col == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col in (1, 3, 4):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="top")
                cell.number_format = _ACT_MONEY
        r += 1

    lbl = ws.cell(row=r, column=5, value="Итого:")
    lbl.font = _LABEL_FONT
    lbl.alignment = Alignment(horizontal="right")
    tot = ws.cell(row=r, column=6, value=float(total))
    tot.font = _LABEL_FONT
    tot.number_format = _ACT_MONEY
    tot.alignment = Alignment(horizontal="right")
    tot.border = _ACT_BORDER
    r += 2

    block(
        f"A{r}:{last}{r}",
        f"Всего оказано услуг {len(rows)}. На сумму {_money_str(total)} руб. "
        f"{rubles_in_words(total)}",
        font=_LABEL_FONT, wrap=True, vtop=True,
    )
    ws.row_dimensions[r].height = 30
    r += 1
    block(f"A{r}:{last}{r}", "Без налога (НДС)")
    r += 2

    block(
        f"A{r}:{last}{r}",
        "Вышеперечисленные услуги выполнены полностью и в срок. "
        "Заказчик по объёму, качеству и срокам претензий не имеет.",
        wrap=True, vtop=True,
    )
    ws.row_dimensions[r].height = 28
    r += 2

    ws.cell(row=r, column=1, value="Исполнитель").font = _LABEL_FONT
    ws.cell(row=r, column=5, value="Заказчик").font = _LABEL_FONT
    r += 1
    ws.cell(row=r, column=1, value=executor.get("full_name") or "").font = _ACT_BASE
    ws.cell(row=r, column=5, value=customer.get("name") or "").font = _ACT_BASE
    r += 2
    ws.cell(row=r, column=1, value=f"__________ / {executor.get('signer_name') or ''}").font = _ACT_BASE
    ws.cell(row=r, column=5, value=f"__________ / {customer.get('signer_name') or ''}").font = _ACT_BASE

    return wb


def workbook_bytes(wb: Workbook) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
