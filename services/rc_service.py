"""Утилиты справочника РЦ: нормализация, поиск адреса и импорт Excel."""
from __future__ import annotations

import io
import re
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook


def route_key(value: str | None) -> str:
    value = (value or "").casefold().replace("ё", "е")
    return re.sub(r"[^0-9a-zа-я]+", " ", value).strip()


def split_aliases(value: str | None) -> list[str]:
    return [p.strip() for p in re.split(r"[\n,;|]+", value or "") if p.strip()]


def distribution_center_lookup(centers: list) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for center in centers:
        address = (getattr(center, "address", None) or getattr(center, "name", None) or "").strip()
        if not address:
            continue
        terms = [
            getattr(center, "name", None),
            getattr(center, "address", None),
            *split_aliases(getattr(center, "aliases", None)),
        ]
        for term in terms:
            key = route_key(term)
            if key:
                lookup.setdefault(key, address)
    return lookup


def canonical_rc_address(destination: str | None, lookup: dict[str, str]) -> str | None:
    key = route_key(destination)
    if not key:
        return None
    if key in lookup:
        return lookup[key]
    for alias_key, address in sorted(lookup.items(), key=lambda item: len(item[0]), reverse=True):
        if len(alias_key) >= 4 and (alias_key in key or key in alias_key):
            return address
    return None


def match_destination_to_center(destination: str | None, centers: list):
    """РЦ из справочника, соответствующий тексту назначения рейса (по
    названию/адресу/алиасам). Нужно для сверки план↔факт. None — не распознали."""
    key = route_key(destination)
    if not key:
        return None
    by_key: dict[str, object] = {}
    for center in centers:
        terms = [
            getattr(center, "name", None),
            getattr(center, "address", None),
            *split_aliases(getattr(center, "aliases", None)),
        ]
        for term in terms:
            k = route_key(term)
            if k:
                by_key.setdefault(k, center)
    if key in by_key:
        return by_key[key]
    for k, center in sorted(by_key.items(), key=lambda x: len(x[0]), reverse=True):
        if len(k) >= 4 and (k in key or key in k):
            return center
    return None


def decimal_or_none(value: str | int | float | Decimal | None) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError("Координаты должны быть числами") from exc


def haversine_m(
    lat1: Decimal | float, lon1: Decimal | float,
    lat2: Decimal | float, lon2: Decimal | float,
) -> float:
    """Расстояние между двумя точками по прямой, в метрах (формула гаверсинуса).
    Для геозон РЦ точности сферической Земли более чем достаточно."""
    from math import asin, cos, radians, sin, sqrt

    lat1, lon1, lat2, lon2 = (radians(float(x)) for x in (lat1, lon1, lat2, lon2))
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
    return 2 * 6_371_000 * asin(sqrt(a))


def _header_token(value) -> str:
    return route_key(str(value or ""))


def _find_xlsx_col(headers: list, variants: tuple[str, ...]) -> int | None:
    needles = tuple(_header_token(v) for v in variants)
    for idx, raw in enumerate(headers):
        token = _header_token(raw)
        if token and any(needle == token or needle in token for needle in needles):
            return idx
    return None


def _xlsx_value(row: tuple, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    return str(value).strip() if value is not None else ""


def _looks_like_rc_header(row: tuple) -> bool:
    headers = list(row)
    return (
        _find_xlsx_col(headers, ("рц", "название", "name", "точка")) is not None
        and _find_xlsx_col(headers, ("адрес", "address")) is not None
    )


def distribution_centers_from_xlsx(data: bytes) -> list[dict]:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось прочитать Excel-файл") from exc
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = next((i for i, row in enumerate(rows[:10]) if _looks_like_rc_header(row)), None)
    if header_idx is None:
        headers: list = []
        data_rows = rows
    else:
        headers = list(rows[header_idx])
        data_rows = rows[header_idx + 1:]

    name_idx = _find_xlsx_col(headers, ("рц", "название", "наименование", "name", "точка"))
    address_idx = _find_xlsx_col(headers, ("адрес", "address"))
    alias_idx = _find_xlsx_col(headers, ("алиас", "синоним", "alias", "вариант"))
    lat_idx = _find_xlsx_col(headers, ("широта", "latitude", "lat"))
    lon_idx = _find_xlsx_col(headers, ("долгота", "longitude", "lon", "lng"))

    parsed: list[dict] = []
    for row in data_rows:
        cells = tuple(row)
        name = _xlsx_value(cells, name_idx) or _xlsx_value(cells, 0)
        address = _xlsx_value(cells, address_idx) or _xlsx_value(cells, 1)
        # Одноколоночный файл (как «РЦ Адреса Спб.xlsx» владельца): название и
        # адрес слиты в одной ячейке — берём её и как адрес.
        if name and not address:
            address = name
        if not name or not address:
            continue
        if header_idx is None and _looks_like_rc_header(cells):
            continue
        parsed.append({
            "name": name,
            "address": address,
            "aliases": _xlsx_value(cells, alias_idx),
            "latitude": _xlsx_value(cells, lat_idx),
            "longitude": _xlsx_value(cells, lon_idx),
        })
    return parsed
