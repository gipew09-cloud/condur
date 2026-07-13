"""
Веб-кабинет владельца.

Архитектура:
  - один FastAPI-app, который монтируется к asyncio.gather в main.py;
  - сессия БД даётся через Depends(get_session);
  - текущий владелец — через Depends(current_owner). На неавторизованных
    запросах кидаем RedirectResponse на /login;
  - все шаблоны Jinja2 рендерятся из app/web/templates;
  - HTMX определяется по заголовку HX-Request: для фильтра рейсов и
    редакта водителей возвращаем не полную страницу, а партиал.
"""
from __future__ import annotations

import io
import re
import uuid
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Annotated, AsyncIterator
from urllib.parse import quote

from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, Request, Response, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, delete, desc, func, or_, select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import async_session
from app.models import (
    Admin,
    Customer,
    DistributionCenter,
    Driver,
    Event,
    Expense,
    ManualEntry,
    Owner,
    RouteTemplate,
    Shift,
    Trip,
    TripDocument,
    Vehicle,
    VehicleState,
    WebSession,
)
from app.config import settings
from app.services import act_service, auth_service, billing, geocode_service, rc_service, telemetry_service
from app.services.event_service import log_event
from app.services.timeutil import (
    RU_MONTHS_SHORT,
    add_months,
    cashflow_buckets,
    fmt_dt,
    month_floor,
    owner_tz,
    smart_since_label,
)
from app.web.insights import generate_insights

# --------- инициализация ----------
BASE_DIR = Path(__file__).parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

# Локализация enum'ов из БД для отображения в шаблонах.
_VEHICLE_TYPE_LABELS = {
    "truck": "Грузовик",
    "gazelle": "Газель / фургон",
    "refrigerator": "Рефрижератор",
}
_TRIP_STATUS_LABELS = {
    "created": "создан",
    "in_transit": "в пути",
    "unloading": "на выгрузке",
    "completed": "завершён",
    "cancelled": "отменён",
}


def _vehicle_type_label(code: str | None) -> str:
    return _VEHICLE_TYPE_LABELS.get(code or "", code or "—")


def _trip_status_label(code: str | None) -> str:
    return _TRIP_STATUS_LABELS.get(code or "", code or "—")


# Цвет статусной плашки по значению статуса (рейс/расход/смена) — стиль «Свежий путь».
_PILL_CLASS = {
    "completed": "pill--success", "approved": "pill--success",
    "created": "pill--info", "in_transit": "pill--info", "started": "pill--info",
    "pending": "pill--warn", "unloading": "pill--warn",
    "rejected": "pill--danger", "cancelled": "pill--danger",
}


def _pill_class(status: str | None) -> str:
    return _PILL_CLASS.get((status or "").lower(), "pill--neutral")


# Русские подписи статусов (расход/смена/рейс) на уровне отображения — Правка 6.
_STATUS_RU = {
    "approved": "одобрен", "pending": "на проверке", "rejected": "отклонён",
    "started": "в смене", "completed": "завершён", "cancelled": "отменён",
    "created": "создан", "in_transit": "в пути", "unloading": "на выгрузке",
}


def _status_ru(status: str | None) -> str:
    return _STATUS_RU.get((status or "").lower(), status or "—")


def _local_dt(value: datetime | None, timezone_name: str | None = None, fmt: str = "%d.%m.%Y %H:%M") -> str:
    return fmt_dt(value, timezone_name, fmt)


templates.env.filters["vtype"] = _vehicle_type_label
templates.env.filters["tstatus"] = _trip_status_label
templates.env.filters["pillclass"] = _pill_class
templates.env.filters["statusru"] = _status_ru
templates.env.filters["localdt"] = _local_dt

app = FastAPI(title="TMS Cabinet")

# static (минимум — favicon/css)
static_dir = BASE_DIR / "static"
static_dir.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")


# --------- зависимости ----------
async def get_session() -> AsyncIterator[AsyncSession]:
    async with async_session() as session:
        yield session


_LOGIN_REDIRECT = HTTPException(status_code=303, headers={"Location": "/login"})


async def _session_from_request(
    request: Request, session: AsyncSession
) -> WebSession | None:
    """Активная веб-сессия по cookie (или None). Обновляет last_seen раз в 5 мин."""
    raw = request.cookies.get(auth_service.SESSION_COOKIE)
    if not raw:
        return None
    ws = (
        await session.execute(
            select(WebSession).where(
                WebSession.token_hash == auth_service.session_token_hash(raw),
                WebSession.revoked_at.is_(None),
            )
        )
    ).scalar_one_or_none()
    if ws is None:
        return None
    now = datetime.now(timezone.utc)
    if ws.last_seen_at is None or (now - ws.last_seen_at).total_seconds() > 300:
        ws.last_seen_at = now
        await session.commit()
    return ws


async def _viewer_telegram_id(request: Request, session: AsyncSession) -> int | None:
    """Telegram ID того, кто сейчас смотрит кабинет (сессия или старый JWT).
    None — старый владельческий JWT без tid; трактуем как владельца."""
    ws = await _session_from_request(request, session)
    if ws is not None:
        return ws.telegram_id
    token = request.cookies.get("auth")
    if token:
        decoded = auth_service.decode_jwt(token)
        if decoded is not None:
            return decoded[1]
    return None


async def current_owner(
    request: Request,
    session: Annotated[AsyncSession, Depends(get_session)],
) -> Owner:
    # 1) Постоянная сессия (основной путь).
    ws = await _session_from_request(request, session)
    if ws is not None:
        owner_id, tid = ws.owner_id, ws.telegram_id
    else:
        # 2) Переходный fallback: старый 7-дневный JWT (чтобы никого не
        # разлогинило при деплое). Истечёт сам — дальше вход уже постоянный.
        token = request.cookies.get("auth")
        if not token:
            raise _LOGIN_REDIRECT
        decoded = auth_service.decode_jwt(token)
        if decoded is None:
            raise _LOGIN_REDIRECT
        owner_id, tid = decoded

    owner = await session.get(Owner, owner_id)
    if owner is None:
        raise _LOGIN_REDIRECT
    # Если вошёл админ (tid не совпадает с владельцем) — проверяем, что доступ
    # ещё не отозван. Удаление админа гасит и все его устройства сразу.
    if tid is not None and tid != owner.telegram_id:
        admin = (
            await session.execute(
                select(Admin).where(Admin.telegram_id == tid, Admin.owner_id == owner_id)
            )
        ).scalar_one_or_none()
        if admin is None:
            raise _LOGIN_REDIRECT
    return owner


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    """Если кинули 303 с Location — редиректим. Иначе стандартное поведение."""
    if exc.status_code == 303 and "Location" in (exc.headers or {}):
        return RedirectResponse(exc.headers["Location"], status_code=303)
    from fastapi.exception_handlers import http_exception_handler as default_handler
    return await default_handler(request, exc)


# --------- утилиты ----------
def _is_htmx(request: Request) -> bool:
    return request.headers.get("HX-Request") == "true"


def _month_window(today: date | None = None) -> tuple[datetime, datetime]:
    today = today or date.today()
    start = datetime(today.year, today.month, 1, tzinfo=timezone.utc)
    return start, datetime.now(timezone.utc)


def _today_window() -> tuple[datetime, datetime]:
    now = datetime.now(timezone.utc)
    start = datetime(now.year, now.month, now.day, tzinfo=timezone.utc)
    return start, now


async def _period_totals(
    session: AsyncSession, owner_id: int, dt_from: datetime, date_from: date
) -> dict:
    """
    Считает доход / расход / прибыль за период.
      доход = SUM(trips.revenue_rub completed) + SUM(manual_entries income)
      расход = SUM(expenses.amount_rub approved) + SUM(manual_entries expense)
              fuel_cost_rub из trips НЕ суммируем — топливо уже в expenses
              (по нашему flow «Расход → Топливо»), иначе двойной счёт.
    """
    trip_revenue = (
        await session.execute(
            select(func.coalesce(func.sum(Trip.revenue_rub), 0)).where(
                Trip.owner_id == owner_id,
                Trip.status == "completed",
                Trip.completed_at >= dt_from,
            )
        )
    ).scalar_one() or Decimal(0)

    manual_income = (
        await session.execute(
            select(func.coalesce(func.sum(ManualEntry.amount_rub), 0)).where(
                ManualEntry.owner_id == owner_id,
                ManualEntry.type == "income",
                ManualEntry.entry_date >= date_from,
            )
        )
    ).scalar_one() or Decimal(0)

    approved_expenses = (
        await session.execute(
            select(func.coalesce(func.sum(Expense.amount_rub), 0)).where(
                Expense.owner_id == owner_id,
                Expense.status == "approved",
                Expense.created_at >= dt_from,
            )
        )
    ).scalar_one() or Decimal(0)

    manual_expense = (
        await session.execute(
            select(func.coalesce(func.sum(ManualEntry.amount_rub), 0)).where(
                ManualEntry.owner_id == owner_id,
                ManualEntry.type == "expense",
                ManualEntry.entry_date >= date_from,
            )
        )
    ).scalar_one() or Decimal(0)

    income = Decimal(trip_revenue) + Decimal(manual_income)
    expense = Decimal(approved_expenses) + Decimal(manual_expense)
    return {
        "income": income,
        "expense": expense,
        "profit": income - expense,
    }


# =========================================================================
# /login
# =========================================================================
@app.get("/login", response_class=HTMLResponse)
async def login_form(request: Request):
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


@app.post("/login")
async def login_submit(
    request: Request,
    telegram_id: Annotated[str, Form()],
    code: Annotated[str, Form()],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    try:
        tg_id = int(telegram_id.strip())
    except ValueError:
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Telegram ID должен быть числом."},
            status_code=400,
        )

    if not auth_service.consume_code(tg_id, code):
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Неверный или истёкший код."},
            status_code=400,
        )

    result = await session.execute(select(Owner).where(Owner.telegram_id == tg_id))
    owner = result.scalar_one_or_none()
    owner_id = owner.id if owner is not None else None
    if owner_id is None:
        # Не владелец — может быть админ чьего-то кабинета (полный доступ).
        admin = (
            await session.execute(select(Admin).where(Admin.telegram_id == tg_id))
        ).scalar_one_or_none()
        if admin is not None:
            owner_id = admin.owner_id
    if owner_id is None:
        return templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "error": "Доступ не найден. Владельцу — /start в боте; "
                "админа добавляет владелец в разделе «Реквизиты».",
            },
            status_code=400,
        )

    # Постоянная сессия: живёт, пока не завершат («Выйти» на устройстве или
    # владелец на «Реквизиты → Устройства»). В cookie — токен, в БД — его hash.
    raw_token = auth_service.new_session_token()
    ip = (request.headers.get("x-forwarded-for") or "").split(",")[0].strip() or (
        request.client.host if request.client else None
    )
    web_session = WebSession(
        owner_id=owner_id,
        telegram_id=tg_id,
        token_hash=auth_service.session_token_hash(raw_token),
        device_label=auth_service.device_label_from_user_agent(
            request.headers.get("user-agent")
        ),
        ip=ip,
        last_seen_at=datetime.now(timezone.utc),
    )
    session.add(web_session)
    await session.commit()

    response = RedirectResponse("/", status_code=303)
    # Secure + явный Expires + Path=/ — иначе на iPhone/Safari cookie слетает
    # и вход постоянно просит код заново (см. auth_service.set_session_cookie).
    auth_service.set_session_cookie(response, raw_token)
    # Старый 7-дневный JWT больше не нужен: после успешного входа оставляем
    # только постоянную web_session, чтобы браузер не путался между схемами.
    response.delete_cookie("auth")
    return response


@app.get("/logout")
async def logout(
    request: Request,
    session: Annotated[AsyncSession, Depends(get_session)],
):
    ws = await _session_from_request(request, session)
    if ws is not None:
        ws.revoked_at = datetime.now(timezone.utc)
        await session.commit()
    response = RedirectResponse("/login", status_code=303)
    auth_service.clear_session_cookie(response)
    response.delete_cookie("auth")
    return response


# =========================================================================
# Dashboard
# =========================================================================
_RU_WEEKDAYS = [
    "Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье",
]
_RU_MONTHS_GEN = [
    "января", "февраля", "марта", "апреля", "мая", "июня",
    "июля", "августа", "сентября", "октября", "ноября", "декабря",
]
_RU_MONTHS_NOM = [
    "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
]
_EXPENSE_CAT_LABELS = {
    "fuel": "Топливо", "repair": "Ремонт", "parking": "Парковка",
    "fine": "Штрафы", "toll": "Дороги", "other": "Прочее",
}


async def _dashboard_overview(session: AsyncSession, owner: Owner) -> dict:
    """Данные для блоков дашборда «Свежий путь»: дата, «требуют внимания»,
    машины в работе, структура расходов, дельты KPI. Телематические поля
    (температура и т.п.) появятся, когда подключим телематику."""
    tz = owner_tz(owner.timezone)
    now_local = datetime.now(tz)
    date_label = (
        f"{_RU_WEEKDAYS[now_local.weekday()]}, {now_local.day} "
        f"{_RU_MONTHS_GEN[now_local.month - 1]} · {now_local:%H:%M}"
    )
    today = now_local.date()
    today_start, _ = _today_window()
    month_start, _ = _month_window()

    # --- машины в работе (активные смены + маршрут активного рейса) ---
    active_res = await session.execute(
        select(Shift.id, Shift.vehicle_id, Driver.full_name, Vehicle.license_plate)
        .join(Driver, Driver.id == Shift.driver_id)
        .join(Vehicle, Vehicle.id == Shift.vehicle_id)
        .where(Shift.owner_id == owner.id, Shift.status == "started")
        .order_by(Vehicle.license_plate)
    )
    active_vehicles = []
    active_shift_by_vehicle: dict[int, dict] = {}
    active_trip_vehicle_ids: set[int] = set()
    for shift_id, vehicle_id, dname, plate in active_res.all():
        trip = (
            await session.execute(
                select(Trip).where(
                    Trip.shift_id == shift_id,
                    Trip.status.in_(("created", "in_transit", "unloading")),
                ).limit(1)
            )
        ).scalar_one_or_none()
        active_vehicles.append({
            "plate": plate, "driver": dname,
            "route": (f"{trip.origin or '—'} → {trip.destination or '—'}" if trip else "без рейса"),
            "status": _TRIP_STATUS_LABELS.get(trip.status, "") if trip else "",
        })
        active_shift_by_vehicle[vehicle_id] = {"driver": dname, "plate": plate}
        if trip is not None:
            active_trip_vehicle_ids.add(vehicle_id)

    # --- требуют внимания: истекающие документы (реальные данные) ---
    attention = []
    cutoff = today + timedelta(days=30)
    veh_res = await session.execute(
        select(Vehicle).where(Vehicle.owner_id == owner.id, Vehicle.is_active.is_(True))
    )
    _DOC = {"osago_expires": "ОСАГО", "inspection_expires": "техосмотр", "tacho_expires": "тахограф"}
    for v in veh_res.scalars().all():
        for field, label in _DOC.items():
            exp = getattr(v, field)
            if exp is None:
                continue
            if exp < today:
                attention.append({
                    "sev": "danger", "icon": "📄", "title": f"Истёк {label}",
                    "pill": f"{(today - exp).days} дн", "sub": f"{v.license_plate} · был до {exp:%d.%m.%Y}",
                })
            elif exp <= cutoff:
                attention.append({
                    "sev": "warn", "icon": "📄", "title": "Истекают документы",
                    "pill": f"{(exp - today).days} дн", "sub": f"{label} · {v.license_plate} · до {exp:%d.%m.%Y}",
                })

    # --- GPS-контроль: простои, заведённый двигатель, потеря связи ---
    now_utc = datetime.now(timezone.utc)
    stale_cutoff = now_utc - timedelta(minutes=30)
    telemetry_rows = (
        await session.execute(
            select(VehicleState, Vehicle.license_plate)
            .join(Vehicle, Vehicle.id == VehicleState.vehicle_id)
            .where(Vehicle.owner_id == owner.id, Vehicle.is_active.is_(True))
        )
    ).all()
    map_normal = map_attention = map_problem = 0
    for st, plate in telemetry_rows:
        status = st.motion_status or telemetry_service.vehicle_motion_status(st.speed_kmh, st.ignition)
        duration = telemetry_service.duration_label(st.motion_since_at, now_utc)
        since = fmt_dt(st.motion_since_at, owner.timezone, "%H:%M")
        signal = telemetry_service.vehicle_control_signal(
            motion_status=status,
            has_active_shift=st.vehicle_id in active_shift_by_vehicle,
            has_active_trip=st.vehicle_id in active_trip_vehicle_ids,
            gps_stale=bool(st.last_seen_at and st.last_seen_at < stale_cutoff),
            gps_invalid=st.is_valid is False,
        )
        if signal == telemetry_service.SIGNAL_GPS_STALE:
            map_problem += 1
            attention.append({
                "sev": "danger", "icon": "📡", "title": "Нет свежего GPS",
                "pill": telemetry_service.duration_label(st.last_seen_at, now_utc),
                "sub": f"{plate} · последний сигнал {fmt_dt(st.last_seen_at, owner.timezone, '%H:%M')}",
            })
        elif signal == telemetry_service.SIGNAL_GPS_INVALID:
            map_problem += 1
            attention.append({
                "sev": "danger", "icon": "📍", "title": "GPS без точных координат",
                "pill": "проверить", "sub": f"{plate} · метка держится на последней нормальной точке",
            })
        elif signal == telemetry_service.SIGNAL_MOVING_WITHOUT_SHIFT:
            map_problem += 1
            attention.append({
                "sev": "danger", "icon": "🚛", "title": "Машина едет без смены",
                "pill": f"{Decimal(st.speed_kmh or 0):.0f} км/ч",
                "sub": f"{plate} · с {since} · водитель не начал смену",
            })
        elif signal == telemetry_service.SIGNAL_MOVING_WITHOUT_TRIP:
            map_attention += 1
            driver = active_shift_by_vehicle.get(st.vehicle_id, {}).get("driver", "водитель в смене")
            attention.append({
                "sev": "warn", "icon": "🛣", "title": "Едет без активного рейса",
                "pill": f"{Decimal(st.speed_kmh or 0):.0f} км/ч",
                "sub": f"{plate} · {driver} · с {since}",
            })
        elif signal == telemetry_service.SIGNAL_IDLE_ENGINE:
            map_attention += 1
            attention.append({
                "sev": "warn", "icon": "⛽", "title": "Стоит с заведённым двигателем",
                "pill": duration, "sub": f"{plate} · с {since}",
            })
        elif status == telemetry_service.MOTION_MOVING:
            map_normal += 1
        else:
            map_attention += 1

    # --- структура расходов за месяц (одобренные, по категориям) ---
    br_res = await session.execute(
        select(Expense.category, func.coalesce(func.sum(Expense.amount_rub), 0))
        .where(
            Expense.owner_id == owner.id,
            Expense.status == "approved",
            Expense.created_at >= month_start,
        )
        .group_by(Expense.category)
    )
    breakdown = [
        {"label": _EXPENSE_CAT_LABELS.get(cat, cat), "amount": float(amt)}
        for cat, amt in br_res.all() if amt
    ]
    breakdown.sort(key=lambda x: x["amount"], reverse=True)

    # --- дельта рейсов к вчера ---
    trips_yesterday = (
        await session.execute(
            select(func.count(Trip.id)).where(
                Trip.owner_id == owner.id,
                Trip.status == "completed",
                Trip.completed_at >= today_start - timedelta(days=1),
                Trip.completed_at < today_start,
            )
        )
    ).scalar_one() or 0

    return {
        "date_label": date_label,
        "in_transit_count": len(active_vehicles),
        "active_vehicles": active_vehicles,
        "attention": attention,
        "breakdown": breakdown,
        "expense_total": sum(b["amount"] for b in breakdown),
        "expense_month": _RU_MONTHS_NOM[now_local.month - 1],
        "trips_yesterday": trips_yesterday,
        "map_normal": map_normal,
        "map_attention": map_attention,
        "map_problem": map_problem,
    }


@app.get("/", response_class=HTMLResponse)
async def dashboard(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    today_start, today_end = _today_window()
    month_start, month_end = _month_window()
    year_start = datetime(today_end.year, 1, 1, tzinfo=timezone.utc)

    trips_today = (
        await session.execute(
            select(func.count(Trip.id)).where(
                Trip.owner_id == owner.id,
                Trip.status == "completed",
                Trip.completed_at >= today_start,
            )
        )
    ).scalar_one() or 0

    km_today = (
        await session.execute(
            select(func.coalesce(func.sum(Shift.distance_km), 0)).where(
                Shift.owner_id == owner.id,
                Shift.status == "completed",
                Shift.ended_at >= today_start,
            )
        )
    ).scalar_one() or 0

    # три периода × доход/расход/прибыль.
    # доход = выручка рейсов + ручные доходы
    # расход = одобренные expenses (любых категорий, включая fuel) + ручные расходы
    #         НЕ суммируем trips.fuel_cost_rub отдельно — это вызывало двойной счёт
    finance = {
        "today": await _period_totals(session, owner.id, today_start, today_start.date()),
        "month": await _period_totals(session, owner.id, month_start, month_start.date()),
        "year": await _period_totals(session, owner.id, year_start, year_start.date()),
    }

    # последние 10 рейсов
    last_trips_res = await session.execute(
        select(Trip, Driver.full_name, Vehicle.license_plate)
        .join(Driver, Driver.id == Trip.driver_id)
        .join(Vehicle, Vehicle.id == Trip.vehicle_id)
        .where(Trip.owner_id == owner.id)
        .order_by(desc(Trip.created_at))
        .limit(10)
    )
    last_trips = list(last_trips_res.all())

    # график: по умолчанию 7 дней; период переключается на странице через /api/dashboard-chart
    chart = await _dashboard_chart(session, owner.id, "7d")

    insights = await generate_insights(session, owner.id)
    overview = await _dashboard_overview(session, owner)

    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "owner": owner,
            "kpi": {
                "trips_today": trips_today,
                "km_today": km_today,
                "trips_yesterday": overview["trips_yesterday"],
            },
            "finance": finance,
            "last_trips": last_trips,
            "chart": chart,
            "insights": insights,
            "overview": overview,
            "active_page": "dashboard",
        },
    )


@app.get("/api/dashboard-chart")
async def api_dashboard_chart(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    period: str = "7d",
):
    """JSON для переключателя периода графика на дашборде (7d / 30d / 12m)."""
    return await _dashboard_chart(session, owner.id, period)


# Календарные помощники и шаг графика — в timeutil (там и тестируются).
_RU_MON = RU_MONTHS_SHORT
_month_floor = month_floor
_add_months = add_months


async def _dashboard_chart(session: AsyncSession, owner_id: int, period: str) -> dict:
    """
    Данные для графика дашборда за период 7д / 30д / 12 мес.
    Доход/расход считаем ТЕМ ЖЕ определением, что и финансовая матрица
    (_period_totals): доход = выручка completed-рейсов + ручные доходы;
    расход = одобренные expenses + ручные расходы. fuel_cost_rub отдельно
    НЕ суммируем — иначе двойной счёт. profit = доход − расход.
    """
    period = period if period in ("7d", "30d", "6m", "12m") else "7d"
    today = date.today()

    if period in ("6m", "12m"):
        months = 12 if period == "12m" else 6
        first = _add_months(_month_floor(today), -(months - 1))
        start = first
        keys = [_add_months(first, i) for i in range(months)]
        # «июл 2026», а не «июл 26» — владелец читал «26» как число месяца.
        labels = [f"{_RU_MON[k.month - 1]} {k:%Y}" for k in keys]
        bucket_keys = [(k.year, k.month) for k in keys]

        def key_of(d):
            return (d.year, d.month)
    else:
        days = 7 if period == "7d" else 30
        start = today - timedelta(days=days - 1)
        keys = [start + timedelta(days=i) for i in range(days)]
        labels = [k.strftime("%d.%m") for k in keys]
        bucket_keys = keys

        def key_of(d):
            return d

    revenue = {k: Decimal(0) for k in bucket_keys}
    expense = {k: Decimal(0) for k in bucket_keys}

    def accumulate(rows, target):
        for d, amount in rows:
            if d is None:
                continue
            k = key_of(d)
            if k in target:
                target[k] += Decimal(amount)

    rev_rows = await session.execute(
        select(func.date(Trip.completed_at), func.coalesce(func.sum(Trip.revenue_rub), 0))
        .where(
            Trip.owner_id == owner_id,
            Trip.status == "completed",
            func.date(Trip.completed_at) >= start,
        )
        .group_by(func.date(Trip.completed_at))
    )
    accumulate(rev_rows.all(), revenue)

    inc_rows = await session.execute(
        select(ManualEntry.entry_date, func.coalesce(func.sum(ManualEntry.amount_rub), 0))
        .where(
            ManualEntry.owner_id == owner_id,
            ManualEntry.type == "income",
            ManualEntry.entry_date >= start,
        )
        .group_by(ManualEntry.entry_date)
    )
    accumulate(inc_rows.all(), revenue)

    exp_rows = await session.execute(
        select(func.date(Expense.created_at), func.coalesce(func.sum(Expense.amount_rub), 0))
        .where(
            Expense.owner_id == owner_id,
            Expense.status == "approved",
            func.date(Expense.created_at) >= start,
        )
        .group_by(func.date(Expense.created_at))
    )
    accumulate(exp_rows.all(), expense)

    mexp_rows = await session.execute(
        select(ManualEntry.entry_date, func.coalesce(func.sum(ManualEntry.amount_rub), 0))
        .where(
            ManualEntry.owner_id == owner_id,
            ManualEntry.type == "expense",
            ManualEntry.entry_date >= start,
        )
        .group_by(ManualEntry.entry_date)
    )
    accumulate(mexp_rows.all(), expense)

    rev_list = [float(revenue[k]) for k in bucket_keys]
    exp_list = [float(expense[k]) for k in bucket_keys]
    profit_list = [round(r - e, 2) for r, e in zip(rev_list, exp_list)]
    return {
        "labels": labels,
        "revenue": rev_list,
        "expenses": exp_list,
        "profit": profit_list,
        "period": period,
    }


async def _cashflow_chart(
    session: AsyncSession, owner_id: int, df: date, dt: date
) -> dict:
    """Денежный поток за выбранный период «с… по…» (страница /finances).
    Те же определения дохода/расхода, что и в финансовой матрице."""
    labels, bucket_keys, key_of = cashflow_buckets(df, dt)
    revenue = {k: Decimal(0) for k in bucket_keys}
    expense = {k: Decimal(0) for k in bucket_keys}

    def accumulate(rows, target):
        for d, amount in rows:
            if d is None:
                continue
            k = key_of(d)
            if k in target:
                target[k] += Decimal(amount)

    rev_rows = await session.execute(
        select(func.date(Trip.completed_at), func.coalesce(func.sum(Trip.revenue_rub), 0))
        .where(
            Trip.owner_id == owner_id,
            Trip.status == "completed",
            func.date(Trip.completed_at) >= df,
            func.date(Trip.completed_at) <= dt,
        )
        .group_by(func.date(Trip.completed_at))
    )
    accumulate(rev_rows.all(), revenue)
    inc_rows = await session.execute(
        select(ManualEntry.entry_date, func.coalesce(func.sum(ManualEntry.amount_rub), 0))
        .where(
            ManualEntry.owner_id == owner_id,
            ManualEntry.type == "income",
            ManualEntry.entry_date >= df,
            ManualEntry.entry_date <= dt,
        )
        .group_by(ManualEntry.entry_date)
    )
    accumulate(inc_rows.all(), revenue)
    exp_rows = await session.execute(
        select(func.date(Expense.created_at), func.coalesce(func.sum(Expense.amount_rub), 0))
        .where(
            Expense.owner_id == owner_id,
            Expense.status == "approved",
            func.date(Expense.created_at) >= df,
            func.date(Expense.created_at) <= dt,
        )
        .group_by(func.date(Expense.created_at))
    )
    accumulate(exp_rows.all(), expense)
    mexp_rows = await session.execute(
        select(ManualEntry.entry_date, func.coalesce(func.sum(ManualEntry.amount_rub), 0))
        .where(
            ManualEntry.owner_id == owner_id,
            ManualEntry.type == "expense",
            ManualEntry.entry_date >= df,
            ManualEntry.entry_date <= dt,
        )
        .group_by(ManualEntry.entry_date)
    )
    accumulate(mexp_rows.all(), expense)

    rev_list = [float(revenue[k]) for k in bucket_keys]
    exp_list = [float(expense[k]) for k in bucket_keys]
    return {
        "labels": labels,
        "revenue": rev_list,
        "expenses": exp_list,
        "profit": [round(r - e, 2) for r, e in zip(rev_list, exp_list)],
        "period": "custom",
    }


# =========================================================================
# /trips
# =========================================================================
@app.get("/trips", response_class=HTMLResponse)
async def trips_page(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    driver_id: Annotated[str | None, Query()] = None,
    date_from: Annotated[str | None, Query()] = None,
    date_to: Annotated[str | None, Query()] = None,
    page: Annotated[int, Query()] = 1,
):
    # driver_id приходит строкой: пустое значение «— все —» не должно ронять
    # запрос в 422 (из-за этого «Применить» выглядел неработающим).
    d_id = int(driver_id) if (driver_id or "").strip().isdigit() else None
    conditions = [Trip.owner_id == owner.id]
    if d_id:
        conditions.append(Trip.driver_id == d_id)
    if date_from:
        try:
            conditions.append(Trip.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            conditions.append(
                Trip.created_at < datetime.fromisoformat(date_to) + timedelta(days=1)
            )
        except ValueError:
            pass

    # Пагинация (Блок G5): по 50 на страницу, тянем +1 чтобы понять есть ли «дальше».
    page = max(1, page)
    page_size = 50
    rows_res = await session.execute(
        select(Trip, Driver.full_name, Vehicle.license_plate)
        .join(Driver, Driver.id == Trip.driver_id)
        .join(Vehicle, Vehicle.id == Trip.vehicle_id)
        .where(and_(*conditions))
        .order_by(desc(Trip.created_at))
        .limit(page_size + 1)
        .offset((page - 1) * page_size)
    )
    rows_all = list(rows_res.all())
    has_next = len(rows_all) > page_size
    rows = rows_all[:page_size]

    # Сводка по фильтру (для KPI-карточек на странице).
    agg = await session.execute(
        select(
            func.count(Trip.id),
            func.coalesce(func.sum(Trip.revenue_rub), 0),
            func.coalesce(func.sum(Trip.fuel_cost_rub), 0),
            func.coalesce(func.sum(Trip.other_costs_rub), 0),
            func.coalesce(func.sum(Trip.profit_rub), 0),
        ).where(and_(*conditions))
    )
    t_count, t_rev, t_fuel, t_other, t_profit = agg.one()
    totals = {
        "count": t_count or 0,
        "revenue": Decimal(t_rev or 0),
        "expenses": Decimal(t_fuel or 0) + Decimal(t_other or 0),
        "profit": Decimal(t_profit or 0),
    }

    drivers_res = await session.execute(
        select(Driver).where(Driver.owner_id == owner.id).order_by(Driver.full_name)
    )
    drivers = list(drivers_res.scalars().all())
    vehicles_res = await session.execute(
        select(Vehicle)
        .where(Vehicle.owner_id == owner.id, Vehicle.is_active.is_(True))
        .order_by(Vehicle.license_plate)
    )
    vehicles = list(vehicles_res.scalars().all())

    ctx = {
        "request": request,
        "owner": owner,
        "rows": rows,
        "drivers": drivers,
        "vehicles": vehicles,
        "today": date.today().isoformat(),
        "filter_driver_id": d_id,
        "filter_date_from": date_from or "",
        "filter_date_to": date_to or "",
        "active_page": "trips",
        "page": page,
        "has_next": has_next,
        "totals": totals,
    }
    template = "_trips_table.html" if _is_htmx(request) else "trips.html"
    return templates.TemplateResponse(template, ctx)


# =========================================================================
# /drivers
# =========================================================================
@app.get("/drivers", response_class=HTMLResponse)
async def drivers_page(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    new: Annotated[int | None, Query()] = None,
):
    rows = await _drivers_stats(session, owner.id)
    invite = None
    if new is not None:
        created = next((r["driver"] for r in rows if r["driver"].id == new), None)
        if created is not None and created.invite_token:
            link = None
            try:
                me = await request.app.state.driver_bot.get_me()
                link = f"https://t.me/{me.username}?start={created.invite_token}"
            except Exception:
                link = None
            invite = {"name": created.full_name, "link": link}
    totals = {
        "count": len(rows),
        "in_shift": sum(1 for r in rows if r.get("active_shift")),
        "idle": sum(1 for r in rows if not r.get("active_shift")),
    }
    return templates.TemplateResponse(
        "drivers.html",
        {
            "request": request, "owner": owner, "rows": rows,
            "active_page": "drivers", "invite": invite, "totals": totals,
        },
    )


async def _drivers_stats(session: AsyncSession, owner_id: int) -> list[dict]:
    # Показатели за ВСЁ ВРЕМЯ (не за месяц): в карточке водителя нужны его
    # суммарные смены / рейсы / пробег / выручка, а не только текущий месяц.
    owner = await session.get(Owner, owner_id)
    tz_name = owner.timezone if owner else None
    drivers_res = await session.execute(
        select(Driver)
        .where(Driver.owner_id == owner_id, Driver.is_active.is_(True))
        .order_by(Driver.full_name)
    )
    drivers = list(drivers_res.scalars().all())

    # гос.номера машин владельца — для подписи «обычная машина» в карточке
    plates: dict[int, str] = dict(
        (
            await session.execute(
                select(Vehicle.id, Vehicle.license_plate).where(Vehicle.owner_id == owner_id)
            )
        ).all()
    )

    rows = []
    for d in drivers:
        # Простой/невыход (Блок F): активна ли смена и с какого времени тишина.
        last_shift = (
            await session.execute(
                select(Shift.status, Shift.started_at)
                .where(Shift.driver_id == d.id)
                .order_by(Shift.started_at.desc())
                .limit(1)
            )
        ).first()
        active_shift = bool(last_shift and last_shift[0] == "started")
        idle_label = None
        if not active_shift:
            ref = (last_shift[1] if last_shift else None) or d.created_at
            if ref is not None:
                idle_label = ref.astimezone(owner_tz(tz_name)).strftime("%d.%m %H:%M")
        agg = await session.execute(
            select(
                func.coalesce(func.sum(Shift.distance_km), 0),
                func.count(Shift.id),
            ).where(
                Shift.driver_id == d.id,
                Shift.status == "completed",
            )
        )
        km, shifts_count = agg.one()
        trips_agg = await session.execute(
            select(
                func.count(Trip.id),
                func.coalesce(func.sum(Trip.revenue_rub), 0),
                func.coalesce(func.sum(Trip.fuel_cost_rub), 0),
            ).where(
                Trip.driver_id == d.id,
                Trip.status == "completed",
            )
        )
        trips_count, revenue, fuel_cost = trips_agg.one()
        rows.append({
            "driver": d,
            "km": km or 0,
            "shifts": shifts_count or 0,
            "trips": trips_count or 0,
            "revenue": Decimal(revenue or 0),
            "fuel_cost": Decimal(fuel_cost or 0),
            "active_shift": active_shift,
            "idle_label": idle_label,
            "default_vehicle_plate": plates.get(d.default_vehicle_id),
        })
    return rows


_SHIFT_TIME_RE_WEB = re.compile(r"^([01]\d|2[0-3]):([0-5]\d)$")


@app.post("/drivers")
async def create_driver(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    full_name: Annotated[str, Form()],
    salary_type: Annotated[str, Form()] = "per_km",
    salary_rate: Annotated[str, Form()] = "0",
    phone: Annotated[str, Form()] = "",
    per_diem_rub: Annotated[str, Form()] = "0",
    shift_start_time: Annotated[str, Form()] = "",
):
    """Создать водителя из веб-кабинета. Генерируем invite-токен — ссылку
    для подключения показываем владельцу на странице после редиректа."""
    if salary_type not in ("per_km", "per_trip", "percent", "fixed_per_shift", "fixed_per_month"):
        raise HTTPException(status_code=400, detail="Bad salary_type")
    if len(full_name.strip()) < 2:
        raise HTTPException(status_code=400, detail="Bad name")
    try:
        rate = Decimal((salary_rate or "0").replace(",", "."))
        per_diem = Decimal((per_diem_rub or "0").replace(",", "."))
        if rate < 0 or per_diem < 0:
            raise InvalidOperation
    except InvalidOperation:
        raise HTTPException(status_code=400, detail="Bad numeric values")

    sst = shift_start_time.strip()
    if sst and not _SHIFT_TIME_RE_WEB.match(sst):
        raise HTTPException(status_code=400, detail="Bad shift_start_time")

    driver = Driver(
        owner_id=owner.id,
        full_name=full_name.strip(),
        phone=phone.strip() or None,
        salary_type=salary_type,
        salary_rate=rate,
        per_diem_rub=per_diem,
        shift_start_time=sst or None,
        invite_token=uuid.uuid4().hex,
        is_active=True,
    )
    session.add(driver)
    await session.commit()
    return RedirectResponse(f"/drivers?new={driver.id}", status_code=303)


@app.post("/drivers/{driver_id}", response_class=HTMLResponse)
async def update_driver(
    request: Request,
    driver_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    full_name: Annotated[str, Form()],
    salary_type: Annotated[str, Form()],
    salary_rate: Annotated[str, Form()],
    phone: Annotated[str, Form()] = "",
    per_diem_rub: Annotated[str, Form()] = "0",
    shift_start_time: Annotated[str, Form()] = "",
    default_vehicle_id: Annotated[str, Form()] = "",
):
    driver = await session.get(Driver, driver_id)
    if driver is None or driver.owner_id != owner.id:
        raise HTTPException(status_code=404, detail="Driver not found")
    if salary_type not in ("per_km", "per_trip", "percent", "fixed_per_shift", "fixed_per_month"):
        raise HTTPException(status_code=400, detail="Bad salary_type")
    if len(full_name.strip()) < 2:
        raise HTTPException(status_code=400, detail="Bad name")
    try:
        rate = Decimal(salary_rate.replace(",", "."))
        per_diem = Decimal(per_diem_rub.replace(",", "."))
        if rate < 0 or per_diem < 0:
            raise InvalidOperation
    except InvalidOperation:
        raise HTTPException(status_code=400, detail="Bad numeric values")

    sst = shift_start_time.strip()
    if sst and not _SHIFT_TIME_RE_WEB.match(sst):
        raise HTTPException(status_code=400, detail="Bad shift_start_time")

    # «обычная машина» (анти-миссклик): пустое значение = не закреплена
    dv_raw = default_vehicle_id.strip()
    if not dv_raw:
        driver.default_vehicle_id = None
    else:
        try:
            dv_id = int(dv_raw)
        except ValueError:
            raise HTTPException(status_code=400, detail="Bad default_vehicle_id")
        dv = await session.get(Vehicle, dv_id)
        if dv is None or dv.owner_id != owner.id:
            raise HTTPException(status_code=400, detail="Bad default_vehicle_id")
        driver.default_vehicle_id = dv_id

    driver.full_name = full_name.strip()
    driver.phone = phone.strip() or None
    driver.salary_type = salary_type
    driver.salary_rate = rate
    driver.per_diem_rub = per_diem
    driver.shift_start_time = sst or None
    await session.commit()

    rows = await _drivers_stats(session, owner.id)
    row = next((r for r in rows if r["driver"].id == driver.id), None)
    return templates.TemplateResponse(
        "_driver_row.html", {"request": request, "row": row, "edit": False}
    )


@app.delete("/drivers/{driver_id}", response_class=HTMLResponse)
async def delete_driver(
    request: Request,
    driver_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """
    Мягкое удаление: is_active=False, telegram_id обнуляем (чтобы тот же
    человек мог быть привязан заново). История смен/рейсов сохраняется по FK.
    Водителю отправляем сообщение что владелец отключил его.
    """
    driver = await session.get(Driver, driver_id)
    if driver is None or driver.owner_id != owner.id:
        raise HTTPException(status_code=404)
    tg_id = driver.telegram_id
    driver.is_active = False
    driver.telegram_id = None  # освобождаем привязку
    driver.invite_token = None
    await session.commit()

    # уведомить водителя что его отключили
    if tg_id is not None:
        from app.bots.notifications import notify_driver
        driver_bot = request.app.state.driver_bot
        await notify_driver(
            driver_bot, session, tg_id,
            "ℹ️ Владелец отключил вас от учёта автопарка. "
            "Если это ошибка — попросите его прислать новую ссылку-приглашение."
        )
    return HTMLResponse("")


@app.get("/drivers/{driver_id}/edit", response_class=HTMLResponse)
async def driver_edit_form(
    request: Request,
    driver_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    driver = await session.get(Driver, driver_id)
    if driver is None or driver.owner_id != owner.id:
        raise HTTPException(status_code=404)
    rows = await _drivers_stats(session, owner.id)
    row = next((r for r in rows if r["driver"].id == driver.id), None)
    vehicles = list(
        (
            await session.execute(
                select(Vehicle)
                .where(Vehicle.owner_id == owner.id, Vehicle.is_active.is_(True))
                .order_by(Vehicle.license_plate)
            )
        ).scalars().all()
    )
    return templates.TemplateResponse(
        "_driver_row.html",
        {"request": request, "row": row, "edit": True, "vehicles": vehicles},
    )


@app.get("/drivers/{driver_id}/cancel", response_class=HTMLResponse)
async def driver_cancel_edit(
    request: Request,
    driver_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """HTMX-эндпоинт: вернуть строку обратно в view-режим без сохранения."""
    driver = await session.get(Driver, driver_id)
    if driver is None or driver.owner_id != owner.id:
        raise HTTPException(status_code=404)
    rows = await _drivers_stats(session, owner.id)
    row = next((r for r in rows if r["driver"].id == driver.id), None)
    return templates.TemplateResponse(
        "_driver_row.html", {"request": request, "row": row, "edit": False}
    )


# =========================================================================
# /vehicles
# =========================================================================
async def _vehicle_row_dict(session: AsyncSession, vehicle: Vehicle, month_start) -> dict:
    """Считаем месячные показатели по одной машине: пробег, рейсы, выручка, расходы, прибыль."""
    km = (
        await session.execute(
            select(func.coalesce(func.sum(Shift.distance_km), 0)).where(
                Shift.vehicle_id == vehicle.id,
                Shift.status == "completed",
                Shift.ended_at >= month_start,
            )
        )
    ).scalar_one() or 0
    trips_agg = await session.execute(
        select(
            func.count(Trip.id),
            func.coalesce(func.sum(Trip.revenue_rub), 0),
            func.coalesce(func.sum(Trip.fuel_cost_rub), 0),
        ).where(
            Trip.vehicle_id == vehicle.id,
            Trip.status == "completed",
            Trip.completed_at >= month_start,
        )
    )
    trips_count, revenue, fuel = trips_agg.one()
    # одобренные расходы водителей по сменам этой машины за период
    approved = (
        await session.execute(
            select(func.coalesce(func.sum(Expense.amount_rub), 0))
            .select_from(Expense)
            .join(Shift, Shift.id == Expense.shift_id)
            .where(
                Shift.vehicle_id == vehicle.id,
                Expense.status == "approved",
                Expense.created_at >= month_start,
            )
        )
    ).scalar_one() or Decimal(0)
    revenue = Decimal(revenue or 0)
    fuel = Decimal(fuel or 0)
    approved = Decimal(approved)
    # топливо уже в одобренных expenses (если водитель его одобрил), но
    # часть может быть pending — суммируем только approved + не дублируем fuel
    total_expense = approved
    profit = revenue - total_expense
    margin = (profit / revenue * Decimal(100)) if revenue > 0 else Decimal(0)
    # машина сейчас в работе? (есть открытая смена) — для бейджа на карточке
    active = (
        await session.execute(
            select(func.count(Shift.id)).where(
                Shift.vehicle_id == vehicle.id, Shift.status == "started"
            )
        )
    ).scalar_one() or 0
    return {
        "vehicle": vehicle,
        "km": km, "fuel": fuel,
        "trips": trips_count or 0,
        "revenue": revenue,
        "expense": total_expense,
        "profit": profit,
        "margin": margin,
        "active": bool(active),
    }


def _clean_stavtrack_object_id(value: str) -> str | None:
    cleaned = value.strip()
    return cleaned or None


@app.get("/vehicles", response_class=HTMLResponse)
async def vehicles_page(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    limit: Annotated[int | None, Query()] = None,
    dup: Annotated[str | None, Query()] = None,
    gps_dup: Annotated[str | None, Query()] = None,
):
    month_start, _ = _month_window()
    vehicles_res = await session.execute(
        select(Vehicle)
        .where(Vehicle.owner_id == owner.id, Vehicle.is_active.is_(True))
        .order_by(Vehicle.license_plate)
    )
    vehicles = list(vehicles_res.scalars().all())
    rows = [await _vehicle_row_dict(session, v, month_start) for v in vehicles]
    in_minus = sum(1 for r in rows if r["profit"] < 0)
    in_work = (
        await session.execute(
            select(func.count(Shift.id)).where(
                Shift.owner_id == owner.id, Shift.status == "started"
            )
        )
    ).scalar_one() or 0
    totals = {
        "count": len(rows),
        "in_work": in_work,
        "revenue": sum((r["revenue"] for r in rows), Decimal(0)),
        "profit": sum((r["profit"] for r in rows), Decimal(0)),
    }
    notice = None
    if limit is not None:
        notice = {"kind": "limit", "limit": limit}
    elif dup:
        notice = {"kind": "dup", "plate": dup}
    elif gps_dup:
        notice = {"kind": "gps_dup", "stavtrack_object_id": gps_dup}
    return templates.TemplateResponse(
        "vehicles.html",
        {
            "request": request, "owner": owner, "rows": rows,
            "active_page": "vehicles", "in_minus": in_minus, "notice": notice,
            "totals": totals,
        },
    )


@app.get("/vehicles/{vehicle_id}/stats", response_class=HTMLResponse)
async def vehicle_stats(
    request: Request,
    vehicle_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Полная статистика по одной машине — только для владельца: пробег,
    рейсы, выручка, расходы, прибыль, маржа за месяц и за всё время."""
    vehicle = await session.get(Vehicle, vehicle_id)
    if vehicle is None or vehicle.owner_id != owner.id:
        raise HTTPException(status_code=404)
    month_start, _ = _month_window()
    epoch = datetime(2000, 1, 1, tzinfo=timezone.utc)
    month = await _vehicle_row_dict(session, vehicle, month_start)
    alltime = await _vehicle_row_dict(session, vehicle, epoch)
    return templates.TemplateResponse(
        "vehicle_stats.html",
        {
            "request": request, "owner": owner, "vehicle": vehicle,
            "month": month, "alltime": alltime, "active_page": "vehicles",
        },
    )


@app.post("/vehicles")
async def create_vehicle(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    license_plate: Annotated[str, Form()],
    brand: Annotated[str, Form()] = "",
    type: Annotated[str, Form()] = "truck",
    fuel_norm_per_100km: Annotated[str, Form()] = "",
    stavtrack_object_id: Annotated[str, Form()] = "",
    osago_expires: Annotated[str, Form()] = "",
    inspection_expires: Annotated[str, Form()] = "",
    tacho_expires: Annotated[str, Form()] = "",
):
    """Добавить машину из веб-кабинета. Проверяем лимит тарифа и уникальность
    гос. номера; мягко удалённую машину с тем же номером — реактивируем."""
    if type not in ("truck", "gazelle", "refrigerator"):
        raise HTTPException(status_code=400, detail="Bad type")
    plate_clean = license_plate.strip().upper().replace(" ", "")
    if len(plate_clean) < 4:
        raise HTTPException(status_code=400, detail="Bad license_plate")

    norm: Decimal | None = None
    if fuel_norm_per_100km.strip():
        try:
            norm = Decimal(fuel_norm_per_100km.replace(",", "."))
            if norm < 0:
                raise InvalidOperation
        except InvalidOperation:
            raise HTTPException(status_code=400, detail="Bad fuel_norm")

    def _parse_date(value: str) -> date | None:
        if not value.strip():
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    can_add, _count, plan_limit = await billing.can_add_vehicle(session, owner.id)
    if not can_add:
        return RedirectResponse(f"/vehicles?limit={plan_limit}", status_code=303)

    existing = (
        await session.execute(
            select(Vehicle).where(
                Vehicle.owner_id == owner.id,
                Vehicle.license_plate == plate_clean,
            )
        )
    ).scalar_one_or_none()
    if existing is not None and existing.is_active:
        return RedirectResponse(f"/vehicles?dup={plate_clean}", status_code=303)

    stavtrack_id = _clean_stavtrack_object_id(stavtrack_object_id)
    if stavtrack_id:
        gps_existing = (
            await session.execute(
                select(Vehicle).where(
                    Vehicle.owner_id == owner.id,
                    Vehicle.stavtrack_object_id == stavtrack_id,
                    Vehicle.is_active.is_(True),
                )
            )
        ).scalar_one_or_none()
        if gps_existing is not None and (existing is None or gps_existing.id != existing.id):
            return RedirectResponse(f"/vehicles?gps_dup={stavtrack_id}", status_code=303)

    target = existing if existing is not None else Vehicle(owner_id=owner.id)
    target.license_plate = plate_clean
    target.brand = brand.strip() or None
    target.type = type
    target.stavtrack_object_id = stavtrack_id
    target.fuel_norm_per_100km = norm
    target.osago_expires = _parse_date(osago_expires)
    target.inspection_expires = _parse_date(inspection_expires)
    target.tacho_expires = _parse_date(tacho_expires)
    target.is_active = True
    if existing is None:
        session.add(target)
    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        return RedirectResponse(f"/vehicles?dup={plate_clean}", status_code=303)
    return RedirectResponse("/vehicles", status_code=303)


async def _load_vehicle_row(
    session: AsyncSession, owner: Owner, vehicle_id: int
) -> tuple[Vehicle, dict]:
    vehicle = await session.get(Vehicle, vehicle_id)
    if vehicle is None or vehicle.owner_id != owner.id:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    month_start, _ = _month_window()
    row = await _vehicle_row_dict(session, vehicle, month_start)
    return vehicle, row


@app.get("/vehicles/{vehicle_id}/edit", response_class=HTMLResponse)
async def vehicle_edit_form(
    request: Request,
    vehicle_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    _, row = await _load_vehicle_row(session, owner, vehicle_id)
    return templates.TemplateResponse(
        "_vehicle_row.html", {"request": request, "row": row, "edit": True}
    )


@app.get("/vehicles/{vehicle_id}/cancel", response_class=HTMLResponse)
async def vehicle_cancel_edit(
    request: Request,
    vehicle_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    _, row = await _load_vehicle_row(session, owner, vehicle_id)
    return templates.TemplateResponse(
        "_vehicle_row.html", {"request": request, "row": row, "edit": False}
    )


@app.post("/vehicles/{vehicle_id}", response_class=HTMLResponse)
async def vehicle_update(
    request: Request,
    vehicle_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    license_plate: Annotated[str, Form()],
    brand: Annotated[str, Form()] = "",
    type: Annotated[str, Form()] = "truck",
    fuel_norm_per_100km: Annotated[str, Form()] = "",
    stavtrack_object_id: Annotated[str, Form()] = "",
    osago_expires: Annotated[str, Form()] = "",
    inspection_expires: Annotated[str, Form()] = "",
    tacho_expires: Annotated[str, Form()] = "",
):
    vehicle = await session.get(Vehicle, vehicle_id)
    if vehicle is None or vehicle.owner_id != owner.id:
        raise HTTPException(status_code=404)
    if type not in ("truck", "gazelle", "refrigerator"):
        raise HTTPException(status_code=400, detail="Bad type")

    plate_clean = license_plate.strip().upper().replace(" ", "")
    if len(plate_clean) < 4:
        raise HTTPException(status_code=400, detail="Bad license_plate")
    vehicle.license_plate = plate_clean
    vehicle.brand = brand.strip() or None
    vehicle.type = type
    stavtrack_id = _clean_stavtrack_object_id(stavtrack_object_id)
    if stavtrack_id:
        gps_existing = (
            await session.execute(
                select(Vehicle).where(
                    Vehicle.owner_id == owner.id,
                    Vehicle.stavtrack_object_id == stavtrack_id,
                    Vehicle.id != vehicle.id,
                    Vehicle.is_active.is_(True),
                )
            )
        ).scalar_one_or_none()
        if gps_existing is not None:
            raise HTTPException(status_code=400, detail="Stavtrack ID уже занят")
    vehicle.stavtrack_object_id = stavtrack_id
    if fuel_norm_per_100km.strip():
        try:
            vehicle.fuel_norm_per_100km = Decimal(fuel_norm_per_100km.replace(",", "."))
        except InvalidOperation:
            raise HTTPException(status_code=400, detail="Bad fuel_norm")
    else:
        vehicle.fuel_norm_per_100km = None

    def _parse_date(value: str) -> date | None:
        if not value.strip():
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    vehicle.osago_expires = _parse_date(osago_expires)
    vehicle.inspection_expires = _parse_date(inspection_expires)
    vehicle.tacho_expires = _parse_date(tacho_expires)

    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=400, detail="Plate уже занят")

    _, row = await _load_vehicle_row(session, owner, vehicle_id)
    return templates.TemplateResponse(
        "_vehicle_row.html", {"request": request, "row": row, "edit": False}
    )


@app.delete("/vehicles/{vehicle_id}", response_class=HTMLResponse)
async def vehicle_delete(
    vehicle_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Мягкое удаление — ставим is_active=False, чтобы история смен/рейсов
    осталась с FK ссылками целой."""
    vehicle = await session.get(Vehicle, vehicle_id)
    if vehicle is None or vehicle.owner_id != owner.id:
        raise HTTPException(status_code=404)
    vehicle.is_active = False
    await session.commit()
    return HTMLResponse("")  # hx-swap='outerHTML' уберёт строку


# =========================================================================
# /finances
# =========================================================================
@app.get("/finances", response_class=HTMLResponse)
async def finances_page(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    period_from: Annotated[str | None, Query()] = None,
    period_to: Annotated[str | None, Query()] = None,
):
    df, dt = _parse_period(period_from, period_to)
    summary = await _finance_summary(session, owner.id, df, dt)

    # Денежный поток за выбранный период: шаг (день/неделя/месяц) сам
    # подстраивается под длину диапазона «с… по…».
    cashflow = await _cashflow_chart(session, owner.id, df, dt)

    # Прибыльность направлений: прибыль завершённых рейсов по маршруту за период.
    dir_res = await session.execute(
        select(
            Trip.origin,
            Trip.destination,
            func.count(Trip.id),
            func.coalesce(func.sum(Trip.revenue_rub), 0),
            func.coalesce(func.sum(Trip.profit_rub), 0),
        )
        .where(
            Trip.owner_id == owner.id,
            Trip.status == "completed",
            func.date(Trip.completed_at) >= df,
            func.date(Trip.completed_at) <= dt,
        )
        .group_by(Trip.origin, Trip.destination)
        .order_by(func.coalesce(func.sum(Trip.profit_rub), 0).desc())
        .limit(8)
    )
    dir_rows = dir_res.all()
    max_abs = max((abs(Decimal(r[4] or 0)) for r in dir_rows), default=Decimal(0)) or Decimal(1)
    directions = [
        {
            "route": f"{o or '—'} → {d or '—'}",
            "trips": cnt,
            "revenue": Decimal(rev or 0),
            "profit": Decimal(pr or 0),
            "bar": int(abs(Decimal(pr or 0)) / max_abs * 100),
        }
        for o, d, cnt, rev, pr in dir_rows
    ]
    inc = summary["total_income"]
    margin = float(summary["profit"] / inc * 100) if inc > 0 else 0.0

    entries_res = await session.execute(
        select(ManualEntry)
        .where(
            ManualEntry.owner_id == owner.id,
            ManualEntry.entry_date >= df,
            ManualEntry.entry_date <= dt,
        )
        .order_by(desc(ManualEntry.entry_date), desc(ManualEntry.id))
    )
    entries = list(entries_res.scalars().all())

    return templates.TemplateResponse(
        "finances.html",
        {
            "request": request,
            "owner": owner,
            "entries": entries,
            "summary": summary,
            "cashflow": cashflow,
            "directions": directions,
            "margin": margin,
            "period_from": df.isoformat(),
            "period_to": dt.isoformat(),
            "today": date.today().isoformat(),
            "active_page": "finances",
        },
    )


@app.post("/finances/add", response_class=HTMLResponse)
async def finances_add(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    type: Annotated[str, Form()],
    amount_rub: Annotated[str, Form()],
    entry_date: Annotated[str, Form()],
    category: Annotated[str, Form()] = "",
    description: Annotated[str, Form()] = "",
):
    if type not in ("income", "expense"):
        raise HTTPException(status_code=400, detail="Bad type")
    try:
        amount = Decimal(amount_rub.replace(",", "."))
        if amount <= 0:
            raise InvalidOperation
        edate = date.fromisoformat(entry_date)
    except (InvalidOperation, ValueError):
        raise HTTPException(status_code=400, detail="Bad input")

    entry = ManualEntry(
        owner_id=owner.id,
        type=type,
        category=category.strip() or None,
        amount_rub=amount,
        description=description.strip() or None,
        entry_date=edate,
    )
    session.add(entry)
    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=400, detail="DB error")

    return templates.TemplateResponse(
        "_manual_entry_row.html", {"request": request, "entry": entry}
    )


@app.post("/finances/delete/{entry_id}")
async def finances_delete(
    entry_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    entry = await session.get(ManualEntry, entry_id)
    if entry is None or entry.owner_id != owner.id:
        raise HTTPException(status_code=404)
    await session.delete(entry)
    await session.commit()
    return Response(status_code=200)


@app.get("/finances/export.xlsx")
async def finances_export(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    period_from: Annotated[str | None, Query()] = None,
    period_to: Annotated[str | None, Query()] = None,
):
    df, dt = _parse_period(period_from, period_to)
    summary = await _finance_summary(session, owner.id, df, dt)

    wb = _build_finance_workbook(summary, df, dt)
    await _fill_entries_sheet(wb, session, owner.id, df, dt)
    await _fill_trips_sheet(wb, session, owner.id, df, dt)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"finances_{df.isoformat()}_{dt.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =========================================================================
# /acts — акты оказанных услуг по РЦ за период (форма 101 РС, .xlsx)
# =========================================================================
def _acts_range(df: date, dt: date) -> tuple[datetime, datetime]:
    start = datetime(df.year, df.month, df.day, tzinfo=timezone.utc)
    end = datetime(dt.year, dt.month, dt.day, tzinfo=timezone.utc) + timedelta(days=1)
    return start, end


def _executor_from_owner(owner: Owner) -> dict:
    """Реквизиты Исполнителя для шапки акта (берём с owner)."""
    return {
        "full_name": owner.executor_name or owner.company_name or "",
        "inn": owner.inn or "",
        "ogrnip": owner.ogrnip or "",
        "address": owner.legal_address or "",
        "bank_name": owner.bank_name or "",
        "account": owner.bank_account or "",
        "corr_account": owner.corr_account or "",
        "bik": owner.bik or "",
        "signer_name": owner.signer_name or owner.full_name or "",
    }


def _customer_to_dict(c: Customer | None) -> dict:
    """Реквизиты Заказчика для шапки акта."""
    if c is None:
        return {
            "name": "", "inn": "", "kpp": "", "address": "", "bank_name": "",
            "account": "", "corr_account": "", "bik": "", "contract_number": "",
            "contract_date": None, "signer_name": "",
        }
    return {
        "name": c.name or "",
        "inn": c.inn or "",
        "kpp": c.kpp or "",
        "address": c.legal_address or "",
        "bank_name": c.bank_name or "",
        "account": c.bank_account or "",
        "corr_account": c.corr_account or "",
        "bik": c.bik or "",
        "contract_number": c.contract_number or "",
        "contract_date": c.contract_date,
        "signer_name": c.signer_name or "",
    }


async def _active_distribution_centers(
    session: AsyncSession, owner_id: int
) -> list[DistributionCenter]:
    res = await session.execute(
        select(DistributionCenter)
        .where(DistributionCenter.owner_id == owner_id, DistributionCenter.is_active.is_(True))
        .order_by(DistributionCenter.name)
    )
    return list(res.scalars().all())


@app.get("/acts", response_class=HTMLResponse)
async def acts_page(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    period_from: Annotated[str | None, Query()] = None,
    period_to: Annotated[str | None, Query()] = None,
    customer_id: Annotated[str | None, Query()] = None,
    title: Annotated[str | None, Query()] = None,
    act_number: Annotated[str | None, Query()] = None,
    act_date: Annotated[str | None, Query()] = None,
):
    df, dt = _parse_period(period_from, period_to)
    start, end = _acts_range(df, dt)
    # Отдельные рейсы за период — для чек-листа «какие включить в акт».
    rows_res = await session.execute(
        select(
            Trip.id, Trip.completed_at, Trip.origin, Trip.destination,
            Driver.full_name, Vehicle.license_plate, Trip.revenue_rub,
        )
        .join(Driver, Driver.id == Trip.driver_id)
        .join(Vehicle, Vehicle.id == Trip.vehicle_id)
        .where(
            Trip.owner_id == owner.id,
            Trip.status == "completed",
            Trip.completed_at >= start,
            Trip.completed_at < end,
            Trip.revenue_rub.is_not(None),
        )
        .order_by(Trip.completed_at, Trip.id)
    )
    rc_lookup = rc_service.distribution_center_lookup(await _active_distribution_centers(session, owner.id))
    trips = [
        {
            "id": tid,
            "date": cat,
            "origin": orig,
            "destination": dest,
            "destination_address": rc_service.canonical_rc_address(dest, rc_lookup),
            "driver": drv,
            "plate": plate,
            "revenue": Decimal(rev or 0),
        }
        for tid, cat, orig, dest, drv, plate, rev in rows_res.all()
    ]
    total_amount = sum((t["revenue"] for t in trips), Decimal(0))
    total_trips = len(trips)

    customers_res = await session.execute(
        select(Customer)
        .where(Customer.owner_id == owner.id, Customer.is_active.is_(True))
        .order_by(Customer.name)
    )
    customers = list(customers_res.scalars().all())
    # реквизиты Исполнителя считаем заполненными, если есть ИНН и наименование
    requisites_ready = bool(owner.inn and (owner.executor_name or owner.company_name))

    return templates.TemplateResponse(
        "acts.html",
        {
            "request": request, "owner": owner, "trips": trips,
            "period_from": df.isoformat(), "period_to": dt.isoformat(),
            "customers": customers,
            "total_amount": total_amount, "total_trips": total_trips,
            "act_date": act_date or date.today().isoformat(),
            "act_title": title or "Акт выполненных работ",
            "act_number_val": act_number or "",
            "sel_customer_id": customer_id or "",
            "requisites_ready": requisites_ready,
            "active_page": "finances",
        },
    )


@app.get("/acts/export.xlsx")
async def acts_export(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    period_from: Annotated[str | None, Query()] = None,
    period_to: Annotated[str | None, Query()] = None,
    customer_id: Annotated[str | None, Query()] = None,
    act_number: Annotated[str, Query()] = "",
    act_date: Annotated[str | None, Query()] = None,
    title: Annotated[str, Query()] = "Акт выполненных работ",
    trip_ids: Annotated[list[int], Query()] = [],
    selection_mode: Annotated[str, Query()] = "",
):
    """Акт (форма 101 РС): один лист с выбранными рейсами за период,
    с реквизитами Исполнителя/Заказчика, итогом и суммой прописью.
    title — название в шапке; trip_ids — какие рейсы включить (пусто = все)."""
    df, dt = _parse_period(period_from, period_to)
    start, end = _acts_range(df, dt)

    # Заказчик: явно выбранный или первый активный у владельца.
    # customer_id приходит строкой (пустая, если заказчиков нет) — парсим мягко.
    customer: Customer | None = None
    cid = int(customer_id) if (customer_id or "").strip().isdigit() else None
    if cid is not None:
        cand = await session.get(Customer, cid)
        if cand is not None and cand.owner_id == owner.id:
            customer = cand
    if customer is None:
        res = await session.execute(
            select(Customer)
            .where(Customer.owner_id == owner.id, Customer.is_active.is_(True))
            .order_by(Customer.id)
            .limit(1)
        )
        customer = res.scalar_one_or_none()

    trip_q = (
        select(Trip, Driver.full_name, Vehicle.license_plate)
        .join(Driver, Driver.id == Trip.driver_id)
        .join(Vehicle, Vehicle.id == Trip.vehicle_id)
        .where(
            Trip.owner_id == owner.id,
            Trip.status == "completed",
            Trip.completed_at >= start,
            Trip.completed_at < end,
            Trip.revenue_rub.is_not(None),
        )
    )
    # Пустой trip_ids = все рейсы периода для прямой ссылки.
    # Если запрос пришёл из чек-листа, пустой выбор должен остаться пустым.
    if trip_ids:
        trip_q = trip_q.where(Trip.id.in_(trip_ids))
    elif selection_mode == "checklist":
        trip_q = trip_q.where(Trip.id.in_([]))
    rows_res = await session.execute(trip_q.order_by(Trip.completed_at, Trip.id))
    rc_lookup = rc_service.distribution_center_lookup(await _active_distribution_centers(session, owner.id))
    rows = [
        {
            "date": trip.completed_at,
            "origin": trip.origin,
            "destination": trip.destination,
            "destination_address": rc_service.canonical_rc_address(trip.destination, rc_lookup),
            "plate": plate,
            "driver": driver_name,
            "amount": trip.revenue_rub,
        }
        for trip, driver_name, plate in rows_res.all()
    ]

    try:
        adate = date.fromisoformat(act_date) if act_date else date.today()
    except ValueError:
        adate = date.today()
    number = (act_number or "").strip() or "б/н"

    wb = act_service.build_act_101rs(
        title=(title or "Акт").strip() or "Акт",
        act_number=number,
        act_date=adate,
        period_from=df,
        period_to=dt,
        executor=_executor_from_owner(owner),
        customer=_customer_to_dict(customer),
        rows=rows,
    )
    buf = act_service.workbook_bytes(wb)
    # Имя файла может содержать кириллицу (напр. «101РС») — заголовок ставим по
    # RFC 5987: ascii-фолбэк + filename* в UTF-8.
    utf8_name = quote(f"Акт_{number}_{adate.isoformat()}.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="akt_{adate.isoformat()}.xlsx"; '
                f"filename*=UTF-8''{utf8_name}"
            )
        },
    )


# =====================================================================
# РЕКВИЗИТЫ: Исполнитель (owner) + заказчики (customers) — шапка акта 101 РС
# =====================================================================
def _norm(v: str | None) -> str | None:
    v = (v or "").strip()
    return v or None


def _apply_customer_form(c: Customer, form: dict) -> None:
    c.name = (form.get("name") or "").strip()
    c.inn = _norm(form.get("inn"))
    c.kpp = _norm(form.get("kpp"))
    c.legal_address = _norm(form.get("legal_address"))
    c.bank_name = _norm(form.get("bank_name"))
    c.bank_account = _norm(form.get("bank_account"))
    c.corr_account = _norm(form.get("corr_account"))
    c.bik = _norm(form.get("bik"))
    c.contract_number = _norm(form.get("contract_number"))
    c.signer_name = _norm(form.get("signer_name"))
    cd = (form.get("contract_date") or "").strip()
    try:
        c.contract_date = date.fromisoformat(cd) if cd else None
    except ValueError:
        c.contract_date = None


@app.get("/requisites", response_class=HTMLResponse)
async def requisites_page(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    res = await session.execute(
        select(Customer).where(Customer.owner_id == owner.id).order_by(Customer.name)
    )
    customers = list(res.scalars().all())
    admins_res = await session.execute(
        select(Admin).where(Admin.owner_id == owner.id).order_by(Admin.created_at)
    )
    admins = list(admins_res.scalars().all())

    # Устройства (активные веб-сессии кабинета): кто, с чего и когда заходил.
    current_ws = await _session_from_request(request, session)
    viewer_tid = await _viewer_telegram_id(request, session)
    is_owner_viewer = viewer_tid is None or viewer_tid == owner.telegram_id
    admin_names = {a.telegram_id: (a.name or f"Админ {a.telegram_id}") for a in admins}
    ws_res = await session.execute(
        select(WebSession)
        .where(WebSession.owner_id == owner.id, WebSession.revoked_at.is_(None))
        .order_by(desc(WebSession.last_seen_at), desc(WebSession.id))
    )
    sessions = [
        {
            "id": ws.id,
            "who": ("Владелец" if ws.telegram_id == owner.telegram_id
                    else admin_names.get(ws.telegram_id, f"Админ {ws.telegram_id}")),
            "device": ws.device_label or "—",
            "ip": ws.ip or "—",
            "created": fmt_dt(ws.created_at, owner.timezone, "%d.%m %H:%M"),
            "seen": fmt_dt(ws.last_seen_at, owner.timezone, "%d.%m %H:%M"),
            "is_current": current_ws is not None and ws.id == current_ws.id,
            # Админ — «второй телефон» владельца с полным доступом, поэтому
            # завершать устройства кабинета может любой вошедший (владелец или админ).
            "can_revoke": True,
        }
        for ws in ws_res.scalars().all()
    ]
    return templates.TemplateResponse(
        "requisites.html",
        {
            "request": request, "owner": owner, "customers": customers,
            "admins": admins, "sessions": sessions,
            "is_owner_viewer": is_owner_viewer, "active_page": "requisites",
        },
    )


@app.post("/requisites/executor")
async def requisites_save_executor(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    executor_name: Annotated[str, Form()] = "",
    inn: Annotated[str, Form()] = "",
    ogrnip: Annotated[str, Form()] = "",
    legal_address: Annotated[str, Form()] = "",
    bank_name: Annotated[str, Form()] = "",
    bank_account: Annotated[str, Form()] = "",
    corr_account: Annotated[str, Form()] = "",
    bik: Annotated[str, Form()] = "",
    signer_name: Annotated[str, Form()] = "",
):
    owner.executor_name = _norm(executor_name)
    owner.inn = _norm(inn)
    owner.ogrnip = _norm(ogrnip)
    owner.legal_address = _norm(legal_address)
    owner.bank_name = _norm(bank_name)
    owner.bank_account = _norm(bank_account)
    owner.corr_account = _norm(corr_account)
    owner.bik = _norm(bik)
    owner.signer_name = _norm(signer_name)
    await session.commit()
    return RedirectResponse("/requisites", status_code=303)


@app.post("/customers/add")
async def customers_add(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    form = dict(await request.form())
    name = (form.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Не указано наименование заказчика")
    customer = Customer(owner_id=owner.id)
    _apply_customer_form(customer, form)
    session.add(customer)
    await session.commit()
    return RedirectResponse("/requisites", status_code=303)


@app.post("/customers/{customer_id}")
async def customers_edit(
    customer_id: int,
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    customer = await session.get(Customer, customer_id)
    if customer is None or customer.owner_id != owner.id:
        raise HTTPException(status_code=404)
    form = dict(await request.form())
    if not (form.get("name") or "").strip():
        raise HTTPException(status_code=400, detail="Не указано наименование заказчика")
    _apply_customer_form(customer, form)
    await session.commit()
    return RedirectResponse("/requisites", status_code=303)


@app.post("/customers/{customer_id}/delete")
async def customers_delete(
    customer_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    customer = await session.get(Customer, customer_id)
    if customer is None or customer.owner_id != owner.id:
        raise HTTPException(status_code=404)
    await session.delete(customer)
    await session.commit()
    return RedirectResponse("/requisites", status_code=303)


@app.post("/admins/add")
async def admins_add(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    telegram_id: Annotated[str, Form()],
    name: Annotated[str, Form()] = "",
):
    """Добавить администратора кабинета по Telegram ID (полный доступ)."""
    try:
        tid = int(telegram_id.strip())
    except ValueError:
        raise HTTPException(status_code=400, detail="Telegram ID должен быть числом")
    # нельзя добавить самого владельца или уже существующего админа
    if tid == owner.telegram_id:
        raise HTTPException(status_code=400, detail="Это Telegram ID владельца")
    exists = (
        await session.execute(select(Admin).where(Admin.telegram_id == tid))
    ).scalar_one_or_none()
    if exists is None:
        session.add(Admin(owner_id=owner.id, telegram_id=tid, name=_norm(name)))
        try:
            await session.commit()
        except IntegrityError:
            await session.rollback()
    return RedirectResponse("/requisites", status_code=303)


@app.post("/admins/{admin_id}/notifications")
async def admins_toggle_notifications(
    admin_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Вкл/выкл дубли уведомлений бота этому админу (второй телефон)."""
    admin = await session.get(Admin, admin_id)
    if admin is None or admin.owner_id != owner.id:
        raise HTTPException(status_code=404)
    admin.notifications_enabled = not admin.notifications_enabled
    await session.commit()
    return RedirectResponse("/requisites", status_code=303)


@app.post("/admins/{admin_id}/delete")
async def admins_delete(
    admin_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    admin = await session.get(Admin, admin_id)
    if admin is None or admin.owner_id != owner.id:
        raise HTTPException(status_code=404)
    # Гасим и все устройства этого админа — доступ закрывается мгновенно.
    await session.execute(
        update(WebSession)
        .where(
            WebSession.owner_id == owner.id,
            WebSession.telegram_id == admin.telegram_id,
            WebSession.revoked_at.is_(None),
        )
        .values(revoked_at=datetime.now(timezone.utc))
    )
    await session.delete(admin)
    await session.commit()
    return RedirectResponse("/requisites", status_code=303)


# =====================================================================
# УСТРОЙСТВА (веб-сессии): завершить одно / все кроме текущего
# =====================================================================
@app.post("/sessions/{session_id}/revoke")
async def sessions_revoke(
    session_id: int,
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    ws = await session.get(WebSession, session_id)
    if ws is None or ws.owner_id != owner.id:
        raise HTTPException(status_code=404)
    current_ws = await _session_from_request(request, session)
    # Админ имеет полный доступ к кабинету (второй телефон владельца),
    # поэтому завершать можно любое устройство ЭТОГО кабинета.
    is_own = current_ws is not None and ws.id == current_ws.id
    ws.revoked_at = datetime.now(timezone.utc)
    await session.commit()
    if is_own:
        response = RedirectResponse("/login", status_code=303)
        auth_service.clear_session_cookie(response)
        response.delete_cookie("auth")
        return response
    return RedirectResponse("/requisites", status_code=303)


@app.post("/sessions/revoke-others")
async def sessions_revoke_others(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Выйти на всех устройствах кабинета, кроме текущего (владелец или админ)."""
    current_ws = await _session_from_request(request, session)
    stmt = update(WebSession).where(
        WebSession.owner_id == owner.id, WebSession.revoked_at.is_(None)
    )
    if current_ws is not None:
        stmt = stmt.where(WebSession.id != current_ws.id)
    await session.execute(stmt.values(revoked_at=datetime.now(timezone.utc)))
    await session.commit()
    return RedirectResponse("/requisites", status_code=303)


# --------- Excel formatting helpers ---------
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_MONEY_FMT = "#,##0 ₽"
_DATE_FMT = "DD.MM.YYYY"


def _style_header(ws, ncols: int) -> None:
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if len(text) > max_len:
                max_len = len(text)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)


def _build_finance_workbook(summary: dict, df: date, dt: date) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Итог"
    ws.append(["Показатель", "Сумма"])
    rows = [
        ("Период", f"{df.strftime('%d.%m.%Y')} — {dt.strftime('%d.%m.%Y')}"),
        ("Выручка по рейсам", float(summary["trip_revenue"])),
        ("Ручной доход", float(summary["manual_income"])),
        ("Топливо по рейсам", float(summary["fuel"])),
        ("Одобренные расходы водителей", float(summary["driver_expenses"])),
        ("Ручной расход", float(summary["manual_expense"])),
        ("Итого выручка", float(summary["total_income"])),
        ("Итого расход", float(summary["total_expense"])),
        ("Прибыль", float(summary["profit"])),
    ]
    for label, value in rows:
        ws.append([label, value])
    _style_header(ws, 2)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = _MONEY_FMT
            cell.alignment = Alignment(horizontal="right")
    # выделим строку «Прибыль»
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    _autosize(ws)
    return wb


async def _fill_entries_sheet(wb: Workbook, session, owner_id: int, df: date, dt: date) -> None:
    ws = wb.create_sheet("Ручные записи")
    ws.append(["Дата", "Тип", "Категория", "Сумма", "Описание"])
    entries_res = await session.execute(
        select(ManualEntry)
        .where(
            ManualEntry.owner_id == owner_id,
            ManualEntry.entry_date >= df,
            ManualEntry.entry_date <= dt,
        )
        .order_by(ManualEntry.entry_date)
    )
    for e in entries_res.scalars().all():
        ws.append([
            e.entry_date,
            "Доход" if e.type == "income" else "Расход",
            e.category or "",
            float(e.amount_rub),
            e.description or "",
        ])
    _style_header(ws, 5)
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).number_format = _DATE_FMT
        amount_cell = ws.cell(row=row_idx, column=4)
        amount_cell.number_format = _MONEY_FMT
        amount_cell.alignment = Alignment(horizontal="right")
    _autosize(ws)


async def _fill_trips_sheet(wb: Workbook, session, owner_id: int, df: date, dt: date) -> None:
    ws = wb.create_sheet("Рейсы")
    ws.append(["Дата", "Маршрут", "Водитель", "Машина", "Выручка", "Топливо", "Прибыль"])
    trips_res = await session.execute(
        select(Trip, Driver.full_name, Vehicle.license_plate)
        .join(Driver, Driver.id == Trip.driver_id)
        .join(Vehicle, Vehicle.id == Trip.vehicle_id)
        .where(
            Trip.owner_id == owner_id,
            Trip.status == "completed",
            func.date(Trip.completed_at) >= df,
            func.date(Trip.completed_at) <= dt,
        )
        .order_by(Trip.completed_at)
    )
    for trip, driver_name, plate in trips_res.all():
        ws.append([
            trip.completed_at.date() if trip.completed_at else None,
            f"{trip.origin or ''} → {trip.destination or ''}",
            driver_name,
            plate,
            float(trip.revenue_rub or 0),
            float(trip.fuel_cost_rub or 0),
            float(trip.profit_rub or 0),
        ])
    _style_header(ws, 7)
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).number_format = _DATE_FMT
        for col_idx in (5, 6, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = _MONEY_FMT
            cell.alignment = Alignment(horizontal="right")
    _autosize(ws)


def _parse_period(period_from: str | None, period_to: str | None) -> tuple[date, date]:
    today = date.today()
    default_from = today.replace(day=1)
    df = today.replace(day=1)
    dt = today
    if period_from:
        try:
            df = date.fromisoformat(period_from)
        except ValueError:
            df = default_from
    if period_to:
        try:
            dt = date.fromisoformat(period_to)
        except ValueError:
            dt = today
    if dt < df:
        df, dt = dt, df
    return df, dt


async def _finance_summary(
    session: AsyncSession, owner_id: int, df: date, dt: date
) -> dict:
    trip_revenue = (
        await session.execute(
            select(func.coalesce(func.sum(Trip.revenue_rub), 0)).where(
                Trip.owner_id == owner_id,
                Trip.status == "completed",
                func.date(Trip.completed_at) >= df,
                func.date(Trip.completed_at) <= dt,
            )
        )
    ).scalar_one() or Decimal(0)
    fuel = (
        await session.execute(
            select(func.coalesce(func.sum(Trip.fuel_cost_rub), 0)).where(
                Trip.owner_id == owner_id,
                Trip.status == "completed",
                func.date(Trip.completed_at) >= df,
                func.date(Trip.completed_at) <= dt,
            )
        )
    ).scalar_one() or Decimal(0)
    driver_expenses = (
        await session.execute(
            select(func.coalesce(func.sum(Expense.amount_rub), 0)).where(
                Expense.owner_id == owner_id,
                Expense.status == "approved",
                func.date(Expense.created_at) >= df,
                func.date(Expense.created_at) <= dt,
            )
        )
    ).scalar_one() or Decimal(0)
    manual_income = (
        await session.execute(
            select(func.coalesce(func.sum(ManualEntry.amount_rub), 0)).where(
                ManualEntry.owner_id == owner_id,
                ManualEntry.type == "income",
                ManualEntry.entry_date >= df,
                ManualEntry.entry_date <= dt,
            )
        )
    ).scalar_one() or Decimal(0)
    manual_expense = (
        await session.execute(
            select(func.coalesce(func.sum(ManualEntry.amount_rub), 0)).where(
                ManualEntry.owner_id == owner_id,
                ManualEntry.type == "expense",
                ManualEntry.entry_date >= df,
                ManualEntry.entry_date <= dt,
            )
        )
    ).scalar_one() or Decimal(0)

    total_income = Decimal(trip_revenue) + Decimal(manual_income)
    # Топливо не суммируем отдельно — оно уже учтено в одобренных expenses
    # (категория fuel). Иначе получаем двойной счёт.
    total_expense = Decimal(driver_expenses) + Decimal(manual_expense)
    return {
        "trip_revenue": Decimal(trip_revenue),
        "fuel": Decimal(fuel),
        "driver_expenses": Decimal(driver_expenses),
        "manual_income": Decimal(manual_income),
        "manual_expense": Decimal(manual_expense),
        "total_income": total_income,
        "total_expense": total_expense,
        "profit": total_income - total_expense,
    }


# =========================================================================
# /routes — прибыль по направлениям (группировка по origin → destination)
# =========================================================================
def _apply_distribution_center_form(center: DistributionCenter, form: dict) -> None:
    center.name = (form.get("name") or "").strip()
    center.address = (form.get("address") or "").strip()
    center.aliases = _norm(form.get("aliases"))
    try:
        center.latitude = rc_service.decimal_or_none(form.get("latitude"))
        center.longitude = rc_service.decimal_or_none(form.get("longitude"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    # радиус геозоны: пусто = глобальный (400 м); большие склады ставят больше
    radius_raw = str(form.get("geofence_radius_m") or "").strip()
    if not radius_raw:
        center.geofence_radius_m = None
    else:
        try:
            radius = int(radius_raw)
        except ValueError:
            raise HTTPException(status_code=400, detail="Радиус — целое число метров")
        if not (50 <= radius <= 5000):
            raise HTTPException(status_code=400, detail="Радиус: от 50 до 5000 м")
        center.geofence_radius_m = radius


@app.get("/routes", response_class=HTMLResponse)
async def routes_page(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    rc_imported: Annotated[int | None, Query()] = None,
):
    # Один запрос: GROUP BY origin, destination для completed рейсов с заполненной выручкой
    rows_res = await session.execute(
        select(
            Trip.origin,
            Trip.destination,
            func.count(Trip.id).label("trips_count"),
            func.coalesce(func.avg(Trip.revenue_rub), 0).label("avg_revenue"),
            func.coalesce(func.avg(Trip.profit_rub), 0).label("avg_profit"),
            func.coalesce(func.sum(Trip.revenue_rub), 0).label("total_revenue"),
            func.coalesce(func.sum(Trip.profit_rub), 0).label("total_profit"),
        )
        .where(
            Trip.owner_id == owner.id,
            Trip.status == "completed",
            Trip.revenue_rub.is_not(None),
            Trip.origin.is_not(None),
            Trip.destination.is_not(None),
        )
        .group_by(Trip.origin, Trip.destination)
    )
    rows = []
    for origin, destination, trips_count, avg_revenue, avg_profit, total_revenue, total_profit in rows_res.all():
        avg_revenue = Decimal(avg_revenue or 0)
        avg_profit = Decimal(avg_profit or 0)
        margin = (avg_profit / avg_revenue * Decimal(100)) if avg_revenue > 0 else Decimal(0)
        rows.append({
            "origin": origin, "destination": destination,
            "trips_count": trips_count,
            "avg_revenue": avg_revenue,
            "avg_profit": avg_profit,
            "avg_margin": margin,
            "total_revenue": Decimal(total_revenue or 0),
            "total_profit": Decimal(total_profit or 0),
        })
    # сортировка: от самых прибыльных к убыточным
    rows.sort(key=lambda r: r["avg_profit"], reverse=True)
    centers = await _active_distribution_centers(session, owner.id)
    # Маршруты-шаблоны (склад → РЦ) для водителя, сгруппированы по складу —
    # чтобы владелец быстро добавлял/видел, а водитель выбирал папками в боте.
    tmpl_rows = (
        await session.execute(
            select(RouteTemplate)
            .where(RouteTemplate.owner_id == owner.id, RouteTemplate.is_active.is_(True))
            .order_by(RouteTemplate.origin, RouteTemplate.destination)
        )
    ).scalars().all()
    templates_by_origin: dict[str, list] = {}
    for t in tmpl_rows:
        templates_by_origin.setdefault((t.origin or "—").strip(), []).append(t)
    return templates.TemplateResponse(
        "routes.html",
        {
            "request": request,
            "owner": owner,
            "rows": rows,
            "centers": centers,
            "templates_by_origin": templates_by_origin,
            "origins_list": sorted(templates_by_origin.keys()),
            "rc_imported": rc_imported,
            "yandex_maps_api_key": settings.yandex_maps_api_key,
            "active_page": "routes",
        },
    )


@app.post("/routes/template/add")
async def routes_template_add(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    origin: Annotated[str, Form()],
    destination: Annotated[str, Form()],
    cargo: Annotated[str, Form()] = "",
):
    """Быстро добавить маршрут склад→РЦ (шаблон для бота водителя).
    Название генерируем сами; РЦ выбирается из справочника (текст его имени)."""
    origin = (origin or "").strip()
    destination = (destination or "").strip()
    if not origin or not destination:
        raise HTTPException(status_code=400, detail="Укажите склад и РЦ")
    # не плодим дубли: тот же склад+РЦ обновляем/оставляем
    existing = (
        await session.execute(
            select(RouteTemplate).where(
                RouteTemplate.owner_id == owner.id,
                RouteTemplate.origin == origin,
                RouteTemplate.destination == destination,
            )
        )
    ).scalar_one_or_none()
    if existing is None:
        session.add(RouteTemplate(
            owner_id=owner.id,
            name=f"{origin} → {destination}"[:100],
            origin=origin, destination=destination,
            default_cargo=_norm(cargo), is_active=True,
        ))
    else:
        existing.is_active = True
        existing.default_cargo = _norm(cargo) or existing.default_cargo
    await session.commit()
    return RedirectResponse("/routes", status_code=303)


@app.post("/routes/template/{template_id}/delete")
async def routes_template_delete(
    template_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    tmpl = await session.get(RouteTemplate, template_id)
    if tmpl is None or tmpl.owner_id != owner.id:
        raise HTTPException(status_code=404)
    tmpl.is_active = False
    await session.commit()
    return RedirectResponse("/routes", status_code=303)


@app.post("/routes/rc/add")
async def routes_rc_add(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    name: Annotated[str, Form()],
    address: Annotated[str, Form()],
    aliases: Annotated[str, Form()] = "",
    latitude: Annotated[str, Form()] = "",
    longitude: Annotated[str, Form()] = "",
    geofence_radius_m: Annotated[str, Form()] = "",
):
    if not name.strip() or not address.strip():
        raise HTTPException(status_code=400, detail="Укажите название и адрес РЦ")
    existing_res = await session.execute(
        select(DistributionCenter).where(DistributionCenter.owner_id == owner.id)
    )
    key = rc_service.route_key(name)
    center = next(
        (c for c in existing_res.scalars().all() if rc_service.route_key(c.name) == key),
        None,
    )
    if center is None:
        center = DistributionCenter(owner_id=owner.id, name=name.strip(), address=address.strip())
        session.add(center)
    _apply_distribution_center_form(center, {
        "name": name, "address": address, "aliases": aliases,
        "latitude": latitude, "longitude": longitude,
        "geofence_radius_m": geofence_radius_m,
    })
    center.is_active = True
    await session.commit()
    return RedirectResponse("/routes", status_code=303)


@app.post("/routes/rc/import")
async def routes_rc_import(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    file: Annotated[UploadFile, File()],
):
    data = await file.read()
    if len(data) > _MAX_DOC_BYTES:
        raise HTTPException(status_code=400, detail="Файл слишком большой (макс 6 МБ)")
    try:
        items = rc_service.distribution_centers_from_xlsx(data)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    existing_res = await session.execute(
        select(DistributionCenter).where(DistributionCenter.owner_id == owner.id)
    )
    existing = {
        rc_service.route_key(center.name): center
        for center in existing_res.scalars().all()
    }
    imported = 0
    for item in items:
        key = rc_service.route_key(item["name"])
        if not key:
            continue
        center = existing.get(key)
        if center is None:
            center = DistributionCenter(owner_id=owner.id, name=item["name"], address=item["address"])
            session.add(center)
            existing[key] = center
        _apply_distribution_center_form(center, item)
        center.is_active = True
        imported += 1
    await session.commit()
    return RedirectResponse(f"/routes?rc_imported={imported}", status_code=303)


@app.post("/routes/rc/geocode")
async def routes_rc_geocode(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Определить координаты РЦ по адресам через бесплатный Nominatim (OSM).

    Берём только РЦ без координат. За один клик — до 20 адресов (между
    запросами пауза ~1 с, политика сервиса), при большом списке владелец
    жмёт кнопку ещё раз — остаток показываем в уведомлении.
    """
    centers_res = await session.execute(
        select(DistributionCenter)
        .where(
            DistributionCenter.owner_id == owner.id,
            DistributionCenter.is_active.is_(True),
            (DistributionCenter.latitude.is_(None))
            | (DistributionCenter.longitude.is_(None)),
        )
        .order_by(DistributionCenter.name)
    )
    pending = [c for c in centers_res.scalars().all() if (c.address or "").strip()]
    if not pending:
        # у всех РЦ координаты уже стоят — кнопке нечего делать
        return RedirectResponse("/routes?geo_none=1", status_code=303)
    batch = pending[:20]
    left = len(pending) - len(batch)
    found = failed = 0
    results = await geocode_service.geocode_many([c.address for c in batch])
    for center, coords in zip(batch, results):
        if coords is None:
            failed += 1
            continue
        center.latitude, center.longitude = coords
        found += 1
    await session.commit()
    return RedirectResponse(
        f"/routes?geo_ok={found}&geo_fail={failed}&geo_left={left}", status_code=303
    )


@app.post("/routes/rc/{center_id}")
async def routes_rc_edit(
    center_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    name: Annotated[str, Form()],
    address: Annotated[str, Form()],
    aliases: Annotated[str, Form()] = "",
    latitude: Annotated[str, Form()] = "",
    longitude: Annotated[str, Form()] = "",
    geofence_radius_m: Annotated[str, Form()] = "",
):
    center = await session.get(DistributionCenter, center_id)
    if center is None or center.owner_id != owner.id:
        raise HTTPException(status_code=404)
    if not name.strip() or not address.strip():
        raise HTTPException(status_code=400, detail="Укажите название и адрес РЦ")
    _apply_distribution_center_form(center, {
        "name": name, "address": address, "aliases": aliases,
        "latitude": latitude, "longitude": longitude,
        "geofence_radius_m": geofence_radius_m,
    })
    try:
        await session.commit()
    except IntegrityError as exc:
        await session.rollback()
        raise HTTPException(status_code=400, detail="РЦ с таким названием уже есть") from exc
    return RedirectResponse("/routes", status_code=303)


@app.post("/routes/rc/{center_id}/delete")
async def routes_rc_delete(
    center_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    center = await session.get(DistributionCenter, center_id)
    if center is None or center.owner_id != owner.id:
        raise HTTPException(status_code=404)
    center.is_active = False
    await session.commit()
    return RedirectResponse("/routes", status_code=303)


# =========================================================================
# /stats — глобальная статистика: журнал простоев на РЦ, сводки
# =========================================================================
def _minutes_label(minutes: int | None) -> str:
    return telemetry_service.minutes_label(minutes)


def _payload_dt(value) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed


def _minutes_between(start: datetime | None, end: datetime | None = None) -> int:
    if start is None:
        return 0
    finish = end or datetime.now(timezone.utc)
    if start.tzinfo is None:
        start = start.replace(tzinfo=timezone.utc)
    if finish.tzinfo is None:
        finish = finish.replace(tzinfo=timezone.utc)
    return max(0, int((finish - start).total_seconds() // 60))


def _issue_counts(issues: list[dict]) -> dict:
    counts = {"total": len(issues), "danger": 0, "warn": 0, "info": 0}
    for issue in issues:
        sev = issue.get("sev")
        if sev in counts:
            counts[sev] += 1
    return counts


@app.post("/stats/downtime/{event_id}/hide")
async def stats_downtime_hide(
    event_id: int,
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    reason: Annotated[str, Form()] = "",
):
    """Скрыть ложную стоянку из статистики (например, машина стояла дома, а
    геозона приписала её РЦ). GPS-данные НЕ удаляем — только помечаем событие
    rc_departed флагом ignored, с автором/временем/причиной для аудита."""
    ev = await session.get(Event, event_id)
    if ev is None or ev.owner_id != owner.id or ev.event_type != "rc_departed":
        raise HTTPException(status_code=404)
    viewer_tid = await _viewer_telegram_id(request, session)
    payload = dict(ev.payload or {})
    payload["ignored"] = True
    payload["ignored_reason"] = (reason or "").strip()[:500]
    payload["ignored_by"] = viewer_tid
    payload["ignored_at"] = datetime.now(timezone.utc).isoformat()
    ev.payload = payload  # переприсваиваем целиком — иначе SQLAlchemy не увидит правку JSONB
    await session.commit()
    return RedirectResponse(f"/stats?{request.url.query}", status_code=303)


@app.get("/stats/downtime/{event_id}/edit", response_class=HTMLResponse)
async def stats_downtime_edit_page(
    event_id: int,
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Страница правки одной стоянки: минуты, сумма к выставлению, водитель."""
    ev = await session.get(Event, event_id)
    if ev is None or ev.owner_id != owner.id or ev.event_type != "rc_departed":
        raise HTTPException(status_code=404)
    payload = ev.payload or {}
    drivers = list(
        (
            await session.execute(
                select(Driver).where(Driver.owner_id == owner.id, Driver.is_active.is_(True))
                .order_by(Driver.full_name)
            )
        ).scalars().all()
    )
    cur_minutes = telemetry_service.int_or_none(payload.get("corrected_waited_minutes"))
    if cur_minutes is None:
        cur_minutes = telemetry_service.int_or_none(payload.get("waited_minutes")) or 0
    cur_billable = telemetry_service.int_or_none(payload.get("corrected_billable_rub"))
    if cur_billable is None:
        cur_billable = telemetry_service.rc_billable_downtime_rub(cur_minutes)
    # 0 в corrected_driver_id = явно «нет водителя» → в форме «не указан»
    if "corrected_driver_id" in payload:
        cd = telemetry_service.int_or_none(payload.get("corrected_driver_id"))
        cur_driver = cd if cd else None
    else:
        cur_driver = telemetry_service.int_or_none(payload.get("driver_id"))
    return templates.TemplateResponse(
        "downtime_edit.html",
        {
            "request": request, "owner": owner, "active_page": "stats",
            "event_id": event_id,
            "rc_name": payload.get("rc_name") or "—",
            "plate": payload.get("plate") or "—",
            "cur_minutes": cur_minutes, "cur_billable": cur_billable,
            "cur_driver": cur_driver, "drivers": drivers,
            "back_query": request.url.query,
        },
    )


@app.post("/stats/downtime/{event_id}/correct")
async def stats_downtime_correct(
    event_id: int,
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    waited_minutes: Annotated[str, Form()],
    billable_rub: Annotated[str, Form()] = "",
    driver_id: Annotated[str, Form()] = "",
    back_query: Annotated[str, Form()] = "",
):
    """Поправить стоянку вручную: минуты, сумму «к выставлению», водителя.
    Аудит: кто/когда правил. Пустая сумма = считать от минут автоматически."""
    ev = await session.get(Event, event_id)
    if ev is None or ev.owner_id != owner.id or ev.event_type != "rc_departed":
        raise HTTPException(status_code=404)
    try:
        minutes = max(0, int(str(waited_minutes).strip()))
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Минуты — целое число")
    viewer_tid = await _viewer_telegram_id(request, session)
    payload = dict(ev.payload or {})
    payload["corrected_waited_minutes"] = minutes
    # сумма: пусто → авто-расчёт (убираем ручную), иначе фиксируем
    bill_raw = str(billable_rub).strip()
    if bill_raw == "":
        payload.pop("corrected_billable_rub", None)
    else:
        try:
            payload["corrected_billable_rub"] = max(0, int(bill_raw))
        except ValueError:
            raise HTTPException(status_code=400, detail="Сумма — целое число рублей")
    # водитель: пусто = ЯВНО «без водителя» (прочерк), число = конкретный
    # водитель. 0 — признак «явно нет водителя» (иначе система заново вычисляла
    # его из смен и прочерк не ставился).
    drv_raw = str(driver_id).strip()
    if not drv_raw:
        payload["corrected_driver_id"] = 0
    else:
        try:
            drv_id = int(drv_raw)
        except ValueError:
            raise HTTPException(status_code=400, detail="Bad driver_id")
        drv = await session.get(Driver, drv_id)
        if drv is None or drv.owner_id != owner.id:
            raise HTTPException(status_code=400, detail="Bad driver_id")
        payload["corrected_driver_id"] = drv_id
    payload["corrected_by"] = viewer_tid
    payload["corrected_at"] = datetime.now(timezone.utc).isoformat()
    ev.payload = payload
    await session.commit()
    q = (back_query or "").lstrip("?")
    return RedirectResponse(f"/stats?{q}" if q else "/stats", status_code=303)


@app.post("/stats/downtime/{event_id}/unhide")
async def stats_downtime_unhide(
    event_id: int,
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Вернуть ошибочно скрытую стоянку обратно в статистику."""
    ev = await session.get(Event, event_id)
    if ev is None or ev.owner_id != owner.id or ev.event_type != "rc_departed":
        raise HTTPException(status_code=404)
    payload = dict(ev.payload or {})
    for k in ("ignored", "ignored_reason", "ignored_by", "ignored_at"):
        payload.pop(k, None)
    ev.payload = payload
    await session.commit()
    return RedirectResponse(f"/stats?{request.url.query}", status_code=303)


@app.get("/stats", response_class=HTMLResponse)
async def stats_page(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    period_from: Annotated[str | None, Query()] = None,
    period_to: Annotated[str | None, Query()] = None,
    show_hidden: Annotated[str | None, Query()] = None,
):
    """Статистика: простои на РЦ (для счетов за простой), сводка по РЦ,
    по водителям, заказы по неделям. Источник простоев — события
    rc_departed от детектора геозон."""
    tz = owner_tz(owner.timezone)
    today = datetime.now(tz).date()

    def _parse_date(raw: str | None, default: date) -> date:
        try:
            return date.fromisoformat((raw or "").strip())
        except ValueError:
            return default

    df = _parse_date(period_from, month_floor(today))
    dt_ = _parse_date(period_to, today)
    if df > dt_:
        df, dt_ = dt_, df
    start_utc = datetime.combine(df, datetime.min.time()).replace(tzinfo=tz).astimezone(timezone.utc)
    end_utc = (
        datetime.combine(dt_, datetime.min.time()).replace(tzinfo=tz) + timedelta(days=1)
    ).astimezone(timezone.utc)

    # --- справочники ---
    plates = dict(
        (
            await session.execute(
                select(Vehicle.id, Vehicle.license_plate).where(Vehicle.owner_id == owner.id)
            )
        ).all()
    )
    rc_names = dict(
        (
            await session.execute(
                select(DistributionCenter.id, DistributionCenter.name).where(
                    DistributionCenter.owner_id == owner.id
                )
            )
        ).all()
    )
    driver_names = dict(
        (
            await session.execute(
                select(Driver.id, Driver.full_name).where(Driver.owner_id == owner.id)
            )
        ).all()
    )

    # --- смены, пересекающие период (для привязки водителя к простою) ---
    shifts_rows = (
        await session.execute(
            select(Shift.vehicle_id, Shift.driver_id, Shift.started_at, Shift.ended_at)
            .where(
                Shift.owner_id == owner.id,
                Shift.started_at <= end_utc,
                (Shift.ended_at.is_(None)) | (Shift.ended_at >= start_utc),
            )
        )
    ).all()
    trip_route_rows = (
        await session.execute(
            select(
                Trip.vehicle_id,
                Trip.driver_id,
                Trip.origin,
                Trip.destination,
                Trip.created_at,
                Trip.completed_at,
            )
            .where(
                Trip.owner_id == owner.id,
                Trip.created_at <= end_utc,
                (Trip.completed_at.is_(None)) | (Trip.completed_at >= start_utc),
            )
            .order_by(Trip.created_at)
        )
    ).all()

    def _driver_at(vehicle_id: int | None, at: datetime | None) -> int | None:
        if vehicle_id is None or at is None:
            return None
        for vid, did, s, e in shifts_rows:
            if vid == vehicle_id and s is not None and s <= at and (e is None or at <= e):
                return did
        return None

    def _route_at(
        vehicle_id: int | None, driver_id: int | None, at: datetime | None
    ) -> str | None:
        if vehicle_id is None or at is None:
            return None
        for vid, did, origin, destination, created_at, completed_at in trip_route_rows:
            if vid != vehicle_id or created_at is None:
                continue
            if driver_id is not None and did != driver_id:
                continue
            finish = completed_at or end_utc
            if created_at <= at <= finish + timedelta(hours=2):
                return f"{origin or '—'} → {destination or '—'}"
        return None

    # --- журнал простоев (rc_departed). Итоги считаем по всему периоду, а
    # в таблицу отдаём последние 500 строк, чтобы страница не тяжелела.
    ev_rows = (
        await session.execute(
            select(Event.id, Event.created_at, Event.payload)
            .where(
                Event.owner_id == owner.id,
                Event.event_type == "rc_departed",
                Event.created_at >= start_utc,
                Event.created_at < end_utc,
            )
            .order_by(desc(Event.created_at))
        )
    ).all()
    journal_all = []
    hidden_rows = []
    for event_id, created_at, payload in ev_rows:
        payload = payload or {}
        # Ложную стоянку владелец скрывает — она исключается из всех итогов
        # (KPI, сводки, «к выставлению», экспорт). Сама GPS-точка не удаляется.
        if payload.get("ignored"):
            hidden_rows.append({
                "event_id": event_id,
                "arrived_at": _payload_dt(payload.get("arrived_at")) or created_at,
                "departed_at": _payload_dt(payload.get("departed_at")) or created_at,
                "plate": payload.get("plate") or plates.get(
                    telemetry_service.int_or_none(payload.get("vehicle_id")), "—"),
                "rc_name": payload.get("rc_name") or rc_names.get(
                    telemetry_service.int_or_none(payload.get("rc_id")), "—"),
                "waited_label": _minutes_label(
                    telemetry_service.int_or_none(payload.get("waited_minutes")) or 0),
                "reason": payload.get("ignored_reason") or "",
            })
            continue
        # Владелец мог поправить строку вручную (GPS ошибся): минуты стоянки,
        # сумму «к выставлению» и водителя. Ручные значения имеют приоритет.
        corrected = telemetry_service.int_or_none(payload.get("corrected_waited_minutes"))
        corrected_billable = telemetry_service.int_or_none(payload.get("corrected_billable_rub"))
        # водитель поправлен вручную, если ключ вообще присутствует (0 = явно «нет»)
        driver_overridden = "corrected_driver_id" in payload
        corrected_driver = telemetry_service.int_or_none(payload.get("corrected_driver_id"))
        is_corrected = driver_overridden or any(
            payload.get(k) is not None
            for k in ("corrected_waited_minutes", "corrected_billable_rub")
        )
        waited = max(0, corrected if corrected is not None
                     else (telemetry_service.int_or_none(payload.get("waited_minutes")) or 0))
        if corrected_billable is not None:
            billable = max(0, corrected_billable)
        elif corrected is not None:
            billable = telemetry_service.rc_billable_downtime_rub(waited)
        else:
            payload_billable = telemetry_service.int_or_none(payload.get("billable_downtime_rub"))
            computed_billable = telemetry_service.rc_billable_downtime_rub(waited)
            billable = max(computed_billable, payload_billable or 0)
        vid = telemetry_service.int_or_none(payload.get("vehicle_id"))
        rcid = telemetry_service.int_or_none(payload.get("rc_id"))
        arrived_at = _payload_dt(payload.get("arrived_at")) or (created_at - timedelta(minutes=waited))
        departed_at = _payload_dt(payload.get("departed_at")) or created_at
        if driver_overridden:
            # ручная правка имеет приоритет; 0 = «водителя нет» → прочерк
            driver_id = corrected_driver if corrected_driver else None
        else:
            driver_id = (
                telemetry_service.int_or_none(payload.get("driver_id"))
                or _driver_at(vid, departed_at)
                or _driver_at(vid, arrived_at)
            )
        driver_name = driver_names.get(driver_id) if driver_id else None
        if driver_name is None and not driver_overridden:
            driver_name = payload.get("driver_name")
        route = (
            payload.get("route")
            or _route_at(vid, driver_id, departed_at)
            or _route_at(vid, driver_id, arrived_at)
        )
        journal_all.append({
            "event_id": event_id,
            "corrected": is_corrected,
            "arrived_at": arrived_at,
            "departed_at": departed_at,
            "plate": payload.get("plate") or plates.get(vid, "—"),
            "rc_name": payload.get("rc_name") or rc_names.get(rcid, "—"),
            "rc_id": rcid,
            "driver_id": driver_id,
            "driver": driver_name,
            "route": route,
            "waited_minutes": waited,
            "waited_label": _minutes_label(waited),
            "engine_off_label": _minutes_label(payload.get("engine_off_minutes")),
            "billable_downtime_rub": billable,
            "billable_label": telemetry_service.rub_label(billable),
            "billable_blocks": billable // telemetry_service.RC_BILLABLE_DOWNTIME_RUB
            if billable else 0,
        })

    journal = journal_all[:500]
    total_wait_min = sum(r["waited_minutes"] for r in journal_all)
    billable_total = sum(r["billable_downtime_rub"] for r in journal_all)
    kpi = {
        "visits": len(journal_all),
        "total_wait": _minutes_label(total_wait_min),
        "avg_wait": _minutes_label(total_wait_min // len(journal_all)) if journal_all else "—",
        "billable_label": telemetry_service.rub_label(billable_total),
    }
    billable_alerts = sorted(
        (r for r in journal_all if r["billable_downtime_rub"] > 0),
        key=lambda x: (x["billable_downtime_rub"], x["waited_minutes"]),
        reverse=True,
    )[:12]

    # --- сводка по РЦ ---
    rc_agg: dict[int, dict] = {}
    for r in journal_all:
        agg = rc_agg.setdefault(
            r["rc_id"], {"name": r["rc_name"], "visits": 0, "total": 0, "billable": 0}
        )
        agg["visits"] += 1
        agg["total"] += r["waited_minutes"]
        agg["billable"] += r["billable_downtime_rub"]
    rc_summary = sorted(
        (
            {
                "name": a["name"], "visits": a["visits"],
                "total_label": _minutes_label(a["total"]),
                "avg_label": _minutes_label(a["total"] // a["visits"]) if a["visits"] else "—",
                "total": a["total"],
                "billable_label": telemetry_service.rub_label(a["billable"]),
                "billable": a["billable"],
            }
            for a in rc_agg.values()
        ),
        key=lambda x: x["total"], reverse=True,
    )

    # --- сводка по водителям: простой + рейсы + км (одометр смен) ---
    idle_by_driver: dict[int, int] = {}
    billable_by_driver: dict[int, int] = {}
    idle_name_only: dict[str, int] = {}
    billable_name_only: dict[str, int] = {}
    for r in journal_all:
        if r["driver_id"]:
            idle_by_driver[r["driver_id"]] = idle_by_driver.get(r["driver_id"], 0) + r["waited_minutes"]
            billable_by_driver[r["driver_id"]] = (
                billable_by_driver.get(r["driver_id"], 0) + r["billable_downtime_rub"]
            )
        elif r["driver"]:
            idle_name_only[r["driver"]] = idle_name_only.get(r["driver"], 0) + r["waited_minutes"]
            billable_name_only[r["driver"]] = (
                billable_name_only.get(r["driver"], 0) + r["billable_downtime_rub"]
            )
    trips_by_driver = dict(
        (
            await session.execute(
                select(Trip.driver_id, func.count(Trip.id))
                .where(
                    Trip.owner_id == owner.id,
                    Trip.status == "completed",
                    Trip.completed_at >= start_utc,
                    Trip.completed_at < end_utc,
                )
                .group_by(Trip.driver_id)
            )
        ).all()
    )
    km_by_driver = dict(
        (
            await session.execute(
                select(Shift.driver_id, func.coalesce(func.sum(Shift.distance_km), 0))
                .where(
                    Shift.owner_id == owner.id,
                    Shift.status == "completed",
                    Shift.ended_at >= start_utc,
                    Shift.ended_at < end_utc,
                )
                .group_by(Shift.driver_id)
            )
        ).all()
    )
    driver_ids = set(trips_by_driver) | set(km_by_driver) | set(idle_by_driver)
    driver_summary = []
    for did in driver_ids:
        name = driver_names.get(did, f"Водитель {did}")
        driver_summary.append({
            "name": name,
            "trips": trips_by_driver.get(did, 0),
            "km": int(km_by_driver.get(did, 0) or 0),
            "idle_label": _minutes_label(idle_by_driver.get(did, 0)),
            "idle_minutes": idle_by_driver.get(did, 0),
            "billable_label": telemetry_service.rub_label(billable_by_driver.get(did, 0)),
            "billable": billable_by_driver.get(did, 0),
        })
    for name, idle_minutes in idle_name_only.items():
        driver_summary.append({
            "name": name,
            "trips": 0,
            "km": 0,
            "idle_label": _minutes_label(idle_minutes),
            "idle_minutes": idle_minutes,
            "billable_label": telemetry_service.rub_label(billable_name_only.get(name, 0)),
            "billable": billable_name_only.get(name, 0),
        })
    driver_summary.sort(key=lambda x: x["idle_minutes"], reverse=True)

    # --- заказы по неделям (завершённые рейсы) ---
    week_rows = (
        await session.execute(
            select(Trip.completed_at, Trip.revenue_rub)
            .where(
                Trip.owner_id == owner.id,
                Trip.status == "completed",
                Trip.completed_at >= start_utc,
                Trip.completed_at < end_utc,
            )
        )
    ).all()
    weeks: dict[date, dict] = {}
    for completed_at, revenue in week_rows:
        local_day = completed_at.astimezone(tz).date()
        monday = local_day - timedelta(days=local_day.weekday())
        agg = weeks.setdefault(monday, {"trips": 0, "revenue": Decimal(0)})
        agg["trips"] += 1
        agg["revenue"] += Decimal(revenue or 0)
    week_summary = [
        {"week": f"нед. {k:%d.%m}", "trips": v["trips"], "revenue": v["revenue"]}
        for k, v in sorted(weeks.items())
    ]

    # --- живой операционный контроль: что владельцу стоит проверить сейчас ---
    now_utc = datetime.now(timezone.utc)
    live_issues: list[dict] = []

    def add_issue(
        sev: str, icon: str, title: str, pill: str, sub: str, href: str | None = None
    ) -> None:
        live_issues.append({
            "sev": sev,
            "icon": icon,
            "title": title,
            "pill": pill,
            "sub": sub,
            "href": href,
        })

    active_shift_rows = (
        await session.execute(
            select(
                Shift.id,
                Shift.vehicle_id,
                Shift.driver_id,
                Shift.started_at,
                Driver.full_name,
                Vehicle.license_plate,
            )
            .join(Driver, Driver.id == Shift.driver_id)
            .join(Vehicle, Vehicle.id == Shift.vehicle_id)
            .where(Shift.owner_id == owner.id, Shift.status == "started")
        )
    ).all()
    active_shift_by_vehicle: dict[int, dict] = {}
    for shift_id, vehicle_id, driver_id, started_at, driver_name, plate in active_shift_rows:
        active_shift_by_vehicle[vehicle_id] = {
            "shift_id": shift_id,
            "driver_id": driver_id,
            "driver": driver_name,
            "plate": plate,
            "started_at": started_at,
        }
        minutes_open = _minutes_between(started_at, now_utc)
        if minutes_open >= 24 * 60:
            add_issue(
                "danger",
                "🕒",
                "Смена открыта больше 24 часов",
                _minutes_label(minutes_open),
                f"{driver_name} · {plate} · {smart_since_label(started_at, owner.timezone)}",
                "/drivers",
            )
        elif minutes_open >= 18 * 60:
            add_issue(
                "warn",
                "🕒",
                "Смена скоро станет длинной",
                _minutes_label(minutes_open),
                f"{driver_name} · {plate} · проверьте, не забыл ли водитель закрыть смену",
                "/drivers",
            )

    active_trip_rows = (
        await session.execute(
            select(
                Trip.id,
                Trip.vehicle_id,
                Trip.origin,
                Trip.destination,
                Trip.status,
                Vehicle.license_plate,
                Driver.full_name,
            )
            .join(Vehicle, Vehicle.id == Trip.vehicle_id)
            .join(Driver, Driver.id == Trip.driver_id)
            .where(
                Trip.owner_id == owner.id,
                Trip.status.in_(("created", "in_transit", "unloading")),
            )
        )
    ).all()
    active_trip_by_vehicle: dict[int, dict] = {}
    for trip_id, vehicle_id, origin, destination, status, plate, driver_name in active_trip_rows:
        active_trip_by_vehicle[vehicle_id] = {
            "trip_id": trip_id,
            "route": f"{origin or '—'} → {destination or '—'}",
            "status": status,
            "plate": plate,
            "driver": driver_name,
        }

    pending_revenue_rows = (
        await session.execute(
            select(
                Trip.id,
                Trip.origin,
                Trip.destination,
                Trip.driver_revenue_pending_rub,
                Driver.full_name,
                Vehicle.license_plate,
            )
            .join(Driver, Driver.id == Trip.driver_id)
            .join(Vehicle, Vehicle.id == Trip.vehicle_id)
            .where(
                Trip.owner_id == owner.id,
                Trip.status == "completed",
                Trip.revenue_rub.is_(None),
                Trip.driver_revenue_pending_rub.is_not(None),
            )
            .order_by(desc(Trip.completed_at))
            .limit(5)
        )
    ).all()
    for trip_id, origin, destination, amount, driver_name, plate in pending_revenue_rows:
        add_issue(
            "warn",
            "💰",
            "Выручка ждёт подтверждения",
            telemetry_service.rub_label(amount),
            f"{driver_name} · {plate} · {origin or '—'} → {destination or '—'}",
            f"/trips/{trip_id}",
        )

    stale_cutoff = now_utc - timedelta(minutes=30)
    vehicle_state_rows = (
        await session.execute(
            select(Vehicle, VehicleState)
            .outerjoin(VehicleState, VehicleState.vehicle_id == Vehicle.id)
            .where(Vehicle.owner_id == owner.id, Vehicle.is_active.is_(True))
            .order_by(Vehicle.license_plate)
        )
    ).all()
    for vehicle, state in vehicle_state_rows:
        if state is None:
            if vehicle.stavtrack_object_id:
                add_issue(
                    "warn",
                    "📡",
                    "По машине ещё нет GPS-точек",
                    vehicle.stavtrack_object_id,
                    f"{vehicle.license_plate} · Stavtrack ID привязан, но поток ещё не пришёл",
                    "/map",
                )
            continue

        gps_stale = state.last_seen_at is None or state.last_seen_at < stale_cutoff
        gps_invalid = state.is_valid is False
        motion_status = state.motion_status or telemetry_service.vehicle_motion_status(
            state.speed_kmh, state.ignition
        )
        signal = telemetry_service.vehicle_control_signal(
            motion_status=motion_status,
            has_active_shift=vehicle.id in active_shift_by_vehicle,
            has_active_trip=vehicle.id in active_trip_by_vehicle,
            gps_stale=gps_stale,
            gps_invalid=gps_invalid,
        )
        since_label = smart_since_label(state.motion_since_at, owner.timezone)
        if signal == telemetry_service.SIGNAL_GPS_STALE:
            add_issue(
                "danger",
                "📡",
                "GPS давно не обновлялся",
                telemetry_service.duration_label(state.last_seen_at, now_utc),
                f"{vehicle.license_plate} · последний сигнал {fmt_dt(state.last_seen_at, owner.timezone, '%d.%m %H:%M')}",
                "/map",
            )
        elif signal == telemetry_service.SIGNAL_GPS_INVALID:
            add_issue(
                "danger",
                "📍",
                "GPS прислал недостоверные координаты",
                "проверить",
                f"{vehicle.license_plate} · метка не должна прыгать на мусорные координаты",
                "/map",
            )
        elif signal == telemetry_service.SIGNAL_MOVING_WITHOUT_SHIFT:
            add_issue(
                "danger",
                "🚛",
                "Машина едет без открытой смены",
                f"{Decimal(state.speed_kmh or 0):.0f} км/ч",
                f"{vehicle.license_plate} · {since_label}",
                "/map",
            )
        elif signal == telemetry_service.SIGNAL_MOVING_WITHOUT_TRIP:
            shift = active_shift_by_vehicle.get(vehicle.id, {})
            add_issue(
                "warn",
                "🛣",
                "Машина едет без активного рейса",
                f"{Decimal(state.speed_kmh or 0):.0f} км/ч",
                f"{vehicle.license_plate} · {shift.get('driver', 'водитель в смене')} · {since_label}",
                "/map",
            )
        elif (
            signal == telemetry_service.SIGNAL_IDLE_ENGINE
            and telemetry_service.parked_long_enough(
                motion_status, state.motion_since_at, now_utc, min_minutes=15
            )
        ):
            add_issue(
                "warn",
                "⛽",
                "Стоит с заведённым двигателем",
                telemetry_service.duration_label(state.motion_since_at, now_utc),
                f"{vehicle.license_plate} · {since_label}",
                "/map",
            )

    rc_events = (
        await session.execute(
            select(Event.event_type, Event.created_at, Event.payload)
            .where(
                Event.owner_id == owner.id,
                Event.event_type.in_(("rc_arrived", "rc_departed", "rc_downtime_alert")),
                Event.created_at >= now_utc - timedelta(days=14),
            )
            .order_by(Event.created_at)
        )
    ).all()
    active_rc: dict[tuple[int, int], dict] = {}
    for event_type, created_at, payload in rc_events:
        payload = payload or {}
        vid = telemetry_service.int_or_none(payload.get("vehicle_id"))
        rcid = telemetry_service.int_or_none(payload.get("rc_id"))
        if vid is None or rcid is None:
            continue
        key = (vid, rcid)
        if event_type == "rc_arrived":
            active_rc[key] = {
                "created_at": created_at,
                "payload": payload,
                "alerted": False,
            }
        elif event_type == "rc_departed":
            active_rc.pop(key, None)
        elif event_type == "rc_downtime_alert" and key in active_rc:
            active_rc[key]["alerted"] = True

    for (_vid, _rcid), state in active_rc.items():
        payload = state["payload"]
        parked_since = _payload_dt(payload.get("parked_since")) or state["created_at"]
        waited_minutes = _minutes_between(parked_since, now_utc)
        billable = telemetry_service.rc_billable_downtime_rub(waited_minutes)
        trip = active_trip_by_vehicle.get(_vid)
        route_part = f" · {trip['route']}" if trip else ""
        if billable:
            add_issue(
                "danger",
                "🏬",
                "Машина стоит на РЦ больше 12 часов",
                telemetry_service.rub_label(billable),
                f"{payload.get('plate') or 'машина'} · {payload.get('rc_name') or 'РЦ'}{route_part} · {_minutes_label(waited_minutes)}",
                "/stats",
            )
        elif waited_minutes >= 6 * 60:
            add_issue(
                "warn",
                "🏬",
                "Долгая стоянка на РЦ",
                _minutes_label(waited_minutes),
                f"{payload.get('plate') or 'машина'} · {payload.get('rc_name') or 'РЦ'} · до платного порога ещё {_minutes_label(telemetry_service.RC_BILLABLE_WAIT_MINUTES - waited_minutes)}",
                "/stats",
            )

    live_issues.sort(
        key=lambda x: {"danger": 0, "warn": 1, "info": 2}.get(x.get("sev"), 3)
    )
    live_issue_counts = _issue_counts(live_issues)

    return templates.TemplateResponse(
        "stats.html",
        {
            "request": request, "owner": owner, "active_page": "stats",
            "period_from": df.isoformat(), "period_to": dt_.isoformat(),
            "kpi": kpi, "journal": journal, "rc_summary": rc_summary,
            "driver_summary": driver_summary, "week_summary": week_summary,
            "billable_alerts": billable_alerts,
            "live_issues": live_issues, "live_issue_counts": live_issue_counts,
            "hidden_rows": hidden_rows, "show_hidden": bool(show_hidden),
        },
    )


# =========================================================================
# /map — карта водителей и машин с последними координатами
# =========================================================================
@app.get("/map", response_class=HTMLResponse)
async def map_page(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
):
    return templates.TemplateResponse(
        "map.html",
        {
            "request": request, "owner": owner, "active_page": "map",
            "yandex_maps_api_key": settings.yandex_maps_api_key,
        },
    )


async def _driver_positions(session: AsyncSession, owner_id: int) -> list[dict]:
    """
    Последняя известная позиция каждого водителя владельца + признак активной
    смены и гос. номер машины в смене.

    Источник координат водителя — ручные геопозиции (события `location_sent`).
    GPS машин Stavtrack отдаётся отдельно из VehicleState в API ниже. Набор
    ключей в словаре — это контракт для `map.html`, его не меняем.
    """
    # Postgres DISTINCT ON через raw — проще, чем рисовать в SQLAlchemy
    from sqlalchemy import text as sa_text
    rows = await session.execute(
        sa_text("""
            SELECT DISTINCT ON (events.driver_id)
                events.driver_id,
                events.payload->>'lat' AS lat,
                events.payload->>'lon' AS lon,
                events.created_at,
                drivers.full_name,
                drivers.phone,
                CASE WHEN EXISTS (
                    SELECT 1 FROM shifts
                    WHERE shifts.driver_id = events.driver_id
                      AND shifts.status = 'started'
                ) THEN true ELSE false END AS active_shift,
                (SELECT vehicles.license_plate
                   FROM shifts
                   JOIN vehicles ON vehicles.id = shifts.vehicle_id
                  WHERE shifts.driver_id = events.driver_id
                    AND shifts.status = 'started'
                  LIMIT 1) AS plate
            FROM events
            JOIN drivers ON drivers.id = events.driver_id
            WHERE events.owner_id = :owner_id
              AND events.event_type = 'location_sent'
              AND events.payload ? 'lat'
            ORDER BY events.driver_id, events.created_at DESC
        """),
        {"owner_id": owner_id},
    )
    result = []
    for row in rows.mappings().all():
        try:
            lat = float(row["lat"])
            lon = float(row["lon"])
        except (TypeError, ValueError):
            continue
        # инициалы для маркера
        name = row["full_name"] or ""
        initials = "".join(w[0].upper() for w in name.split()[:2]) or "?"
        result.append({
            "driver_id": row["driver_id"],
            "name": name,
            "phone": row["phone"] or "",
            "plate": row["plate"] or "",
            "initials": initials,
            "lat": lat,
            "lon": lon,
            "active_shift": bool(row["active_shift"]),
            "updated_at": row["created_at"].isoformat() if row["created_at"] else None,
        })
    return result


@app.get("/api/drivers-locations")
async def api_drivers_locations(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Для карты: ручные координаты водителей + GPS машин (Stavtrack)."""
    vehicles_res = await session.execute(
        select(VehicleState, Vehicle.license_plate)
        .join(Vehicle, Vehicle.id == VehicleState.vehicle_id)
        .where(
            Vehicle.owner_id == owner.id,
            Vehicle.is_active.is_(True),
            VehicleState.latitude.is_not(None),
            VehicleState.longitude.is_not(None),
        )
    )
    # машины с открытой сменой — для балуна «в смене / без смены»
    busy_res = await session.execute(
        select(Shift.vehicle_id).where(
            Shift.owner_id == owner.id, Shift.status == "started"
        )
    )
    busy_vehicle_ids = {row[0] for row in busy_res.all()}
    vehicles = [
        {
            "vehicle_id": st.vehicle_id,
            "plate": plate,
            "lat": float(st.latitude),
            "lon": float(st.longitude),
            "speed_kmh": float(st.speed_kmh or 0),
            "ignition": st.ignition,
            "ignition_known": st.ignition is not None,
            "motion_status": st.motion_status,
            "motion_status_text": telemetry_service.motion_status_text(st.motion_status, st.speed_kmh),
            "motion_since_at": st.motion_since_at.isoformat() if st.motion_since_at else None,
            # «с 21:52» / «со вчера, 21:52» / «с 01.07, 21:52»
            "motion_since_label": smart_since_label(st.motion_since_at, owner.timezone),
            "motion_duration_label": telemetry_service.duration_label(st.motion_since_at),
            "has_active_shift": st.vehicle_id in busy_vehicle_ids,
            "is_valid": st.is_valid,
            "updated_at": st.last_seen_at.isoformat() if st.last_seen_at else None,
            "updated_label": fmt_dt(st.last_seen_at, owner.timezone, "%d.%m, %H:%M"),
        }
        for st, plate in vehicles_res.all()
        # «нулевой остров» (потеря GPS у трекера) на карту не выносим
        if abs(float(st.latitude)) > 0.001 or abs(float(st.longitude)) > 0.001
    ]
    # РЦ с координатами — статичные синие точки (границы геозон)
    rcs_res = await session.execute(
        select(DistributionCenter).where(
            DistributionCenter.owner_id == owner.id,
            DistributionCenter.is_active.is_(True),
            DistributionCenter.latitude.is_not(None),
            DistributionCenter.longitude.is_not(None),
        )
    )
    rcs = [
        {
            "id": rc.id,
            "name": rc.name,
            "address": rc.address,
            "lat": float(rc.latitude),
            "lon": float(rc.longitude),
        }
        for rc in rcs_res.scalars().all()
    ]
    return {
        "drivers": await _driver_positions(session, owner.id),
        "vehicles": vehicles,
        "rcs": rcs,
    }


# =========================================================================
# /api/photo/{file_id} — прокси для просмотра фото из Telegram в кабинете
# =========================================================================
async def _owner_owns_photo(session: AsyncSession, owner_id: int, file_id: str) -> bool:
    """Проверяем что фото принадлежит владельцу (есть в trips/expenses/shifts)."""
    in_trips = await session.execute(
        select(func.count(Trip.id)).where(
            Trip.owner_id == owner_id, Trip.waybill_photo_url == file_id
        )
    )
    if (in_trips.scalar_one() or 0) > 0:
        return True
    in_expenses = await session.execute(
        select(func.count(Expense.id)).where(
            Expense.owner_id == owner_id, Expense.receipt_photo_url == file_id
        )
    )
    if (in_expenses.scalar_one() or 0) > 0:
        return True
    in_shifts = await session.execute(
        select(func.count(Shift.id)).where(
            Shift.owner_id == owner_id,
            (Shift.odometer_start_photo_url == file_id)
            | (Shift.odometer_end_photo_url == file_id),
        )
    )
    return (in_shifts.scalar_one() or 0) > 0


@app.get("/api/photo/{file_id}")
async def api_photo(
    request: Request,
    file_id: str,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    if not await _owner_owns_photo(session, owner.id, file_id):
        raise HTTPException(status_code=403, detail="Forbidden")
    driver_bot = request.app.state.driver_bot
    try:
        buf = await driver_bot.download(file_id)
    except Exception:
        raise HTTPException(status_code=404, detail="Not found")
    if buf is None:
        raise HTTPException(status_code=404)
    return Response(content=buf.read(), media_type="image/jpeg")


# =========================================================================
# /trips/{id} — детали рейса с фото ТТН и связанными расходами
# =========================================================================
def _web_date_to_utc(value: str, tz_name: str | None) -> datetime:
    """Дата из <input type=date> → UTC-datetime (полдень местного времени)."""
    tz = owner_tz(tz_name)
    try:
        d = date.fromisoformat(value) if value else datetime.now(tz).date()
    except ValueError:
        d = datetime.now(tz).date()
    return datetime(d.year, d.month, d.day, 12, 0, tzinfo=tz).astimezone(timezone.utc)


@app.post("/trips/add")
async def create_trip_manual(
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    driver_id: Annotated[str, Form()],
    vehicle_id: Annotated[str, Form()],
    origin: Annotated[str, Form()],
    destination: Annotated[str, Form()],
    revenue_rub: Annotated[str, Form()] = "",
    trip_date: Annotated[str, Form()] = "",
):
    """Владелец добавляет рейс с сайта (Блок: форма владельца). Рейс ручной —
    живёт в ручной (завершённой) смене, чтобы FK shift_id был валиден; помечается
    is_manual («вручную, км неизвестен»)."""
    try:
        d_id, v_id = int(driver_id), int(vehicle_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Bad driver/vehicle")
    driver = await session.get(Driver, d_id)
    vehicle = await session.get(Vehicle, v_id)
    if driver is None or driver.owner_id != owner.id or vehicle is None or vehicle.owner_id != owner.id:
        raise HTTPException(status_code=400, detail="Bad driver/vehicle")
    if not origin.strip() or not destination.strip():
        raise HTTPException(status_code=400, detail="Need route")

    revenue = None
    if revenue_rub.strip():
        try:
            revenue = Decimal(revenue_rub.replace(",", ".").replace(" ", ""))
            if revenue < 0:
                raise InvalidOperation
        except InvalidOperation:
            raise HTTPException(status_code=400, detail="Bad revenue")

    dt = _web_date_to_utc(trip_date, owner.timezone)
    shift = Shift(
        owner_id=owner.id, driver_id=d_id, vehicle_id=v_id,
        status="completed", started_at=dt, ended_at=dt, is_manual=True,
    )
    session.add(shift)
    await session.flush()
    trip = Trip(
        owner_id=owner.id, shift_id=shift.id, driver_id=d_id, vehicle_id=v_id,
        status="completed", origin=origin.strip(), destination=destination.strip(),
        completed_at=dt, is_manual=True,
        revenue_rub=(revenue.quantize(Decimal("0.01")) if revenue is not None else None),
    )
    session.add(trip)
    await session.commit()
    return RedirectResponse("/trips", status_code=303)


@app.post("/trips/{trip_id}/revenue")
async def update_trip_revenue(
    trip_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    revenue_rub: Annotated[str, Form()],
):
    """Владелец правит выручку рейса на сайте (Блок G1). Прибыль (profit_rub) —
    вычисляемая колонка Postgres, пересчитается сама; финитоги и графики берут
    revenue_rub из тех же рейсов, поэтому обновятся автоматически."""
    trip = await session.get(Trip, trip_id)
    if trip is None or trip.owner_id != owner.id:
        raise HTTPException(status_code=404)
    try:
        rev = Decimal(revenue_rub.replace(",", ".").replace(" ", ""))
        if rev < 0:
            raise InvalidOperation
    except InvalidOperation:
        raise HTTPException(status_code=400, detail="Bad revenue")
    trip.revenue_rub = rev.quantize(Decimal("0.01"))
    trip.driver_revenue_pending_rub = None
    await session.commit()
    return RedirectResponse(f"/trips/{trip_id}", status_code=303)


_MAX_DOC_BYTES = 6 * 1024 * 1024  # 6 МБ на документ


@app.post("/trips/{trip_id}/document")
async def upload_trip_document(
    trip_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    file: Annotated[UploadFile, File()],
):
    """Владелец загружает фото/скан документа к рейсу прямо на сайте. Байты
    кладём в Postgres (без S3). Лимит 6 МБ."""
    trip = await session.get(Trip, trip_id)
    if trip is None or trip.owner_id != owner.id:
        raise HTTPException(status_code=404)
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Пустой файл")
    if len(data) > _MAX_DOC_BYTES:
        raise HTTPException(status_code=400, detail="Файл слишком большой (макс 6 МБ)")
    doc = TripDocument(
        trip_id=trip.id, owner_id=owner.id,
        filename=file.filename,
        content_type=file.content_type or "application/octet-stream",
        data=data,
    )
    session.add(doc)
    await session.commit()
    return RedirectResponse(f"/trips/{trip_id}", status_code=303)


@app.get("/api/trip-doc/{doc_id}")
async def get_trip_document(
    doc_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    doc = await session.get(TripDocument, doc_id)
    if doc is None or doc.owner_id != owner.id:
        raise HTTPException(status_code=404)
    return Response(content=doc.data, media_type=doc.content_type)


@app.post("/trip-doc/{doc_id}/delete")
async def delete_trip_document(
    doc_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Удалить документ рейса, загруженный владельцем."""
    doc = await session.get(TripDocument, doc_id)
    if doc is None or doc.owner_id != owner.id:
        raise HTTPException(status_code=404)
    trip_id = doc.trip_id
    await session.delete(doc)
    await session.commit()
    return RedirectResponse(f"/trips/{trip_id}", status_code=303)


@app.post("/trips/{trip_id}/waybill/delete")
async def delete_trip_waybill(
    trip_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Удалить фото ТТН, присланное водителем (Telegram file_id обнуляем)."""
    trip = await session.get(Trip, trip_id)
    if trip is None or trip.owner_id != owner.id:
        raise HTTPException(status_code=404)
    trip.waybill_photo_url = None
    await session.commit()
    return RedirectResponse(f"/trips/{trip_id}", status_code=303)


async def _route_travel_estimate(
    session: AsyncSession, owner_id: int, origin: str | None, destination: str | None
) -> dict | None:
    """Среднее время в пути по маршруту origin→destination из истории завершённых
    рейсов. Длительность = завершение − выезд (событие trip_in_transit, иначе
    создание рейса). Возвращает {minutes, count} или None, если данных нет."""
    if not origin or not destination:
        return None
    trips = (
        await session.execute(
            select(Trip.id, Trip.created_at, Trip.completed_at)
            .where(
                Trip.owner_id == owner_id,
                Trip.origin == origin,
                Trip.destination == destination,
                Trip.status == "completed",
                Trip.completed_at.is_not(None),
            )
            .order_by(desc(Trip.completed_at))
            .limit(50)
        )
    ).all()
    durations = []
    for tid, created, completed in trips:
        dep = (
            await session.execute(
                select(Event.created_at)
                .where(Event.trip_id == tid, Event.event_type == "trip_in_transit")
                .order_by(Event.created_at)
                .limit(1)
            )
        ).scalar_one_or_none() or created
        if dep and completed and completed > dep:
            durations.append((completed - dep).total_seconds())
    if not durations:
        return None
    avg_min = int(sum(durations) / len(durations) / 60)
    return {"minutes": avg_min, "count": len(durations)}


def _money_label(value) -> str:
    return telemetry_service.rub_label(value or 0)


def _event_payload_amount(payload: dict | None, key: str) -> str | None:
    payload = payload or {}
    raw = payload.get(key)
    if raw in (None, ""):
        return None
    try:
        return _money_label(Decimal(str(raw)))
    except (InvalidOperation, ValueError):
        return None


def _timeline_item(
    *,
    kind: str,
    order: int,
    at: datetime | None,
    title: str,
    subtitle: str | None = None,
    meta: str | None = None,
    amount: str | None = None,
    action: str | None = None,
) -> dict:
    return {
        "kind": kind,
        "order": order,
        "at": at,
        "title": title,
        "subtitle": subtitle,
        "meta": meta,
        "amount": amount,
        "action": action,
        "_sort_at": at or datetime.min.replace(tzinfo=timezone.utc),
    }


async def _trip_timeline(
    session: AsyncSession,
    owner: Owner,
    trip: Trip,
    shift: Shift | None,
    driver: Driver | None,
    vehicle: Vehicle | None,
    waybill_uploaded_at: datetime | None,
    documents: list[dict],
) -> list[dict]:
    """Собрать понятную владельцу историю рейса из событий и GPS-геозон."""
    plate = vehicle.license_plate if vehicle else "—"
    driver_name = driver.full_name if driver else "—"
    route = f"{trip.origin or '—'} → {trip.destination or '—'}"
    items: list[dict] = []
    seen: set[tuple[str, str, str]] = set()

    def add(item: dict) -> None:
        at = item.get("at")
        key = (
            item["title"],
            at.isoformat() if isinstance(at, datetime) else "",
            item.get("subtitle") or "",
        )
        if key in seen:
            return
        seen.add(key)
        items.append(item)

    if shift is not None:
        add(_timeline_item(
            kind="info", order=10, at=shift.started_at, title="Смена открыта",
            subtitle=f"{plate} · {driver_name}",
            meta=(f"одометр {shift.odometer_start} км" if shift.odometer_start else None),
        ))

    add(_timeline_item(
        kind="info", order=20, at=trip.created_at,
        title="Рейс создан" if not trip.is_manual else "Рейс добавлен вручную",
        subtitle=route,
        meta=trip.cargo_name or ("ручной рейс · км неизвестен" if trip.is_manual else None),
    ))

    event_rows = (
        await session.execute(
            select(Event.event_type, Event.created_at, Event.payload)
            .where(
                Event.owner_id == owner.id,
                or_(
                    Event.trip_id == trip.id,
                    and_(
                        Event.shift_id == trip.shift_id,
                        Event.event_type.in_(("shift_started", "shift_reassigned")),
                    ),
                ),
            )
            .order_by(Event.created_at)
        )
    ).all()
    saw_completed_event = False
    for event_type, created_at, payload in event_rows:
        payload = payload or {}
        if event_type == "trip_in_transit":
            add(_timeline_item(
                kind="info", order=30, at=created_at, title="Выехал по маршруту",
                subtitle=route,
            ))
        elif event_type == "trip_unloading":
            add(_timeline_item(
                kind="warn", order=50, at=created_at, title="На выгрузке",
                subtitle=f"{trip.destination or '—'} · водитель отметил сдачу груза",
            ))
        elif event_type == "waybill_uploaded":
            add(_timeline_item(
                kind="success", order=70, at=created_at, title="ТТН загружена",
                subtitle="фото от водителя",
            ))
        elif event_type == "trip_completed":
            saw_completed_event = True
            fuel = _event_payload_amount(payload, "fuel_cost")
            add(_timeline_item(
                kind="success", order=80, at=created_at, title="Рейс завершён",
                subtitle=route,
                meta=f"топливо {fuel}" if fuel else None,
            ))
        elif event_type == "trip_revenue_from_driver":
            amount = _event_payload_amount(payload, "revenue")
            pending = trip.revenue_rub is None
            add(_timeline_item(
                # Пока владелец не решил — жёлтое «ждёт подтверждения». После
                # решения это просто историческая отметка «водитель предложил».
                kind="warn" if pending else "info",
                order=90, at=created_at,
                title=(
                    f"Выручка {amount or '—'} — ждёт подтверждения" if pending
                    else f"Водитель предложил выручку: {amount or '—'}"
                ),
                subtitle=(
                    "указал водитель · проверьте документы и подтвердите сумму" if pending
                    else "предложение водителя · владелец уже принял решение ниже"
                ),
                action="confirm_revenue" if pending else None,
            ))
        elif event_type == "trip_revenue_approved":
            amount = _event_payload_amount(payload, "revenue")
            add(_timeline_item(
                kind="success", order=92, at=created_at,
                title=f"Выручка подтверждена{f': {amount}' if amount else ''}",
                subtitle="сумма вошла в финансы рейса",
            ))
        elif event_type == "trip_revenue_set":
            amount = _event_payload_amount(payload, "revenue")
            add(_timeline_item(
                kind="success", order=92, at=created_at,
                title=f"Выручка изменена владельцем{f': {amount}' if amount else ''}",
                subtitle="сумма вошла в финансы рейса",
            ))
        elif event_type == "expense_submitted":
            amount = _event_payload_amount(payload, "amount")
            add(_timeline_item(
                kind="warn", order=74, at=created_at,
                title=f"Расход отправлен{f': {amount}' if amount else ''}",
                subtitle=payload.get("category") or "чек/расход от водителя",
            ))
        elif event_type == "trip_rc_confirmed":
            add(_timeline_item(
                kind="success", order=44, at=created_at,
                title="✅ GPS подтвердил прибытие на РЦ",
                subtitle=payload.get("rc_name") or trip.destination or "—",
            ))
        elif event_type == "trip_rc_mismatch":
            add(_timeline_item(
                kind="danger", order=44, at=created_at,
                title="⚠️ Приехал не на тот РЦ",
                subtitle=(
                    f"план: {payload.get('planned_rc_name') or '—'} · "
                    f"факт: {payload.get('actual_rc_name') or '—'}"
                ),
            ))

    # Окно GPS-событий = только время ЭТОГО рейса, иначе в карточку попадают
    # визиты на РЦ из других поездок за день (владелец видел «уехал в 17:30»
    # у рейса, начатого в 20:54). Небольшие поля: до старта на выезд, после
    # завершения — на выгрузку/отъезд с РЦ.
    _now = datetime.now(timezone.utc)
    trip_start_ref = trip.created_at or (shift.started_at if shift else None) or _now
    if shift is not None and shift.started_at is not None and shift.started_at < trip_start_ref:
        trip_start_ref = shift.started_at
    window_start = trip_start_ref - timedelta(minutes=30)
    window_end = (trip.completed_at or _now) + timedelta(hours=2)
    if vehicle is not None:
        rc_rows = (
            await session.execute(
                select(Event.event_type, Event.created_at, Event.payload)
                .where(
                    Event.owner_id == owner.id,
                    Event.event_type.in_(("rc_arrived", "rc_departed", "rc_downtime_alert")),
                    Event.created_at >= window_start,
                    Event.created_at <= window_end,
                )
                .order_by(Event.created_at)
            )
        ).all()
        for event_type, created_at, payload in rc_rows:
            payload = payload or {}
            if telemetry_service.int_or_none(payload.get("vehicle_id")) != vehicle.id:
                continue
            rc_name = payload.get("rc_name") or "РЦ"
            if event_type == "rc_arrived":
                parked_since = _payload_dt(payload.get("parked_since")) or created_at
                add(_timeline_item(
                    kind="warn", order=42, at=parked_since,
                    title="Приехал на РЦ",
                    subtitle=rc_name,
                    meta=f"начало стоянки {fmt_dt(parked_since, owner.timezone, '%d.%m %H:%M')}",
                ))
            elif event_type == "rc_downtime_alert":
                waited = telemetry_service.int_or_none(payload.get("waited_minutes")) or 0
                billable = telemetry_service.int_or_none(payload.get("suggested_amount_rub")) or telemetry_service.rc_billable_downtime_rub(waited)
                blocks = billable // telemetry_service.RC_BILLABLE_DOWNTIME_RUB if billable else 0
                add(_timeline_item(
                    kind="danger", order=58, at=created_at,
                    title="Простой на РЦ перешёл платный порог",
                    subtitle=f"{rc_name} · стоял {_minutes_label(waited)}",
                    meta=(f"{blocks} блок(а) по 12 часов" if blocks else None),
                    amount=_money_label(billable) if billable else None,
                ))
            elif event_type == "rc_departed":
                waited = telemetry_service.int_or_none(payload.get("waited_minutes")) or 0
                engine_off = telemetry_service.int_or_none(payload.get("engine_off_minutes"))
                payload_billable = telemetry_service.int_or_none(payload.get("billable_downtime_rub"))
                billable = max(telemetry_service.rc_billable_downtime_rub(waited), payload_billable or 0)
                blocks = billable // telemetry_service.RC_BILLABLE_DOWNTIME_RUB if billable else 0
                arrived_at = _payload_dt(payload.get("arrived_at"))
                departed_at = _payload_dt(payload.get("departed_at")) or created_at
                meta_parts = []
                if arrived_at:
                    meta_parts.append(f"заехал {fmt_dt(arrived_at, owner.timezone, '%d.%m %H:%M')}")
                meta_parts.append(f"уехал {fmt_dt(departed_at, owner.timezone, '%d.%m %H:%M')}")
                if engine_off is not None:
                    meta_parts.append(f"мотор заглушен {_minutes_label(engine_off)}")
                if blocks:
                    meta_parts.append(f"{blocks} блок(а) по 12 часов")
                add(_timeline_item(
                    kind="danger" if billable else "warn",
                    order=60,
                    at=departed_at,
                    title=f"Простой на РЦ {rc_name}: {_minutes_label(waited)}",
                    subtitle=" · ".join(meta_parts),
                    amount=_money_label(billable) if billable else None,
                ))

    if waybill_uploaded_at:
        add(_timeline_item(
            kind="success", order=70, at=waybill_uploaded_at, title="ТТН загружена",
            subtitle="фото от водителя",
        ))

    for doc in documents:
        add(_timeline_item(
            kind="success", order=76, at=doc.get("uploaded_at"),
            title="Документ добавлен владельцем",
            subtitle=doc.get("filename") or "документ",
        ))

    if trip.completed_at and not saw_completed_event:
        # синтетическую строку добавляем ТОЛЬКО если не было события
        # trip_completed (иначе «Рейс завершён» дублировался)
        add(_timeline_item(
            kind="success", order=80, at=trip.completed_at,
            title="Рейс завершён", subtitle=route,
        ))
    elif not trip.completed_at:
        add(_timeline_item(
            kind="info", order=95, at=datetime.now(timezone.utc),
            title="Рейс ещё открыт",
            subtitle="пока водитель не завершил рейс, финальные цифры могут меняться",
        ))

    if trip.driver_revenue_pending_rub is not None and trip.revenue_rub is None:
        add(_timeline_item(
            kind="warn", order=90, at=trip.completed_at or datetime.now(timezone.utc),
            title=f"Выручка {_money_label(trip.driver_revenue_pending_rub)} — ждёт подтверждения",
            subtitle="указал водитель · нажмите «Подтвердить» или измените сумму",
            action="confirm_revenue",
        ))
    elif trip.revenue_rub is not None:
        add(_timeline_item(
            kind="success", order=92, at=trip.completed_at or trip.created_at,
            title=f"Выручка подтверждена: {_money_label(trip.revenue_rub)}",
            subtitle="сумма учитывается в финансах",
        ))

    items.sort(key=lambda x: (x["_sort_at"], x["order"]))
    for item in items:
        item.pop("_sort_at", None)
    return items


async def _vehicle_now_info(session: AsyncSession, owner: Owner, vehicle: Vehicle | None):
    """Где машина СЕЙЧАС по GPS (для карточки рейса — отдельно от истории рейса).
    None, если у машины нет свежего GPS-состояния."""
    if vehicle is None:
        return None
    st = (
        await session.execute(
            select(VehicleState).where(VehicleState.vehicle_id == vehicle.id)
        )
    ).scalar_one_or_none()
    if st is None or st.last_seen_at is None:
        return None
    now = datetime.now(timezone.utc)
    last_seen = st.last_seen_at
    if last_seen.tzinfo is None:
        last_seen = last_seen.replace(tzinfo=timezone.utc)
    fresh = last_seen >= now - timedelta(minutes=30)
    status_text = telemetry_service.motion_status_text(st.motion_status, st.speed_kmh)
    # на каком РЦ машина сейчас (последний rc_arrived без более позднего rc_departed)
    at_rc = None
    ev_rows = (
        await session.execute(
            select(Event.event_type, Event.created_at, Event.payload)
            .where(
                Event.owner_id == owner.id,
                Event.event_type.in_(("rc_arrived", "rc_departed")),
                Event.created_at >= now - timedelta(days=2),
            )
            .order_by(desc(Event.created_at))
        )
    ).all()
    for et, _cat, pl in ev_rows:
        pl = pl or {}
        if telemetry_service.int_or_none(pl.get("vehicle_id")) != vehicle.id:
            continue
        at_rc = pl.get("rc_name") if et == "rc_arrived" else None
        break  # берём самое свежее событие по этой машине
    return {
        "status_text": status_text,
        "since_label": smart_since_label(st.motion_since_at, owner.timezone),
        "last_seen_label": fmt_dt(st.last_seen_at, owner.timezone, "%d.%m %H:%M"),
        "fresh": fresh,
        "at_rc": at_rc,
    }


@app.get("/trips/{trip_id}", response_class=HTMLResponse)
async def trip_detail(
    request: Request,
    trip_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    trip = await session.get(Trip, trip_id)
    if trip is None or trip.owner_id != owner.id:
        raise HTTPException(status_code=404)
    driver = await session.get(Driver, trip.driver_id)
    vehicle = await session.get(Vehicle, trip.vehicle_id)
    shift = await session.get(Shift, trip.shift_id)
    expenses_res = await session.execute(
        select(Expense).where(Expense.trip_id == trip.id).order_by(Expense.created_at)
    )
    expenses = list(expenses_res.scalars().all())
    # время когда водитель загрузил ТТН (берём последнее событие)
    waybill_uploaded_at = (
        await session.execute(
            select(Event.created_at).where(
                Event.trip_id == trip.id,
                Event.event_type == "waybill_uploaded",
            ).order_by(desc(Event.created_at)).limit(1)
        )
    ).scalar_one_or_none()
    # документы, загруженные владельцем на сайте (метаданные без байтов)
    docs_res = await session.execute(
        select(TripDocument.id, TripDocument.filename, TripDocument.content_type, TripDocument.uploaded_at)
        .where(TripDocument.trip_id == trip.id)
        .order_by(desc(TripDocument.uploaded_at))
    )
    documents = [
        {"id": did, "filename": fn, "content_type": ct, "uploaded_at": ua}
        for did, fn, ct, ua in docs_res.all()
    ]
    travel = await _route_travel_estimate(session, owner.id, trip.origin, trip.destination)
    travel_label = None
    if travel:
        h, m = divmod(travel["minutes"], 60)
        travel_label = (f"{h} ч " if h else "") + f"{m} мин"
    vehicle_now = await _vehicle_now_info(session, owner, vehicle)
    timeline = await _trip_timeline(
        session, owner, trip, shift, driver, vehicle, waybill_uploaded_at, documents
    )
    trip_duration_label = _minutes_label(
        _minutes_between(trip.created_at, trip.completed_at or datetime.now(timezone.utc))
    )
    all_drivers, all_vehicles = await _reassign_options(session, owner.id)
    return templates.TemplateResponse(
        "trip_detail.html",
        {
            "request": request, "owner": owner,
            "trip": trip, "shift": shift, "driver": driver, "vehicle": vehicle,
            "expenses": expenses,
            "waybill_uploaded_at": waybill_uploaded_at,
            "documents": documents,
            "travel": travel, "travel_label": travel_label,
            "vehicle_now": vehicle_now,
            "timeline": timeline,
            "trip_duration_label": trip_duration_label,
            "all_drivers": all_drivers, "all_vehicles": all_vehicles,
            "active_page": "trips",
        },
    )


@app.post("/trips/{trip_id}/delete")
async def trip_delete(
    trip_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Безопасно удалить рейс из рабочего контура владельца.

    Расходы и события оставляем в истории, но отвязываем от рейса: так база не
    падает на внешних ключах, а финансовые ручные записи не исчезают бесследно.
    """
    trip = await session.get(Trip, trip_id)
    if trip is None or trip.owner_id != owner.id:
        raise HTTPException(status_code=404)

    await session.execute(
        update(Expense)
        .where(Expense.owner_id == owner.id, Expense.trip_id == trip.id)
        .values(trip_id=None)
    )
    await session.execute(
        update(Event)
        .where(Event.owner_id == owner.id, Event.trip_id == trip.id)
        .values(trip_id=None)
    )
    await session.execute(
        delete(TripDocument).where(
            TripDocument.owner_id == owner.id,
            TripDocument.trip_id == trip.id,
        )
    )
    await session.delete(trip)
    await session.commit()
    return RedirectResponse("/trips?deleted=1", status_code=303)


async def _reassign_options(
    session: AsyncSession, owner_id: int
) -> tuple[list[Driver], list[Vehicle]]:
    """Активные водители и машины — для селектов «исправить миссклик»."""
    drivers = list(
        (
            await session.execute(
                select(Driver)
                .where(Driver.owner_id == owner_id, Driver.is_active.is_(True))
                .order_by(Driver.full_name)
            )
        ).scalars().all()
    )
    vehicles = list(
        (
            await session.execute(
                select(Vehicle)
                .where(Vehicle.owner_id == owner_id, Vehicle.is_active.is_(True))
                .order_by(Vehicle.license_plate)
            )
        ).scalars().all()
    )
    return drivers, vehicles


# =========================================================================
# /shifts — список смен с фото одометров
# =========================================================================
@app.get("/shifts", response_class=HTMLResponse)
async def shifts_page(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    rows_res = await session.execute(
        select(Shift, Driver.full_name, Vehicle.license_plate)
        .join(Driver, Driver.id == Shift.driver_id)
        .join(Vehicle, Vehicle.id == Shift.vehicle_id)
        .where(Shift.owner_id == owner.id)
        .order_by(desc(Shift.started_at))
        .limit(200)
    )
    rows = list(rows_res.all())
    return templates.TemplateResponse(
        "shifts.html",
        {"request": request, "owner": owner, "rows": rows, "active_page": "trips"},
    )


@app.get("/shifts/{shift_id}", response_class=HTMLResponse)
async def shift_detail(
    request: Request,
    shift_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    shift = await session.get(Shift, shift_id)
    if shift is None or shift.owner_id != owner.id:
        raise HTTPException(status_code=404)
    driver = await session.get(Driver, shift.driver_id)
    vehicle = await session.get(Vehicle, shift.vehicle_id)
    trips_res = await session.execute(
        select(Trip).where(Trip.shift_id == shift.id).order_by(Trip.created_at)
    )
    trips = list(trips_res.scalars().all())
    expenses_res = await session.execute(
        select(Expense).where(Expense.shift_id == shift.id).order_by(Expense.created_at)
    )
    expenses = list(expenses_res.scalars().all())
    # время фото одометров: shift_started → начало, shift_completed → конец
    events_times = await session.execute(
        select(Event.event_type, Event.created_at).where(
            Event.shift_id == shift.id,
            Event.event_type.in_(("shift_started", "shift_completed")),
        )
    )
    times = {et: dt for et, dt in events_times.all()}
    all_drivers, all_vehicles = await _reassign_options(session, owner.id)
    return templates.TemplateResponse(
        "shift_detail.html",
        {
            "request": request, "owner": owner,
            "shift": shift, "driver": driver, "vehicle": vehicle,
            "trips": trips, "expenses": expenses,
            "photo_start_at": times.get("shift_started"),
            "photo_end_at": times.get("shift_completed"),
            "all_drivers": all_drivers, "all_vehicles": all_vehicles,
            "active_page": "trips",
        },
    )


@app.post("/shifts/{shift_id}/delete")
async def shift_delete(
    shift_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Удалить смену вместе с её рейсами, не теряя ручные расходы и историю."""
    shift = await session.get(Shift, shift_id)
    if shift is None or shift.owner_id != owner.id:
        raise HTTPException(status_code=404)

    trip_ids = list(
        (
            await session.execute(
                select(Trip.id).where(Trip.owner_id == owner.id, Trip.shift_id == shift.id)
            )
        ).scalars().all()
    )
    if trip_ids:
        await session.execute(
            update(Expense)
            .where(Expense.owner_id == owner.id, Expense.trip_id.in_(trip_ids))
            .values(trip_id=None)
        )
        await session.execute(
            update(Event)
            .where(Event.owner_id == owner.id, Event.trip_id.in_(trip_ids))
            .values(trip_id=None)
        )
        await session.execute(
            delete(TripDocument).where(
                TripDocument.owner_id == owner.id,
                TripDocument.trip_id.in_(trip_ids),
            )
        )
        await session.execute(
            delete(Trip).where(Trip.owner_id == owner.id, Trip.id.in_(trip_ids))
        )

    await session.execute(
        update(Expense)
        .where(Expense.owner_id == owner.id, Expense.shift_id == shift.id)
        .values(shift_id=None)
    )
    await session.execute(
        update(Event)
        .where(Event.owner_id == owner.id, Event.shift_id == shift.id)
        .values(shift_id=None)
    )
    await session.delete(shift)
    await session.commit()
    return RedirectResponse("/shifts?deleted=1", status_code=303)


# =========================================================================
# Исправление миссклика: поменять водителя/машину у смены и рейса
# =========================================================================
async def _reassign_targets(
    session: AsyncSession, owner: Owner, driver_id: str, vehicle_id: str
) -> tuple[Driver, Vehicle]:
    """Проверить, что выбранные водитель и машина существуют и принадлежат владельцу."""
    try:
        d_id, v_id = int(driver_id), int(vehicle_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Bad driver/vehicle id")
    driver = await session.get(Driver, d_id)
    vehicle = await session.get(Vehicle, v_id)
    if (
        driver is None or driver.owner_id != owner.id
        or vehicle is None or vehicle.owner_id != owner.id
    ):
        raise HTTPException(status_code=400, detail="Bad driver/vehicle id")
    return driver, vehicle


@app.post("/shifts/{shift_id}/reassign")
async def shift_reassign(
    shift_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    driver_id: Annotated[str, Form()],
    vehicle_id: Annotated[str, Form()],
):
    """Поменять водителя/машину у смены задним числом (исправить миссклик).
    Рейсы смены обновляются каскадно; зарплата пересчитывается на лету."""
    shift = await session.get(Shift, shift_id)
    if shift is None or shift.owner_id != owner.id:
        raise HTTPException(status_code=404)
    driver, vehicle = await _reassign_targets(session, owner, driver_id, vehicle_id)
    if driver.id == shift.driver_id and vehicle.id == shift.vehicle_id:
        return RedirectResponse(f"/shifts/{shift.id}", status_code=303)

    # активную смену нельзя перевесить на водителя/машину из другой активной смены
    if shift.status == "started":
        if driver.id != shift.driver_id:
            busy = (
                await session.execute(
                    select(Shift.id).where(
                        Shift.driver_id == driver.id,
                        Shift.status == "started",
                        Shift.id != shift.id,
                    )
                )
            ).scalar_one_or_none()
            if busy is not None:
                return RedirectResponse(f"/shifts/{shift.id}?err=driver_busy", status_code=303)
        if vehicle.id != shift.vehicle_id:
            busy = (
                await session.execute(
                    select(Shift.id).where(
                        Shift.vehicle_id == vehicle.id,
                        Shift.status == "started",
                        Shift.id != shift.id,
                    )
                )
            ).scalar_one_or_none()
            if busy is not None:
                return RedirectResponse(f"/shifts/{shift.id}?err=vehicle_busy", status_code=303)

    old = {"driver_id": shift.driver_id, "vehicle_id": shift.vehicle_id}
    shift.driver_id = driver.id
    shift.vehicle_id = vehicle.id
    # каскад: рейсы этой смены переезжают на нового водителя/машину
    await session.execute(
        update(Trip)
        .where(Trip.shift_id == shift.id)
        .values(driver_id=driver.id, vehicle_id=vehicle.id)
    )
    await log_event(
        session, owner_id=owner.id, driver_id=driver.id, shift_id=shift.id,
        event_type="shift_reassigned",
        payload={
            "old": old,
            "new": {"driver_id": driver.id, "vehicle_id": vehicle.id},
            "source": "web",
        },
    )
    await session.commit()
    return RedirectResponse(f"/shifts/{shift.id}?saved=1", status_code=303)


@app.post("/trips/{trip_id}/reassign")
async def trip_reassign(
    trip_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    driver_id: Annotated[str, Form()],
    vehicle_id: Annotated[str, Form()],
):
    """Поменять водителя/машину у ОДНОГО рейса (смена не трогается)."""
    trip = await session.get(Trip, trip_id)
    if trip is None or trip.owner_id != owner.id:
        raise HTTPException(status_code=404)
    driver, vehicle = await _reassign_targets(session, owner, driver_id, vehicle_id)
    if driver.id == trip.driver_id and vehicle.id == trip.vehicle_id:
        return RedirectResponse(f"/trips/{trip.id}", status_code=303)
    old = {"driver_id": trip.driver_id, "vehicle_id": trip.vehicle_id}
    trip.driver_id = driver.id
    trip.vehicle_id = vehicle.id
    await log_event(
        session, owner_id=owner.id, driver_id=driver.id, trip_id=trip.id,
        event_type="trip_reassigned",
        payload={
            "old": old,
            "new": {"driver_id": driver.id, "vehicle_id": vehicle.id},
            "source": "web",
        },
    )
    await session.commit()
    return RedirectResponse(f"/trips/{trip.id}?saved=1", status_code=303)


# =========================================================================
# /expenses — все расходы любых категорий с фото чеков и фильтром
# =========================================================================
_EXPENSE_CATEGORIES = ("fuel", "repair", "parking", "fine", "toll", "other")


@app.get("/expenses", response_class=HTMLResponse)
async def expenses_page(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    category: Annotated[str | None, Query()] = None,
    status: Annotated[str | None, Query()] = None,
    driver_id: Annotated[str | None, Query()] = None,
    vehicle_id: Annotated[str | None, Query()] = None,
    date_from: Annotated[str | None, Query()] = None,
    date_to: Annotated[str | None, Query()] = None,
):
    # id приходят строками: пустое «— все —» не должно ронять запрос в 422.
    d_id = int(driver_id) if (driver_id or "").strip().isdigit() else None
    v_id = int(vehicle_id) if (vehicle_id or "").strip().isdigit() else None
    conditions = [Expense.owner_id == owner.id]
    if category and category in _EXPENSE_CATEGORIES:
        conditions.append(Expense.category == category)
    if status and status in ("pending", "approved", "rejected"):
        conditions.append(Expense.status == status)
    if d_id:
        conditions.append(Expense.driver_id == d_id)
    if v_id:
        # машина у расхода определяется через смену; расходы без смены
        # при фильтре по машине не показываем
        conditions.append(Shift.vehicle_id == v_id)
    if date_from:
        try:
            conditions.append(Expense.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            conditions.append(
                Expense.created_at < datetime.fromisoformat(date_to) + timedelta(days=1)
            )
        except ValueError:
            pass
    rows_res = await session.execute(
        select(Expense, Driver.full_name, Vehicle.license_plate)
        .join(Driver, Driver.id == Expense.driver_id)
        .outerjoin(Shift, Shift.id == Expense.shift_id)
        .outerjoin(Vehicle, Vehicle.id == Shift.vehicle_id)
        .where(and_(*conditions))
        .order_by(desc(Expense.created_at))
        .limit(300)
    )
    rows = list(rows_res.all())
    totals = {
        "count": len(rows),
        "sum": sum((e.amount_rub or Decimal(0) for e, _, _ in rows), Decimal(0)),
        "pending": sum(1 for e, _, _ in rows if e.status == "pending"),
        "approved": sum(1 for e, _, _ in rows if e.status == "approved"),
    }
    # Разбивка по категориям для доната (из уже загруженных строк).
    _cat_ru = {"fuel": "Топливо", "repair": "Ремонт", "parking": "Парковка",
               "fine": "Штрафы", "toll": "Платные дороги", "other": "Прочее"}
    cat_sums: dict[str, Decimal] = {}
    for e, _, _ in rows:
        cat_sums[e.category] = cat_sums.get(e.category, Decimal(0)) + (e.amount_rub or Decimal(0))
    breakdown = [
        {"label": _cat_ru.get(c, c), "amount": float(v)}
        for c, v in sorted(cat_sums.items(), key=lambda kv: kv[1], reverse=True)
    ]
    # списки для фильтров «Водитель» и «Машина»
    drivers = list((await session.execute(
        select(Driver).where(Driver.owner_id == owner.id).order_by(Driver.full_name)
    )).scalars().all())
    vehicles = list((await session.execute(
        select(Vehicle)
        .where(Vehicle.owner_id == owner.id, Vehicle.is_active.is_(True))
        .order_by(Vehicle.license_plate)
    )).scalars().all())
    return templates.TemplateResponse(
        "expenses.html",
        {
            "request": request, "owner": owner, "rows": rows,
            "filter_category": category or "",
            "filter_status": status or "",
            "filter_driver_id": d_id,
            "filter_vehicle_id": v_id,
            "filter_date_from": date_from or "",
            "filter_date_to": date_to or "",
            "drivers": drivers,
            "vehicles": vehicles,
            "categories": _EXPENSE_CATEGORIES,
            "active_page": "trips", "totals": totals,
            "breakdown": breakdown,
        },
    )


@app.get("/expenses/{expense_id}", response_class=HTMLResponse)
async def expense_edit_page(
    request: Request,
    expense_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Страница правки расхода владельцем (Правка 5)."""
    expense = await session.get(Expense, expense_id)
    if expense is None or expense.owner_id != owner.id:
        raise HTTPException(status_code=404)
    driver = await session.get(Driver, expense.driver_id)
    return templates.TemplateResponse(
        "expense_edit.html",
        {
            "request": request, "owner": owner, "expense": expense, "driver": driver,
            "categories": _EXPENSE_CATEGORIES, "active_page": "trips",
        },
    )


@app.post("/expenses/{expense_id}/delete")
async def expense_delete(
    expense_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Полностью удалить расход (штраф/топливо/прочее) с сайта."""
    expense = await session.get(Expense, expense_id)
    if expense is None or expense.owner_id != owner.id:
        raise HTTPException(status_code=404)
    await session.delete(expense)
    await session.commit()
    return RedirectResponse("/expenses", status_code=303)


@app.post("/expenses/{expense_id}")
async def expense_edit_save(
    expense_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
    amount_rub: Annotated[str, Form()],
    category: Annotated[str, Form()],
    status: Annotated[str, Form()],
    file: Annotated[UploadFile | None, File()] = None,
):
    """Сохранить правку расхода: сумма, категория, статус, (опц.) фото чека."""
    expense = await session.get(Expense, expense_id)
    if expense is None or expense.owner_id != owner.id:
        raise HTTPException(status_code=404)
    if category not in _EXPENSE_CATEGORIES:
        raise HTTPException(status_code=400, detail="Bad category")
    if status not in ("pending", "approved", "rejected"):
        raise HTTPException(status_code=400, detail="Bad status")
    try:
        amt = Decimal(amount_rub.replace(",", ".").replace(" ", ""))
        if amt < 0:
            raise InvalidOperation
    except InvalidOperation:
        raise HTTPException(status_code=400, detail="Bad amount")

    expense.amount_rub = amt.quantize(Decimal("0.01"))
    expense.category = category
    if expense.status != status:
        expense.status = status
        expense.decided_at = (
            datetime.now(timezone.utc) if status in ("approved", "rejected") else None
        )
    if file is not None and file.filename:
        data = await file.read()
        if data:
            if len(data) > _MAX_DOC_BYTES:
                raise HTTPException(status_code=400, detail="Файл слишком большой (макс 6 МБ)")
            expense.receipt_web_data = data
            expense.receipt_web_type = file.content_type or "image/jpeg"
    await session.commit()
    return RedirectResponse("/expenses", status_code=303)


@app.post("/expenses/{expense_id}/receipt/delete")
async def delete_expense_receipt(
    expense_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Удалить чек у расхода (и фото от водителя, и загруженный владельцем).
    Сам расход остаётся — удаляется только прикреплённый чек."""
    expense = await session.get(Expense, expense_id)
    if expense is None or expense.owner_id != owner.id:
        raise HTTPException(status_code=404)
    expense.receipt_photo_url = None
    expense.receipt_web_data = None
    expense.receipt_web_type = None
    await session.commit()
    return RedirectResponse(f"/expenses/{expense_id}", status_code=303)


@app.get("/api/expense-receipt/{expense_id}")
async def expense_receipt(
    expense_id: int,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    expense = await session.get(Expense, expense_id)
    if expense is None or expense.owner_id != owner.id or not expense.receipt_web_data:
        raise HTTPException(status_code=404)
    return Response(content=expense.receipt_web_data, media_type=expense.receipt_web_type or "image/jpeg")


# =========================================================================
# /fuel-history — все заправки с фото чеков
# =========================================================================
@app.get("/fuel-history", response_class=HTMLResponse)
async def fuel_history(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    rows_res = await session.execute(
        select(Expense, Driver.full_name, Vehicle.license_plate)
        .join(Driver, Driver.id == Expense.driver_id)
        .outerjoin(Shift, Shift.id == Expense.shift_id)
        .outerjoin(Vehicle, Vehicle.id == Shift.vehicle_id)
        .where(
            Expense.owner_id == owner.id,
            Expense.category == "fuel",
        )
        .order_by(desc(Expense.created_at))
        .limit(200)
    )
    rows = list(rows_res.all())
    return templates.TemplateResponse(
        "fuel_history.html",
        {"request": request, "owner": owner, "rows": rows, "active_page": "trips"},
    )


# =========================================================================
# /documents — фото ТТН по рейсам
# =========================================================================
@app.get("/documents", response_class=HTMLResponse)
async def documents_page(
    request: Request,
    owner: Annotated[Owner, Depends(current_owner)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    rows_res = await session.execute(
        select(Trip, Driver.full_name, Vehicle.license_plate)
        .join(Driver, Driver.id == Trip.driver_id)
        .join(Vehicle, Vehicle.id == Trip.vehicle_id)
        .where(
            Trip.owner_id == owner.id,
            Trip.waybill_photo_url.is_not(None),
        )
        .order_by(desc(Trip.created_at))
        .limit(200)
    )
    rows = list(rows_res.all())
    # Времена загрузки ТТН по каждому trip — одним SQL
    trip_ids = [t.id for t, _, _ in rows]
    waybill_times: dict[int, datetime] = {}
    if trip_ids:
        wb_res = await session.execute(
            select(Event.trip_id, func.max(Event.created_at)).where(
                Event.owner_id == owner.id,
                Event.event_type == "waybill_uploaded",
                Event.trip_id.in_(trip_ids),
            ).group_by(Event.trip_id)
        )
        waybill_times = {tid: dt for tid, dt in wb_res.all()}
    return templates.TemplateResponse(
        "documents.html",
        {
            "request": request, "owner": owner, "rows": rows,
            "waybill_times": waybill_times,
            "active_page": "trips",
        },
    )


# =========================================================================
# Health
# =========================================================================
@app.get("/health")
async def health():
    return {"status": "ok"}
