"""
Акт оказанных услуг по РЦ за период (близко к форме 101 РС).

Группируем завершённые рейсы по пункту назначения (РЦ) и за выбранные даты
делаем .xlsx: по каждому РЦ — отдельный лист с рейсами, итогом и суммой прописью.
Только openpyxl (уже в проекте), без внешних ключей и сервисов.
"""
import io
import re
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=13)
_MONEY = "#,##0 ₽"
_THIN = Side(style="thin", color="C9CFD9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# --- сумма прописью (рубли) ---
_ONES = [
    "", "один", "два", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять",
    "десять", "одиннадцать", "двенадцать", "тринадцать", "четырнадцать", "пятнадцать",
    "шестнадцать", "семнадцать", "восемнадцать", "девятнадцать",
]
_TENS = ["", "", "двадцать", "тридцать", "сорок", "пятьдесят", "шестьдесят", "семьдесят", "восемьдесят", "девяносто"]
_HUNDREDS = ["", "сто", "двести", "триста", "четыреста", "пятьсот", "шестьсот", "семьсот", "восемьсот", "девятьсот"]


def _ones_word(o: int, female: bool) -> str:
    if female and o == 1:
        return "одна"
    if female and o == 2:
        return "две"
    return _ONES[o]


def _triplet(n: int, female: bool = False) -> list[str]:
    out: list[str] = []
    h, rem = divmod(n, 100)
    t, o = divmod(rem, 10)
    if h:
        out.append(_HUNDREDS[h])
    if t >= 2:
        out.append(_TENS[t])
        if o:
            out.append(_ones_word(o, female))
    elif t == 1:
        out.append(_ONES[10 + o])
    elif o:
        out.append(_ones_word(o, female))
    return out


def _plural(n: int, forms: tuple[str, str, str]) -> str:
    n = abs(n) % 100
    if 10 < n < 20:
        return forms[2]
    n %= 10
    if n == 1:
        return forms[0]
    if 2 <= n <= 4:
        return forms[1]
    return forms[2]


def rubles_in_words(amount: Decimal) -> str:
    amount = Decimal(amount)
    rub = int(amount)
    kop = int(round((amount - rub) * 100))
    parts: list[str] = []
    if rub == 0:
        parts = ["ноль"]
    else:
        millions, rest = divmod(rub, 1_000_000)
        thousands, ones = divmod(rest, 1000)
        if millions:
            parts += _triplet(millions)
            parts.append(_plural(millions, ("миллион", "миллиона", "миллионов")))
        if thousands:
            parts += _triplet(thousands, female=True)
            parts.append(_plural(thousands, ("тысяча", "тысячи", "тысяч")))
        if ones:
            parts += _triplet(ones)
    rub_word = _plural(rub, ("рубль", "рубля", "рублей"))
    s = " ".join(w for w in parts if w)
    s = s[:1].upper() + s[1:]
    return f"{s} {rub_word} {kop:02d} коп."


def _safe_sheet_title(name: str, used: set[str]) -> str:
    title = re.sub(r"[\\/\?\*\[\]:]", " ", name or "РЦ").strip()[:28] or "РЦ"
    base, i = title, 2
    while title in used:
        title = f"{base[:26]} {i}"
        i += 1
    used.add(title)
    return title


def build_acts_workbook(
    groups: dict[str, list[dict]], df: date, dt: date, company: str
) -> Workbook:
    """groups: РЦ → список рейсов {date, origin, destination, plate, driver, revenue}."""
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    period = f"{df.strftime('%d.%m.%Y')} — {dt.strftime('%d.%m.%Y')}"

    if not groups:
        ws = wb.create_sheet(_safe_sheet_title("Нет данных", used))
        ws["A1"] = f"За период {period} завершённых рейсов с выручкой нет."
        return wb

    for rc, trips in groups.items():
        ws = wb.create_sheet(_safe_sheet_title(rc, used))
        ws["A1"] = company or "Перевозчик"
        ws["A1"].font = _TITLE_FONT
        ws["A2"] = "АКТ оказанных транспортных услуг"
        ws["A2"].font = Font(bold=True)
        ws["A3"] = f"Заказчик (РЦ): {rc}"
        ws["A4"] = f"Период: {period}"

        header_row = 6
        headers = ["№", "Дата", "Маршрут", "Машина", "Водитель", "Сумма"]
        for col, title in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col, value=title)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _BORDER

        total = Decimal(0)
        r = header_row + 1
        for i, t in enumerate(trips, start=1):
            revenue = Decimal(t["revenue"] or 0)
            total += revenue
            values = [
                i,
                t["date"].strftime("%d.%m.%Y") if t["date"] else "—",
                f"{t['origin'] or '—'} → {t['destination'] or '—'}",
                t["plate"] or "—",
                t["driver"] or "—",
                float(revenue),
            ]
            for col, v in enumerate(values, start=1):
                c = ws.cell(row=r, column=col, value=v)
                c.border = _BORDER
                if col == 6:
                    c.number_format = _MONEY
                    c.alignment = Alignment(horizontal="right")
            r += 1

        # ИТОГО
        ws.cell(row=r, column=5, value="ИТОГО:").font = Font(bold=True)
        total_cell = ws.cell(row=r, column=6, value=float(total))
        total_cell.font = Font(bold=True)
        total_cell.number_format = _MONEY
        total_cell.alignment = Alignment(horizontal="right")
        ws.cell(row=r + 2, column=1, value=f"Всего оказано услуг на сумму: {rubles_in_words(total)}")

        widths = [5, 13, 38, 14, 22, 15]
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w

    return wb


def workbook_bytes(wb: Workbook) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
